# Importa o módulo tkinter e o renomeia como tk, facilitando
#       seu uso posterior no código.
import tkinter as tk

# Importa componentes específicos do tkinter: ttk para estilizar widgets e
#       messagebox para exibir mensagens de erro ou informações.
from tkinter import ttk, messagebox, Toplevel, Button
from tkinter.ttk import Label

from PIL.ImageOps import expand
from openpyxl.styles.builtins import total
from pandas.core.dtypes.cast import infer_dtype_from
# Importa o MongoClient do módulo pymongo, que é usado para trabalhar com MongoDB.
from pymongo import MongoClient

# Importa ObjectId do módulo bson, que é útil para trabalhar com IDs de MongoDB.
from bson import ObjectId

# Importa o módulo datetime, que permite manipular datas e horários.
import datetime

# Importa a classe FPDF do módulo fpdf, que é usada para criar arquivos PDF.
from fpdf import FPDF

# Importa o módulo os, que fornece uma maneira portátil de usar funcionalidades
#       dependentes do sistema operacional.
import os

# Importa Image e ImageTk do módulo PIL, que são usadas para trabalhar com
#       imagens em aplicações Tkinter.
from PIL import Image, ImageTk
from pymongo.synchronous.network import command

# Tenta importar Calendar e DateEntry do módulo tkcalendar.
try:
    from tkcalendar import Calendar, DateEntry

# Se houver um erro de importação (ImportError), o bloco except será executado.
except ImportError:

    # Mostra uma caixa de mensagem de erro informando que a biblioteca
    #       tkcalendar não está instalada.
    messagebox.showerror("Erro",
                         "A biblioteca 'tkcalendar' não está instalada.\nUse: pip install tkcalendar")

    # Encerra o programa lançando SystemExit, impedindo que o restante do
    #       código seja executado sem a biblioteca necessária.
    raise SystemExit


# ---------------------------------------------------------------------
# Conexão com MongoDB
# ---------------------------------------------------------------------
# Define uma função chamada 'conectar_mongodb'. Funções são
#       usadas para agrupar código que realiza uma tarefa específica.
def conectar_mongodb():

    # O bloco 'try' tentará executar o código dentro dele, que é
    #       usado para tentativas que podem falhar.
    try:

        # Cria uma conexão com o servidor MongoDB localizado em 'localhost' na porta '27017'.
        client = MongoClient("mongodb://localhost:27017/")

        # Acessa o banco de dados chamado 'clinica_veterinaria' dentro do servidor MongoDB.
        db = client["clinica_veterinaria"]

        # Retorna o objeto de banco de dados que permite fazer operações no banco.
        return db

    # O bloco 'except' captura qualquer exceção (erro) que aconteça no bloco 'try'.
    except Exception as e:

        # Mostra uma mensagem de erro se a conexão falhar. 'e' é a
        #       descrição do erro capturado.
        messagebox.showerror("Erro", f"Erro ao conectar no MongoDB:\n{e}")

        # Retorna None para indicar que houve um erro na conexão.
        return None


# Chama a função 'conectar_mongodb' e armazena o resultado na variável 'db'.
db = conectar_mongodb()

# Verifica se o valor de 'db' é None, o que indica que a conexão falhou.
if db is None:

    # Lança uma exceção, interrompendo o programa e mostrando uma mensagem de erro.
    raise Exception("Não foi possível conectar ao MongoDB.")


def centralizar_janela(janela, largura=None, altura=None):

    """
    Centraliza a janela 'janela'. Se largura/altura não forem informados,
            usa as dimensões atuais da janela.
    """

    # Atualiza as tarefas pendentes da janela para garantir
    #       que temos as dimensões corretas.
    janela.update_idletasks()

    # Se 'largura' não for especificado, obtém a largura atual da janela.
    if not largura:
        largura = janela.winfo_width()

    # Se 'altura' não for especificado, obtém a altura atual da janela.
    if not altura:
        altura = janela.winfo_height()

    # Obtém a largura da tela do dispositivo onde a janela está sendo exibida.
    largura_tela = janela.winfo_screenwidth()

    # Obtém a altura da tela do dispositivo onde a janela está sendo exibida.
    altura_tela = janela.winfo_screenheight()

    # Calcula a posição 'x' para centralizar a janela. Isso é feito
    #       subtraindo metade da largura da janela de metade da largura da tela.
    x = (largura_tela // 2) - (largura // 2)

    # Calcula a posição 'y' para centralizar a janela. Isso é feito
    #       subtraindo metade da altura da janela de metade da altura da tela.
    y = (altura_tela // 2) - (altura // 2)

    # Define a geometria da janela para que ela apareça centralizada na tela,
    #       usando as coordenadas 'x' e 'y' calculadas.
    janela.geometry(f"{largura}x{altura}+{x}+{y}")


# ---------------------------------------------------------------------
# Janela de Cadastro de Novo Médico (centralizada)
# ---------------------------------------------------------------------

# A classe JanelaCadastroMedico herda de tk.Toplevel.
# Isso significa que ela utilizará todas as funcionalidades
#       de uma janela de nível superior fornecidas pelo Tkinter,
#       que é uma biblioteca de interface gráfica.
class JanelaCadastroMedico(tk.Toplevel):

    # Este é o construtor da classe. Quando você cria uma instância
    #       desta classe, esse método é automaticamente chamado.
    # 'self' refere-se à instância atual da classe.
    # 'parent' refere-se ao widget pai ao qual esta janela (Toplevel)
    #       está associada, que geralmente é a janela principal da aplicação.

    # *args é uma convenção em Python usada para passar uma lista de
    #       argumentos variáveis não nomeados para uma função.
    # **kwargs permite passar argumentos com palavras-chave como um
    #       dicionário para uma função.
    def __init__(self, parent, *args, **kwargs):

        # Inicializa a classe pai, tk.Toplevel, que é uma janela de nível superior.
        super().__init__(parent, *args, **kwargs)

        # Configura o título da janela, que aparecerá na barra de título da janela.
        self.title("Cadastrar Novo Usuário (Médico)")

        # Este método resizable() controla se o usuário pode ou não
        #       redimensionar a janela.
        # False, False significa que a janela não pode ser redimensionada
        #       nem horizontalmente nem verticalmente.
        self.resizable(False, False)

        # Define as dimensões da janela para 500x400 pixels e
        #       centraliza a janela na tela.
        self.geometry("500x400")

        # Chamada da função 'centralizar_janela' para ajustar a
        #       posição da janela ao centro da tela.
        centralizar_janela(self, 500, 400)

        # Define a cor de fundo da janela como um cinza claro (#F0F0F0).
        self.configure(bg="#F0F0F0")

        # Cria um widget de rótulo (Label) que serve como título dentro da janela.
        titulo = tk.Label(

            self,  # 'self' indica que o rótulo pertence à instância atual da janela.
            text="Cadastro de Médico",  # Texto que será exibido no rótulo.
            font=("Arial", 18, "bold"),  # Define a fonte do texto, tamanho 18 e em negrito.
            bg="#F0F0F0",  # A cor de fundo do rótulo é cinza claro, igual ao da janela, para um design uniforme.
            fg="#333333"  # Cor do texto do rótulo, um cinza escuro, para contraste com o fundo claro.

        )

        # Posiciona o rótulo no topo da janela com um espaço vertical (pady)
        #       de 15 pixels para separá-lo do topo da janela.
        titulo.pack(pady=15)

        # Container principal
        # Cria um frame que será o container principal na janela.
        # 'bg="#FFFFFF"' define a cor de fundo do frame como branco.
        # 'bd=2' especifica uma borda de 2 pixels ao redor do frame.
        # 'relief="groove"' cria um efeito de sulco na borda para um visual 3D.
        frame = tk.Frame(self,
                         bg="#FFFFFF",
                         bd=2,
                         relief="groove")

        # Posiciona o frame no grid.
        # 'fill="both"' permite que o frame preencha tanto horizontal
        #       quanto verticalmente dentro do espaço disponível.
        # 'expand=True' permite que o frame cresça caso haja espaço adicional disponível.
        # 'padx=20' adiciona 20 pixels de espaçamento horizontal externo ao frame.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical externo ao frame.
        frame.pack(fill="both",
                   expand=True,
                   padx=20,
                   pady=10)

        # Nome do Médico
        # Cria um label para o nome do médico.
        # 'frame' especifica que o label será um filho do container principal.
        # 'text="Nome do Médico:"' define o texto exibido no label.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do label como branco.
        # 'anchor="w"' alinha o texto do label à esquerda.
        label_nome = tk.Label(frame,
                              text="Nome do Médico:",
                              font=("Arial", 12),
                              bg="#FFFFFF",
                              anchor="w")

        # Posiciona o label no grid.
        # 'row=0' posiciona o label na primeira linha do grid.
        # 'column=0' posiciona o label na primeira coluna do grid.
        # 'sticky="w"' alinha o label à esquerda dentro da célula do grid.
        # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor do label.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor do label.
        label_nome.grid(row=0,
                        column=0,
                        sticky="w",
                        padx=10,
                        pady=10)

        # Cria uma entrada de texto para o nome do médico.
        # 'frame' especifica que o entry será um filho do container principal.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'width=30' define que o campo de texto terá 30 caracteres de largura.
        self.entry_nome = tk.Entry(frame,
                                   font=("Arial", 12),
                                   width=30)

        # Posiciona a entrada de texto no grid.
        # 'row=0' posiciona a entrada de texto na primeira linha do grid.
        # 'column=1' posiciona a entrada de texto na segunda coluna do grid.
        # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor da entrada de texto.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor da entrada de texto.
        self.entry_nome.grid(row=0,
                             column=1,
                             padx=10,
                             pady=10)

        # Especialidade
        # Cria um label para o campo de especialidade.
        # 'frame' especifica que o label será um filho do container principal.
        # 'text="Especialidade:"' define o texto exibido no label.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do label como branco.
        # 'anchor="w"' alinha o texto do label à esquerda.
        tk.Label(
            frame,
            text="Especialidade:",
            font=("Arial", 12),
            bg="#FFFFFF",
            anchor="w"
            ).grid(

                # Posiciona o label no grid.
                # 'row=1' coloca o label na segunda linha do grid.
                # 'column=0' coloca o label na primeira coluna do grid.
                # 'sticky="w"' alinha o label à esquerda dentro da célula do grid.
                # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor do label.
                # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor do label.
                row=1,
                column=0,
                sticky="w",
                padx=10,
                pady=10)

        # Cria uma entrada de texto para a especialidade.
        # 'frame' especifica que o entry será um filho do container principal.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'width=30' define que o campo de texto terá 30 caracteres de largura.
        self.entry_especialidade = tk.Entry(frame,
                                            font=("Arial", 12),
                                            width=30)

        # Posiciona a entrada de texto no grid.
        # 'row=1' coloca a entrada de texto na segunda linha do grid.
        # 'column=1' coloca a entrada de texto na segunda coluna do grid.
        # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor da entrada de texto.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor da entrada de texto.
        self.entry_especialidade.grid(row=1,
                                      column=1,
                                      padx=10,
                                      pady=10)

        # CRMV
        # Cria um label para o campo de CRMV (Conselho Regional de Medicina Veterinária).
        # 'frame' especifica que o label será um filho do container principal.
        # 'text="CRMV:"' define o texto exibido no label.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do label como branco.
        # 'anchor="w"' alinha o texto do label à esquerda.
        tk.Label(frame,
                text="CRMV:",
                font=("Arial", 12),
                bg="#FFFFFF",
                anchor="w"
        ).grid(
                # Posiciona o label no grid.
                # 'row=2' coloca o label na terceira linha do grid.
                # 'column=0' coloca o label na primeira coluna do grid.
                # 'sticky="w"' alinha o label à esquerda dentro da célula do grid.
                # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor do label.
                # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor do label.
                row=2, column=0, sticky="w", padx=10, pady=10
        )

        # Cria uma entrada de texto para o CRMV.
        # 'frame' especifica que o entry será um filho do container principal.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'width=30' define que o campo de texto terá 30 caracteres de largura.
        self.entry_crmv = tk.Entry(frame, font=("Arial", 12), width=30)

        # Posiciona a entrada de texto no grid.
        # 'row=2' coloca a entrada de texto na terceira linha do grid.
        # 'column=1' coloca a entrada de texto na segunda coluna do grid.
        # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor da entrada de texto.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor da entrada de texto.
        self.entry_crmv.grid(row=2,
                             column=1,
                             padx=10,
                             pady=10)

        # Senha
        # Cria um label para o campo de Senha.
        # 'frame' especifica que o label será um filho do container principal.
        # 'text="Senha:"' define o texto exibido no label.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do label como branco.
        # 'anchor="w"' alinha o texto do label à esquerda.
        tk.Label(frame,
                text="Senha:",
                font=("Arial", 12),
                bg="#FFFFFF",
                anchor="w"
        ).grid(

            # Posiciona o label no grid.
            # 'row=3' coloca o label na quarta linha do grid.
            # 'column=0' coloca o label na primeira coluna do grid.
            # 'sticky="w"' alinha o label à esquerda dentro da célula do grid.
            # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor do label.
            # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor do label.
            row=3, column=0, sticky="w", padx=10, pady=10
        )

        # Cria uma entrada de texto para o campo de Senha.
        # 'frame' especifica que o entry será um filho do container principal.
        # 'font=("Arial", 12)' configura a fonte como Arial, tamanho 12.
        # 'width=30' define que o campo de texto terá 30 caracteres de largura.
        # 'show="*"' exibe um asterisco (*) no lugar de cada caractere digitado, ocultando a senha.
        self.entry_senha = tk.Entry(frame,
                                    font=("Arial", 12),
                                    width=30,
                                    show="*")

        # Posiciona a entrada de texto no grid.
        # 'row=3' coloca a entrada de texto na quarta linha do grid.
        # 'column=1' coloca a entrada de texto na segunda coluna do grid.
        # 'padx=10' adiciona 10 pixels de espaçamento horizontal ao redor da entrada de texto.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical ao redor da entrada de texto.
        self.entry_senha.grid(row=3,
                              column=1,
                              padx=10,
                              pady=10)

        # Botão de Cadastro
        # Cria um botão para a funcionalidade de cadastro.
        # 'self' especifica que o botão será um filho da janela atual.
        # 'text="Cadastrar"' define o texto exibido no botão como "Cadastrar".
        # 'font=("Arial", 14, "bold")' configura a fonte como Arial, tamanho 14, em negrito.
        # 'bg="#4CAF50"' define a cor de fundo do botão como verde.
        # 'fg="#FFFFFF"' define a cor do texto do botão como branco.
        # 'activebackground="#45A049"' define a cor de fundo do botão ao
        #       ser pressionado como um verde mais escuro.
        # 'command=self.cadastrar' especifica que a função 'self.cadastrar'
        #       será chamada quando o botão for clicado.
        self.btn_cadastrar = tk.Button(self,
                                        text="Cadastrar",
                                        font=("Arial", 14, "bold"),
                                        bg="#4CAF50",
                                        fg="#FFFFFF",
                                        activebackground="#45A049",
                                        command=self.cadastrar)

        # Posiciona o botão na janela usando o gerenciador de layout 'pack'.
        # 'pady=15' adiciona 15 pixels de espaçamento vertical acima e abaixo do botão.
        self.btn_cadastrar.pack(pady=15)


    def cadastrar(self):

        # Obtém o texto inserido no campo 'Nome do Médico', removendo
        #       espaços em branco no início e no final.
        nome = self.entry_nome.get().strip()

        # Obtém o texto inserido no campo 'Especialidade', removendo
        #       espaços em branco no início e no final.
        espec = self.entry_especialidade.get().strip()

        # Obtém o texto inserido no campo 'CRMV', removendo espaços em
        #       branco no início e no final.
        crmv = self.entry_crmv.get().strip()

        # Obtém o texto inserido no campo 'Senha', removendo espaços em
        #       branco no início e no final.
        senha = self.entry_senha.get().strip()

        # Verifica se os campos 'Nome' ou 'Senha' estão vazios.
        # Caso estejam, exibe uma mensagem de aviso para o usuário e
        #       encerra a função.
        if not nome or not senha:
            messagebox.showwarning("Aviso",
                                   "Nome do médico e Senha são obrigatórios.")
            return

        # Verifica no banco de dados MongoDB se já existe um médico com o mesmo nome.
        existe = db.medicos.find_one({"nome": nome})

        # Se já existir um médico com o mesmo nome, exibe uma mensagem de erro
        #       e encerra a função para evitar duplicação.
        if existe:
            messagebox.showerror("Erro",
                                 "Já existe um médico com este nome.")
            return

        # Cria um dicionário com os dados do médico a serem
        #       cadastrados no banco de dados.
        doc = {

            "nome": nome,  # Adiciona o nome do médico ao dicionário.
            "especialidade": espec,  # Adiciona a especialidade do médico ao dicionário.
            "crmv": crmv,  # Adiciona o número CRMV do médico ao dicionário.
            "senha": senha,  # Adiciona a senha do médico ao dicionário.
            "cadastrado_em": datetime.datetime.now()  # Adiciona a data e hora atuais como o momento do cadastro.

        }

        # Tenta inserir o documento no banco de dados.
        try:

            # Insere o dicionário 'doc' na coleção 'medicos' do banco de dados.
            db.medicos.insert_one(doc)

            # Exibe uma mensagem de sucesso informando que o médico foi cadastrado.
            messagebox.showinfo("Sucesso",
                                f"Médico '{nome}' cadastrado com sucesso!")

            # Fecha a janela de cadastro após o sucesso da operação.
            self.destroy()

        # Trata possíveis exceções que possam ocorrer durante a tentativa de inserção.
        except Exception as e:

            # Exibe uma mensagem de erro informando que o cadastro não foi possível
            # e exibe o motivo detalhado (conteúdo da exceção 'e').
            messagebox.showerror("Erro",
                                 f"Não foi possível cadastrar:\n{e}")




# ---------------------------------------------------------------------
# JanelaPrincipal
# ---------------------------------------------------------------------
# Define a classe `JanelaPrincipal`, que representa a janela principal do sistema.
class JanelaPrincipal(tk.Tk):

    # Método inicializador da classe, chamado ao criar uma nova instância.
    # O parâmetro `medico_logado` contém os dados do médico autenticado no login.
    def __init__(self, medico_logado):

        # Chama o inicializador da classe pai (`tk.Tk`) para configurar a
        #       janela como um objeto do Tkinter.
        super().__init__()

        # Define o título da janela principal como "Sistema de Clínica Veterinária".
        self.title("Sistema de Clínica Veterinária")

        # Ajusta a janela para abrir em tela cheia.
        # O método `state('zoomed')` faz isso em sistemas suportados.
        self.state('zoomed')

        # Cria um widget `Notebook`, que permite a criação de múltiplas abas na janela.
        # Um `Notebook` é um contêiner de abas onde cada aba pode
        #       conter widgets distintos.
        self.notebook = ttk.Notebook(self)

        # Posiciona o `Notebook` na janela principal.
        # `fill="both"` faz com que ele ocupe todo o espaço disponível,
        # e `expand=True` permite que ele se ajuste ao redimensionamento da janela.
        self.notebook.pack(fill="both", expand=True)

        # Cria a aba "Principal", que representa a página inicial do sistema.
        # `TelaPrincipal` é uma classe que define o conteúdo desta aba.
        # O primeiro parâmetro é o pai (`self.notebook`), e o segundo
        #       são os dados do médico logado.
        aba_principal = TelaPrincipal(self.notebook, medico_logado)

        # Adiciona a aba "Principal" ao `Notebook` com o título "Principal".
        self.notebook.add(aba_principal, text="Principal")

        # Cria a aba "Clientes", que gerencia informações dos clientes da clínica.
        # `TelaClientes` define o conteúdo da aba, passando o `Notebook`
        #       como pai e os dados do médico logado.
        aba_clientes = TelaClientes(self.notebook, medico_logado)

        # Adiciona a aba "Clientes" ao `Notebook` com o título "Clientes".
        self.notebook.add(aba_clientes, text="Clientes")

        # Cria a aba "Animais", que gerencia os dados dos animais registrados na clínica.
        # `TelaAnimais` define o conteúdo da aba, recebendo o `Notebook` e
        #       os dados do médico logado.
        aba_animais = TelaAnimais(self.notebook, medico_logado)

        # Adiciona a aba "Animais" ao `Notebook` com o título "Animais".
        self.notebook.add(aba_animais, text="Animais")

        # Cria a aba "Médicos", que gerencia as informações dos médicos da clínica.
        # `TelaMedicos` é a classe responsável por configurar essa aba.
        aba_medicos = TelaMedicos(self.notebook, medico_logado)

        # Adiciona a aba "Médicos" ao `Notebook` com o título "Médicos".
        self.notebook.add(aba_medicos, text="Médicos")

        # Cria a aba "Agenda", que gerencia os compromissos e horários da clínica.
        # `TelaAgenda` é a classe que define o conteúdo dessa aba.
        aba_agenda = TelaAgenda(self.notebook, medico_logado)

        # Adiciona a aba "Agenda" ao `Notebook` com o título "Agenda".
        self.notebook.add(aba_agenda, text="Agenda")

        # Cria a aba "Estoque", que gerencia os itens disponíveis na clínica.
        # A referência desta aba é armazenada em `self.aba_estoque` para
        #       uso posterior, se necessário.
        # `TelaEstoque` é a classe responsável pelo conteúdo dessa aba.
        self.aba_estoque = TelaEstoque(self.notebook, medico_logado)

        # Adiciona a aba "Estoque" ao `Notebook` com o título "Estoque".
        self.notebook.add(self.aba_estoque, text="Estoque")

        # Cria a aba "Relatórios", que exibe relatórios analíticos sobre a clínica.
        # `TelaRelatorios` é a classe que define o conteúdo dessa aba.
        aba_relatorios = TelaRelatorios(self.notebook, medico_logado)

        # Adiciona a aba "Relatórios" ao `Notebook` com o título "Relatórios".
        self.notebook.add(aba_relatorios, text="Relatórios")


# ---------------------------------------------------------------------
# Janela de Atendimento
# ---------------------------------------------------------------------

# Define a classe JanelaAtendimento, que herda de tk.Toplevel,
#       para criar uma janela separada.
class JanelaAtendimento(tk.Toplevel):

    # Método de inicialização da janela de atendimento.
    # 'parent': janela pai.
    # 'id_agendamento_str': identificador do agendamento em formato de string.
    # 'nome_animal': nome do animal a ser atendido.
    # 'medico_logado': dados do médico que está realizando o atendimento.
    # 'referencia_estoque': referência à tela de estoque, caso necessário.
    # '*args': argumentos posicionais adicionais para o construtor da classe base.
    # '**kwargs': argumentos nomeados adicionais para o construtor da classe base.
    def __init__(self,
                 parent,
                 id_agendamento_str,
                 nome_animal,
                 medico_logado,
                 referencia_estoque=None,
                 *args,
                 **kwargs):

        # Inicializa a classe base (tk.Toplevel), que representa uma nova janela,
        # passando os argumentos e parâmetros adicionais recebidos.
        super().__init__(parent, *args, **kwargs)

        # Define o título da janela como "Atendimento do Animal".
        self.title("Atendimento do Animal")

        # Armazena a referência para a tela de estoque, caso seja usada.
        self.referencia_estoque = referencia_estoque

        # Define o tamanho da janela para 800x700 pixels.
        self.geometry("800x700")

        # Centraliza a janela na tela com a função 'centralizar_janela'.
        centralizar_janela(self, 800, 700)

        # Torna esta janela modal em relação à janela pai, impedindo interação
        #       com a janela principal enquanto essa estiver aberta.
        self.transient(parent)

        # Armazena os dados do médico logado para uso no atendimento.
        self.medico_logado = medico_logado

        # Armazena o identificador do agendamento para rastrear o atendimento atual.
        self.id_agendamento_str = id_agendamento_str

        # Armazena o nome do animal sendo atendido.
        self.nome_animal = nome_animal

        # Carrega o documento do agendamento correspondente ao ID fornecido.
        # Usa o MongoDB para buscar um documento na coleção 'agenda'
        #       cujo '_id' corresponda ao 'id_agendamento_str'.
        self.agendamento_doc = db.agenda.find_one({"_id": ObjectId(id_agendamento_str)})

        # Verifica se o agendamento foi encontrado no banco de dados.
        if not self.agendamento_doc:

            # Exibe uma mensagem de erro se o agendamento não foi encontrado.
            messagebox.showerror("Erro", "Agendamento não encontrado.")

            # Fecha a janela atual, pois não há dados para exibir.
            self.destroy()

            # Encerra o método, pois não faz sentido continuar sem o agendamento.
            return

        # Carrega os dados do animal associado ao agendamento.
        # Busca no banco de dados MongoDB na coleção 'animais' usando o
        #       ID do animal armazenado no agendamento.
        self.animal_doc = db.animais.find_one({"_id": self.agendamento_doc["id_animal"]})

        # Verifica se os dados do animal foram encontrados no banco de dados.
        if not self.animal_doc:

            # Exibe uma mensagem de erro se o animal não foi encontrado.
            messagebox.showerror("Erro", "Animal não encontrado.")

            # Fecha a janela atual, pois não há dados do animal para exibir.
            self.destroy()

            # Encerra o método, pois não faz sentido continuar sem os dados do animal.
            return

        # Obtém o ID do dono do animal a partir do documento do animal carregado.
        # Utiliza o método 'get' para evitar erros caso a chave 'id_dono' não exista.
        dono_id = self.animal_doc.get("id_dono")

        # Busca no banco de dados MongoDB, na coleção 'clientes', o
        #       documento do dono correspondente ao 'dono_id'.
        # Se 'dono_id' for None, 'dono_doc' será definido como None.
        dono_doc = db.clientes.find_one({"_id": dono_id}) if dono_id else None

        # Define o nome do dono do animal com base no documento encontrado.
        # Se 'dono_doc' for None (ou seja, não encontrado), o nome será "Desconhecido".
        self.nome_dono = dono_doc["nome"] if dono_doc else "Desconhecido"

        # Inicializa o carrinho de compras ou procedimentos como uma lista vazia.
        # Essa lista será usada para armazenar os itens
        #       adicionados durante o atendimento.
        self.carrinho = []

        # Cria um widget Canvas, que é uma área de desenho na interface gráfica.
        # O Canvas será usado para exibir informações com possibilidade de rolagem.
        # 'borderwidth=0' remove a borda padrão do Canvas.
        canvas = tk.Canvas(self, borderwidth=0)

        # Cria um widget Scrollbar para permitir a rolagem vertical do Canvas.
        # 'orient="vertical"' define que a barra de rolagem será vertical.
        # 'command=canvas.yview' vincula a barra de rolagem ao movimento do Canvas.
        scrollbar = tk.Scrollbar(self,
                                 orient="vertical",
                                 command=canvas.yview)

        # Configura o Canvas para que sua barra de rolagem seja
        #       atualizada sempre que houver movimento.
        # Isso cria a associação entre o Canvas e a barra de rolagem.
        canvas.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do Canvas.
        # 'side="right"' coloca a barra de rolagem na borda direita do
        #       widget pai (a janela).
        # 'fill="y"' ajusta a altura da barra para preencher todo o
        #       espaço vertical disponível.
        scrollbar.pack(side="right", fill="y")

        # Posiciona o Canvas no lado esquerdo da janela, onde será
        #       exibido o conteúdo principal.
        # 'side="left"' posiciona o Canvas na borda esquerda do widget pai (a janela).
        # 'fill="both"' ajusta tanto a largura quanto a altura do Canvas
        #       para ocupar todo o espaço disponível.
        # 'expand=True' permite que o Canvas se expanda e ajuste
        #       automaticamente ao redimensionar a janela.
        canvas.pack(side="left", fill="both", expand=True)

        # Cria um frame interno dentro do Canvas.
        # Esse frame será usado para conter os widgets exibidos no Canvas.
        # Ele permite que o conteúdo tenha rolagem quando exceder o
        #       tamanho visível do Canvas.
        self.inner_frame = tk.Frame(canvas)

        # Adiciona o frame interno ao Canvas como uma janela.
        # 'create_window((0, 0))' posiciona o frame no canto superior esquerdo do Canvas.
        # 'window=self.inner_frame' especifica que o frame interno será
        #       exibido dentro do Canvas.
        # 'anchor="nw"' alinha o frame ao canto superior esquerdo (noroeste) do Canvas.
        canvas.create_window((0, 0),
                             window=self.inner_frame,
                             anchor="nw")

        # Associa um evento ao frame interno para atualizar a região de rolagem do Canvas.
        # '<Configure>' é acionado sempre que o tamanho ou a posição do frame interno muda.
        # O lambda executa uma função para configurar a região de rolagem do Canvas.
        # 'scrollregion=canvas.bbox("all")' ajusta a área de rolagem para abranger todo o conteúdo do Canvas.
        self.inner_frame.bind("<Configure>",
                              lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Cria um rótulo (Label) para exibir o nome do animal atendido.
        # O texto exibe "Animal:" seguido do nome do animal (`self.nome_animal`).
        # 'font=("Arial", 14, "bold")' define o estilo da fonte
        #       como Arial, tamanho 14 e negrito.
        lbl_animal = tk.Label(self.inner_frame,
                              text=f"Animal: {self.nome_animal}",
                              font=("Arial", 14, "bold"))

        # Adiciona o rótulo ao frame interno.
        # 'pack(pady=10)' posiciona o rótulo com um espaçamento
        #       vertical (padding) de 10 pixels acima e abaixo.
        lbl_animal.pack(pady=10)

        # Cria outro rótulo para exibir o nome do dono do animal.
        # O texto exibe "Dono:" seguido do nome do dono (`self.nome_dono`).
        # 'font=("Arial", 12)' define o estilo da fonte como Arial, tamanho 12.
        lbl_dono = tk.Label(self.inner_frame,
                            text=f"Dono: {self.nome_dono}",
                            font=("Arial", 12))

        # Adiciona o rótulo do dono ao frame interno.
        # 'pack(pady=5)' posiciona o rótulo com um espaçamento
        #       vertical (padding) de 5 pixels acima e abaixo.
        lbl_dono.pack(pady=5)

        # Cria um frame para organizar os widgets relacionados ao valor da consulta.
        # 'pady=5' adiciona um espaçamento vertical de 5 pixels acima e abaixo do frame.
        frame_valor = tk.Frame(self.inner_frame, pady=5)

        # Adiciona o frame ao frame interno (`self.inner_frame`).
        # 'fill="x"' faz com que o frame ocupe toda a largura disponível horizontalmente.
        # 'padx=20' adiciona um espaçamento horizontal de 20 pixels nas laterais do frame.
        frame_valor.pack(fill="x", padx=20)

        # Cria um rótulo para identificar o campo de entrada do valor da consulta.
        # O texto "Valor da Consulta (R$):" é exibido para o usuário.
        # 'font=("Arial", 12)' define a fonte do texto como Arial, tamanho 12.
        tk.Label(frame_valor,
                text="Valor da Consulta (R$):",
                font=("Arial", 12)).pack(side="left",  # Posiciona o rótulo à esquerda dentro do frame.
                                        padx=5)  # Adiciona um espaçamento horizontal de 5 pixels ao redor do rótulo.


        # Cria um campo de entrada (Entry) para que o usuário
        #       possa digitar o valor da consulta.
        # 'width=10' define a largura do campo de entrada em caracteres.
        # 'font=("Arial", 12)' define a fonte do texto digitado no
        #       campo como Arial, tamanho 12.
        self.entry_valor_consulta = tk.Entry(frame_valor,
                                            width=10,
                                            font=("Arial", 12))

        # Adiciona o campo de entrada ao frame de valor da consulta.
        # 'side="left"' posiciona o campo de entrada à esquerda do frame, ao lado do rótulo.
        # 'padx=5' adiciona um espaçamento horizontal de 5 pixels ao redor do campo de entrada.
        self.entry_valor_consulta.pack(side="left", padx=5)

        # Cria um frame rotulado para adicionar produtos ao atendimento.
        # 'text="Adicionar Produtos ao Atendimento"' define o título visível do frame.
        # 'padx=10' e 'pady=10' adicionam espaçamento interno
        #       horizontal e vertical dentro do frame.
        frame_produtos_top = tk.LabelFrame(self.inner_frame,
                                            text="Adicionar Produtos ao Atendimento",
                                            padx=10,
                                            pady=10)

        # Posiciona o frame dentro do frame interno (`self.inner_frame`).
        # 'fill="x"' faz com que o frame ocupe toda a largura disponível.
        # 'padx=20' e 'pady=10' adicionam espaçamento externo ao redor do frame.
        frame_produtos_top.pack(fill="x", padx=20, pady=10)

        # Cria um rótulo para identificar o campo onde será selecionado o produto.
        # O texto "Produto:" é exibido ao lado do campo de seleção.
        tk.Label(

            frame_produtos_top,  # O rótulo pertence ao frame rotulado `frame_produtos_top`.
            text="Produto:"  # Texto que será exibido ao lado do campo de seleção.

        ).grid(

            row=0,  # Define que o rótulo será colocado na primeira linha do grid.
            column=0,  # Define que o rótulo será colocado na primeira coluna do grid.
            sticky="e",  # Alinha o rótulo à direita (east) dentro de sua célula.
            padx=5,  # Adiciona um espaçamento horizontal de 5 pixels ao redor do rótulo.
            pady=5  # Adiciona um espaçamento vertical de 5 pixels ao redor do rótulo.

        )

        # Cria um combobox para que o usuário possa selecionar um produto da lista.
        # 'width=30' define a largura do combobox em caracteres.
        self.combo_produto = ttk.Combobox(

            frame_produtos_top,  # O combobox pertence ao frame rotulado `frame_produtos_top`.
            width=30  # Define a largura do combobox.

        )

        # Posiciona o combobox ao lado do rótulo dentro do frame.
        # 'row=0' posiciona o combobox na mesma linha do rótulo.
        # 'column=1' posiciona o combobox na segunda coluna ao lado do rótulo.
        # 'padx=5' e 'pady=5' adicionam espaçamento ao redor do combobox.
        self.combo_produto.grid(row=0, column=1, padx=5, pady=5)

        # Posiciona o rótulo "Preço Unitário (R$)" no grid.
        # 'row=0' coloca o rótulo na primeira linha.
        # 'column=2' coloca o rótulo na terceira coluna.
        # 'sticky="e"' alinha o rótulo à direita dentro da célula.
        # 'padx=5' adiciona 5 pixels de espaço horizontal ao redor do rótulo.
        # 'pady=5' adiciona 5 pixels de espaço vertical ao redor do rótulo.
        tk.Label(frame_produtos_top,
                 text="Preço Unitário (R$):").grid(row=0,
                                                   column=2,
                                                   sticky="e",
                                                   padx=5,
                                                   pady=5)

        # Cria e posiciona o rótulo para exibir o preço unitário no grid.
        # 'text="0.00"' define o valor inicial exibido no rótulo como 0.00.
        # 'width=10' define a largura do rótulo em caracteres.
        # 'anchor="e"' alinha o texto do rótulo à direita.
        self.label_preco_unit = tk.Label(frame_produtos_top,
                                         text="0.00",
                                         width=10,
                                         anchor="e")

        # 'row=0' coloca o rótulo na primeira linha.
        # 'column=3' coloca o rótulo na quarta coluna.
        # 'padx=5' adiciona 5 pixels de espaço horizontal ao redor do rótulo.
        # 'pady=5' adiciona 5 pixels de espaço vertical ao redor do rótulo.
        self.label_preco_unit.grid(row=0,
                                   column=3,
                                   padx=5,
                                   pady=5)

        # Posiciona o rótulo "Qtd:" no grid.
        # 'text="Qtd:"' define o texto exibido no rótulo como "Qtd:".
        # 'row=0' coloca o rótulo na primeira linha.
        # 'column=4' coloca o rótulo na quinta coluna.
        # 'sticky="e"' alinha o rótulo à direita dentro da célula.
        # 'padx=5' adiciona 5 pixels de espaço horizontal ao redor do rótulo.
        # 'pady=5' adiciona 5 pixels de espaço vertical ao redor do rótulo.
        tk.Label(frame_produtos_top, text="Qtd:").grid(row=0,
                                                       column=4,
                                                       sticky="e",
                                                       padx=5,
                                                       pady=5)

        # Cria e posiciona o campo de entrada para a quantidade no grid.
        # 'width=5' define a largura do campo de entrada como 5 caracteres.
        # 'font=("Arial", 12)' define a fonte e o tamanho do texto no campo de entrada.
        # 'row=0' coloca o campo de entrada na primeira linha.
        # 'column=5' coloca o campo de entrada na sexta coluna.
        # 'padx=5' adiciona 5 pixels de espaço horizontal ao redor do campo de entrada.
        # 'pady=5' adiciona 5 pixels de espaço vertical ao redor do campo de entrada.
        self.entry_qtd = tk.Entry(frame_produtos_top, width=5, font=("Arial", 12))
        self.entry_qtd.grid(row=0, column=5, padx=5, pady=5)

        # Define o valor inicial do campo de entrada para "1".
        # O método 'insert(0, "1")' insere o texto "1" na posição
        #       inicial (índice 0) do campo de entrada.
        self.entry_qtd.insert(0, "1")

        # Cria e posiciona o botão "Adicionar" no grid.
        # 'text="Adicionar"' define o texto exibido no botão como "Adicionar".
        # 'command=self.adicionar_produto_ao_carrinho' vincula o botão à
        #       função 'adicionar_produto_ao_carrinho' que será executada quando clicado.
        # 'row=0' coloca o botão na primeira linha.
        # 'column=6' coloca o botão na sétima coluna.
        # 'padx=5' adiciona 5 pixels de espaço horizontal ao redor do botão.
        # 'pady=5' adiciona 5 pixels de espaço vertical ao redor do botão.
        self.btn_add_produto = tk.Button(frame_produtos_top, text="Adicionar",
                                         command=self.adicionar_produto_ao_carrinho)
        self.btn_add_produto.grid(row=0, column=6, padx=5, pady=5)

        # Cria e posiciona o botão "Remover Produto Selecionado" no grid.
        # 'text="Remover Produto Selecionado"' define o texto exibido no
        #       botão como "Remover Produto Selecionado".
        # 'command=self.remover_produto_selecionado' vincula o botão à
        #       função 'remover_produto_selecionado' que será executada quando clicado.
        # 'bg="#F44336"' define a cor de fundo do botão como vermelho (código hexadecimal #F44336).
        # 'fg="white"' define a cor do texto do botão como branco.
        # 'row=0' coloca o botão na primeira linha.
        # 'column=7' coloca o botão na oitava coluna.
        # 'padx=10' adiciona 10 pixels de espaço horizontal ao redor do botão.
        # 'pady=5' adiciona 5 pixels de espaço vertical ao redor do botão.
        self.btn_remover_produto = tk.Button(frame_produtos_top,
                                             text="Remover Produto Selecionado",
                                             command=self.remover_produto_selecionado,
                                             bg="#F44336", fg="white")
        self.btn_remover_produto.grid(row=0, column=7, padx=10, pady=5)

        # Vincula o evento de seleção de um item na combobox à função '_ao_selecionar_produto'.
        # 'bind("<<ComboboxSelected>>")' detecta quando um item da combobox é selecionado.
        # 'self._ao_selecionar_produto' é a função que será executada ao selecionar um item.
        self.combo_produto.bind("<<ComboboxSelected>>", self._ao_selecionar_produto)

        # Cria um frame com borda e título para listar os produtos no atendimento.
        # 'text="Produtos no Atendimento"' define o título do frame
        #       como "Produtos no Atendimento".
        # 'padx=10' e 'pady=10' adicionam 10 pixels de espaço interno ao
        #       redor do conteúdo do frame.
        frame_produtos_list = tk.LabelFrame(self.inner_frame,
                                            text="Produtos no Atendimento",
                                            padx=10,
                                            pady=10)

        # Posiciona o frame dentro do layout.
        # 'fill="both"' faz com que o frame expanda tanto horizontalmente
        #       quanto verticalmente para preencher o espaço disponível.
        # 'expand=True' permite que o frame ocupe todo o espaço adicional
        #       disponível na direção definida por 'fill'.
        # 'padx=20' e 'pady=10' adicionam 20 pixels de espaço horizontal e 10 pixels de
        #       espaço vertical ao redor do frame.
        frame_produtos_list.pack(fill="both", expand=True, padx=20, pady=10)

        # Define as colunas que serão exibidas na tabela do Treeview.
        # 'colunas = ("nome_produto", "preco_unit", "qtd", "subtotal")' cria
        #       uma tupla com os nomes das colunas:
        # - "nome_produto": Nome do produto.
        # - "preco_unit": Preço unitário do produto.
        # - "qtd": Quantidade do produto.
        # - "subtotal": Subtotal (preço unitário multiplicado pela quantidade).
        colunas = ("nome_produto", "preco_unit", "qtd", "subtotal")


        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria um widget Treeview para exibir os produtos no atendimento.
        # 'frame_produtos_list' é o container onde o Treeview será colocado.
        # 'columns=colunas' especifica as colunas definidas anteriormente.
        # 'show="headings"' exibe apenas os cabeçalhos das colunas, sem a coluna de hierarquia.
        self.tree_carrinho = ttk.Treeview(frame_produtos_list, columns=colunas, show="headings")

        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado



        # Define os cabeçalhos das colunas no Treeview.
        # O loop percorre cada coluna em 'colunas' e define o texto do
        #       cabeçalho como o próprio nome da coluna.
        # 'self.tree_carrinho.heading(c, text=c)' ajusta o texto visível no
        #       cabeçalho da tabela para cada coluna.
        for c in colunas:
            self.tree_carrinho.heading(c, text=c)

        # Define a largura da coluna "nome_produto" no Treeview como 200 pixels.
        self.tree_carrinho.column("nome_produto", width=200)

        # Define a largura da coluna "preco_unit" no Treeview como 100 pixels.
        self.tree_carrinho.column("preco_unit", width=100)

        # Define a largura da coluna "qtd" no Treeview como 80 pixels.
        self.tree_carrinho.column("qtd", width=80)

        # Define a largura da coluna "subtotal" no Treeview como 100 pixels.
        self.tree_carrinho.column("subtotal", width=100)

        # Posiciona o Treeview dentro do layout.
        # 'fill="both"' faz com que o Treeview expanda tanto horizontalmente
        #       quanto verticalmente para preencher o espaço disponível.
        # 'expand=True' permite que o Treeview ocupe todo o espaço
        #       adicional disponível no frame.
        self.tree_carrinho.pack(fill="both", expand=True)

        # Cria um frame para exibir os totais relacionados aos produtos no atendimento.
        # 'frame_totais' será usado para exibir informações como o valor total do atendimento.
        frame_totais = tk.Frame(self.inner_frame)

        # Posiciona o frame de totais no layout.
        # 'fill="x"' faz com que o frame se expanda horizontalmente para
        #       preencher todo o espaço disponível.
        # 'padx=20' adiciona 20 pixels de espaço horizontal externo ao redor do frame.
        # 'pady=5' adiciona 5 pixels de espaço vertical externo ao redor do frame.
        frame_totais.pack(fill="x", padx=20, pady=5)

        # Cria um rótulo dentro do frame de totais para exibir o texto "Total Produtos (R$):".
        # Este rótulo serve para identificar o valor total dos produtos no atendimento.
        # 'font=("Arial", 12)' define a fonte do texto como Arial com tamanho 12.
        tk.Label(frame_totais,
                 text="Total Produtos (R$):",
                 font=("Arial", 12)).pack(side="left",  # Posiciona o rótulo à esquerda dentro do frame.
                                          padx=5)  # Adiciona 5 pixels de espaço horizontal ao redor do rótulo.

        # Cria um rótulo para exibir o valor total dos produtos no atendimento.
        # Este rótulo será atualizado dinamicamente conforme os
        #       produtos são adicionados ou removidos.
        # 'text="0.00"' define o texto inicial do rótulo como "0.00".
        # 'font=("Arial", 12, "bold")' define a fonte como Arial, tamanho 12, em negrito.
        # 'width=10' define a largura do rótulo como 10 caracteres.
        # 'anchor="e"' alinha o texto à direita dentro do rótulo.
        self.label_total_produtos = tk.Label(frame_totais,
                                            text="0.00",
                                            font=("Arial", 12, "bold"),
                                            width=10,
                                            anchor="e")

        # Posiciona o rótulo de total de produtos dentro do frame.
        # 'side="left"' posiciona o rótulo à esquerda do frame, ao
        #       lado do texto descritivo.
        # 'padx=5' adiciona 5 pixels de espaço horizontal ao redor do rótulo.
        self.label_total_produtos.pack(side="left", padx=5)

        # Cria um rótulo dentro do frame de totais para exibir o texto "Total Geral (R$):".
        # Este rótulo identifica o valor total geral, incluindo todos os custos do atendimento.
        # 'text="|  Total Geral (R$):"' adiciona uma barra separadora (|) para melhor visualização.
        # 'font=("Arial", 12)' define a fonte do texto como Arial com tamanho 12.
        tk.Label(frame_totais,
                 text="|  Total Geral (R$):",
                 font=("Arial", 12)).pack(

                                    # Posiciona o rótulo à esquerda dentro do frame.
                                    side="left",

                                    # Adiciona 15 pixels de espaço horizontal ao redor do rótulo
                                    #       para separação do conteúdo anterior.
                                    padx=15)

        # Cria um rótulo para exibir o valor total geral do atendimento.
        # Este rótulo será atualizado dinamicamente conforme os valores
        #       dos produtos ou consultas mudam.
        # 'text="0.00"' define o texto inicial do rótulo como "0.00".
        # 'font=("Arial", 12, "bold")' define a fonte como Arial, tamanho 12, em negrito.
        # 'width=10' define a largura do rótulo como 10 caracteres.
        # 'anchor="e"' alinha o texto à direita dentro do rótulo.
        self.label_total_geral = tk.Label(
            frame_totais,
            text="0.00",
            font=("Arial", 12, "bold"),
            width=10,
            anchor="e"
        )

        # Posiciona o rótulo de total geral dentro do frame.
        # 'side="left"' posiciona o rótulo à esquerda do frame, ao lado do texto descritivo.
        # 'padx=5' adiciona 5 pixels de espaço horizontal ao redor do rótulo.
        self.label_total_geral.pack(side="left", padx=5)

        # Cria um rótulo para identificar o campo de observações.
        # 'text="Observações:"' define o texto do rótulo como "Observações:", que
        #       será exibido acima do campo de texto.
        # 'font=("Arial", 12)' define a fonte do texto como Arial com tamanho 12.
        lbl_obs = tk.Label(
            self.inner_frame,  # Adiciona o rótulo dentro do frame interno (inner_frame).
            text="Observações:",
            font=("Arial", 12)
        )

        # Posiciona o rótulo no frame.
        # 'pady=5' adiciona 5 pixels de espaço vertical acima e abaixo do rótulo.
        # 'padx=20' adiciona 20 pixels de espaço horizontal ao redor do rótulo.
        # 'anchor="w"' alinha o rótulo à esquerda do frame.
        lbl_obs.pack(pady=5, padx=20, anchor="w")

        # Cria um widget de entrada de texto multilinhas para permitir
        #       que o usuário escreva observações.
        # 'width=80' define a largura do campo de texto como 80 caracteres.
        # 'height=6' define a altura do campo de texto como 6 linhas.
        # 'font=("Arial", 12)' define a fonte do texto digitado como Arial com tamanho 12.
        self.text_obs = tk.Text(self.inner_frame,  # Adiciona o widget de texto dentro do frame interno (inner_frame).
                                width=80,
                                height=6,
                                font=("Arial", 12))

        # Posiciona o campo de texto no frame.
        # 'pady=5' adiciona 5 pixels de espaço vertical acima e abaixo do campo de texto.
        # 'padx=20' adiciona 20 pixels de espaço horizontal ao redor do campo de texto.
        # 'fill="x"' faz com que o campo de texto se expanda horizontalmente
        #       para preencher o espaço disponível.
        self.text_obs.pack(pady=5, padx=20, fill="x")

        # Cria um botão para finalizar o atendimento.
        # 'text="Finalizar Atendimento"' define o texto exibido no botão.
        # 'font=("Arial", 14, "bold")' define a fonte como Arial, tamanho 14 e em negrito.
        # 'bg="#4CAF50"' define a cor de fundo do botão como verde (#4CAF50).
        # 'fg="white"' define a cor do texto do botão como branco.
        # 'command=self.finalizar_atendimento' associa o clique do botão à
        #       função 'finalizar_atendimento'.
        self.btn_finalizar = tk.Button(self.inner_frame,
                                        text="Finalizar Atendimento",
                                        font=("Arial", 14, "bold"),
                                        bg="#4CAF50",
                                        fg="white",
                                        command=self.finalizar_atendimento)

        # Posiciona o botão no frame interno.
        # 'pady=15' adiciona 15 pixels de espaço vertical acima e abaixo do botão.
        self.btn_finalizar.pack(pady=15)

        # Inicializa uma lista para armazenar os produtos disponíveis no estoque.
        # Esta lista será preenchida com os dados carregados do banco de dados.
        self.lista_estoque = []

        # Chama o método privado '_carregar_lista_estoque' para
        #       carregar os produtos do estoque.
        # Este método busca os dados do banco de dados e popula a
        #       lista 'lista_estoque'.
        self._carregar_lista_estoque()

        # Insere um valor inicial padrão no campo de entrada de valor da consulta.
        # 'insert(0, "80.00")' insere o texto "80.00" na posição inicial (índice 0) do campo.
        self.entry_valor_consulta.insert(0, "80.00")

        # Chama o método 'atualizar_totais' para calcular e atualizar os
        #       totais da consulta e dos produtos.
        # Este método soma os valores dos produtos e o valor da consulta
        #       para exibir o total geral.
        self.atualizar_totais()


    def _carregar_lista_estoque(self):

        # Inicializa a lista de produtos do estoque como vazia.
        self.lista_estoque = []

        # Cria uma lista para armazenar os nomes dos produtos
        #       disponíveis no estoque.
        nomes = []

        # Itera sobre cada documento encontrado na coleção 'estoque' do banco de dados.
        for prod in db.estoque.find():

            # Obtém o nome do produto do campo 'nome_produto'.
            # Caso o campo não exista, retorna uma string vazia.
            nome_produto = prod.get("nome_produto", "")

            # Obtém o preço unitário do produto do campo 'preco_unitario'.
            # Caso o campo não exista, retorna o valor padrão 0.0.
            preco_unit = prod.get("preco_unitario", 0.0)

            # Adiciona o nome do produto e o preço unitário como uma
            #       tupla na lista de estoque.
            self.lista_estoque.append((nome_produto, preco_unit))

            # Adiciona apenas o nome do produto à lista de nomes.
            nomes.append(nome_produto)

        # Atualiza os valores da combobox 'combo_produto' com a
        #       lista de nomes dos produtos.
        self.combo_produto["values"] = nomes

        # Se houver produtos na lista de nomes:
        if nomes:

            # Define o primeiro produto da lista como o selecionado
            #       inicialmente na combobox.
            self.combo_produto.current(0)

            # Chama o método '_ao_selecionar_produto' para carregar os
            #       detalhes do produto selecionado.
            # O parâmetro 'None' é passado porque este método também pode
            #       ser chamado como um callback de evento.
            self._ao_selecionar_produto(None)


    def atualizar_totais(self):

        # Calcula o total de produtos no carrinho.
        # 'item[3]' refere-se ao subtotal de cada item no carrinho.
        # 'sum()' soma os subtotais de todos os itens no carrinho.
        total_produtos = sum(item[3] for item in self.carrinho)

        # Atualiza o texto do rótulo 'label_total_produtos' para exibir o total calculado.
        # O valor é formatado com duas casas decimais.
        self.label_total_produtos.config(text=f"{total_produtos:.2f}")

        try:

            # Obtém o valor da consulta digitado pelo usuário no
            #       campo de entrada 'entry_valor_consulta'.
            # Converte o valor para um número do tipo float, removendo
            #       espaços desnecessários com 'strip()'.
            valor_consulta = float(self.entry_valor_consulta.get().strip())

        except:

            # Caso ocorra um erro na conversão (ex.: valor não numérico),
            #       define o valor da consulta como 0.0.
            valor_consulta = 0.0

        # Calcula o total geral somando o valor da consulta com o total de produtos.
        total_geral = valor_consulta + total_produtos

        # Atualiza o texto do rótulo 'label_total_geral' para
        #       exibir o total geral calculado.
        # O valor é formatado com duas casas decimais.
        self.label_total_geral.config(text=f"{total_geral:.2f}")



    def finalizar_atendimento(self):

        try:

            # Tenta obter e converter o valor da consulta fornecido pelo usuário.
            # O valor é retirado do campo de entrada 'entry_valor_consulta',
            #       removendo espaços extras com 'strip()'.
            # Em seguida, converte o valor para um número decimal (float).
            valor_consulta = float(self.entry_valor_consulta.get().strip())

        except:

            # Caso ocorra um erro durante a conversão (por exemplo, se o campo
            #       estiver vazio ou contiver caracteres inválidos),
            #       exibe uma mensagem de erro informando que o valor da consulta é inválido.
            messagebox.showerror("Erro", "Valor da consulta inválido.")

            # Interrompe o processo de finalização do atendimento.
            return

        # Obtém o texto das observações fornecidas pelo usuário.
        # 'get("1.0", tk.END)' captura todo o texto do widget 'Text' desde o
        #       início ("1.0") até o final ("tk.END").
        # 'strip()' remove quaisquer espaços extras ou quebras de linha no
        #       início e no final do texto.
        obs_text = self.text_obs.get("1.0", tk.END).strip()

        # Inicializa uma lista vazia para armazenar os produtos
        #       utilizados no atendimento.
        produtos_usados = []

        # Itera sobre os itens do carrinho (uma lista contendo tuplas de
        #       produtos com nome, preço unitário, quantidade e subtotal).
        for (nome, preco_unit, qtd, subtotal) in self.carrinho:

            # Adiciona cada produto ao dicionário `produtos_usados` com as seguintes informações:
            # - "nome": nome do produto.
            # - "preco_unitario": preço unitário do produto.
            # - "quantidade": quantidade do produto adquirida no atendimento.
            # - "subtotal": preço total para este produto (quantidade * preço unitário).
            produtos_usados.append({"nome": nome,
                                    "preco_unitario": preco_unit,
                                    "quantidade": qtd,
                                    "subtotal": subtotal})

        # Calcula o total dos valores dos produtos no atendimento.
        # Usa uma compreensão de lista para somar os subtotais de
        #       todos os itens em `produtos_usados`.
        total_produtos = sum(item["subtotal"] for item in produtos_usados)

        # Calcula o total geral do atendimento somando o valor da
        #       consulta ao total dos produtos.
        total_geral = valor_consulta + total_produtos

        # Cria um dicionário para registrar as informações detalhadas do atendimento.
        registro = {

            # Armazena a data e hora atual do atendimento usando `datetime.datetime.now()`.
            "data": datetime.datetime.now(),

            # Indica o tipo de registro como "Atendimento".
            "tipo": "Atendimento",

            # Salva o valor da consulta como parte do registro.
            "valor_consulta": valor_consulta,

            # Armazena a lista de produtos utilizados durante o atendimento.
            "produtos_usados": produtos_usados,

            # Registra o total do valor dos produtos utilizados.
            "total_produtos": total_produtos,

            # Calcula e armazena o total geral do atendimento (consulta + produtos).
            "total_geral": total_geral,

            # Inclui as observações fornecidas pelo médico durante o atendimento.
            "observacoes": obs_text,

            # Salva o ID do médico que realizou o atendimento para referência.
            "realizado_por_id": self.medico_logado["_id"],

            # Salva o nome do médico responsável pelo atendimento.
            "realizado_por_nome": self.medico_logado["nome"]

        }

        try:

            # Atualiza o registro do animal no banco de dados.
            # `update_one` localiza o animal pelo seu `_id` e adiciona o novo
            #       registro de atendimento ao campo "historicos".
            db.animais.update_one(

                # Localiza o documento do animal pelo seu ID.
                {"_id": self.animal_doc["_id"]},

                # Adiciona o registro ao array "historicos" usando `$push`.
                {"$push": {"historicos": registro}}

            )

            # Atualiza o status do agendamento no banco de dados para "Realizado".
            # `update_one` localiza o agendamento pelo seu `_id` e modifica o campo "status".
            db.agenda.update_one(

                # Localiza o documento do agendamento pelo seu ID.
                {"_id": self.agendamento_doc["_id"]},

                # Define o status como "Realizado" usando `$set`.
                {"$set": {"status": "Realizado"}}

            )

            # Reduz estoque + registrar transações
            # Para cada item na lista `produtos_usados`, realiza as operações de
            #       atualização de estoque e registro de transações.
            for item in produtos_usados:

                # Obtém o nome do produto e a quantidade utilizada no atendimento.
                nome_produto = item["nome"]
                qtd_utilizada = item["quantidade"]

                # Consulta o banco de dados para encontrar o produto correspondente no estoque.
                produto_estoque = db.estoque.find_one({"nome_produto": nome_produto})

                if produto_estoque:

                    # Calcula a nova quantidade do produto no estoque após a saída.
                    nova_qtd = produto_estoque.get("quantidade", 0) - qtd_utilizada

                    # Verifica se a nova quantidade seria negativa.
                    if nova_qtd < 0:

                        # Mostra um aviso ao usuário caso o estoque fique negativo.
                        messagebox.showwarning(
                            "Estoque insuficiente",
                            f"O produto '{nome_produto}' ficaria com estoque negativo!"
                        )

                    # Atualiza o estoque no banco de dados para refletir a nova quantidade.
                    db.estoque.update_one(

                        # Filtra o produto pelo seu ID único.
                        {"_id": produto_estoque["_id"]},

                        # Define o novo valor de `quantidade`.
                        {"$set": {"quantidade": nova_qtd}}

                    )

                    # Insere um registro de transação no banco de dados `estoque_transacoes`.
                    db.estoque_transacoes.insert_one({

                        "id_produto": produto_estoque["_id"],  # Salva o ID do produto.
                        "data": datetime.datetime.now(),  # Registra a data e hora da transação.
                        "tipo": "Saída",  # Define o tipo da transação como "Saída".
                        "quantidade": qtd_utilizada,  # Salva a quantidade retirada do estoque.
                        "observacoes": f"Saída no atendimento do animal: {self.nome_animal}"
                        # Adiciona uma observação detalhada.

                    })

                else:

                    # Mostra um aviso caso o produto não seja encontrado no
                    #       banco de dados de estoque.
                    messagebox.showwarning(
                        "Produto não encontrado",
                        f"Produto '{nome_produto}' não foi localizado no estoque."
                    )

            # Abre a janela de comprovante do atendimento.
            # Passa as informações do atendimento, como o nome do animal,
            #       nome do dono, e o registro detalhado.
            JanelaComprovanteAtendimento(self, self.nome_animal, self.nome_dono, registro)

            # Atualiza a tela de estoque em tempo real.
            # Verifica se há uma referência para a tela de estoque (self.referencia_estoque).
            # Caso exista, chama o método `carregar_estoque` para recarregar os dados do estoque.
            if self.referencia_estoque:
                self.referencia_estoque.carregar_estoque()

        # Lida com exceções durante o processo de salvar o atendimento.
        # Exibe uma mensagem de erro detalhada para o usuário caso ocorra alguma falha.
        except Exception as e:
            messagebox.showerror("Erro",
                                 f"Não foi possível salvar o atendimento:\n{e}")



    def _ao_selecionar_produto(self, event):

        # Obtém o nome do produto selecionado atualmente na combobox 'combo_produto'.
        prod_nome = self.combo_produto.get()

        # Itera sobre a lista de produtos carregados do estoque.
        # Cada item da lista é uma tupla contendo o nome do produto e seu preço unitário.
        for (nome, preco) in self.lista_estoque:

            # Verifica se o nome do produto na lista corresponde ao nome selecionado.
            if nome == prod_nome:

                # Atualiza o texto do rótulo 'label_preco_unit' com o preço do produto,
                # formatado com duas casas decimais.
                self.label_preco_unit.config(text=f"{preco:.2f}")

                # Interrompe o loop, já que o produto foi encontrado e
                #       não há necessidade de continuar a busca.
                break


    def adicionar_produto_ao_carrinho(self):

        # Obtém o nome do produto selecionado atualmente na combobox 'combo_produto'.
        prod_nome = self.combo_produto.get()

        # Obtém o preço unitário do produto a partir do texto do rótulo 'label_preco_unit'.
        preco_str = self.label_preco_unit.cget("text")

        # Obtém a quantidade digitada no campo de entrada 'entry_qtd' e remove espaços extras.
        qtd_str = self.entry_qtd.get().strip()

        # Converte o preço unitário de string para número decimal (float).
        # Caso ocorra erro na conversão, o valor padrão será 0.0.
        try:
            preco_unit = float(preco_str)
        except:
            preco_unit = 0.0

        # Converte a quantidade de string para número decimal (float).
        # Caso ocorra erro na conversão, o valor padrão será 0.0.
        try:
            qtd = float(qtd_str)
        except:
            qtd = 0.0

        # Verifica se a quantidade é menor ou igual a zero.
        # Mostra um aviso se a quantidade for inválida e interrompe o processo.
        if qtd <= 0:
            messagebox.showwarning("Aviso", "Quantidade deve ser maior que zero.")
            return

        # Calcula o subtotal do produto (preço unitário multiplicado pela quantidade).
        subtotal = preco_unit * qtd

        # Adiciona o produto ao carrinho como uma tupla contendo nome,
        #       preço unitário, quantidade e subtotal.
        self.carrinho.append((prod_nome, preco_unit, qtd, subtotal))

        # Insere o produto no Treeview 'tree_carrinho', formatando os
        #       valores numéricos com duas casas decimais.
        self.tree_carrinho.insert("",
                                  tk.END,
                                  values=(prod_nome, f"{preco_unit:.2f}", f"{qtd:.2f}", f"{subtotal:.2f}"))

        # Atualiza os totais exibidos na interface, como o total de
        #       produtos e o total geral.
        self.atualizar_totais()


    def remover_produto_selecionado(self):

        # Obtém a seleção atual no Treeview 'tree_carrinho'.
        # 'selection()' retorna uma lista de itens selecionados.
        selection = self.tree_carrinho.selection()

        # Verifica se nenhum item foi selecionado no Treeview.
        # Caso a lista 'selection' esteja vazia, exibe uma mensagem de
        #       aviso ao usuário e interrompe o processo.
        if not selection:
            messagebox.showwarning("Aviso",
                                   "Selecione um produto para remover.")
            return

        # Obtém os detalhes do item selecionado no Treeview usando 'item(selection[0])'.
        # 'selection[0]' refere-se ao primeiro (e único) item selecionado, já
        #       que apenas um item pode ser removido por vez.
        item = self.tree_carrinho.item(selection[0])

        # Obtém os valores das colunas do item selecionado.
        # Os valores correspondem a uma tupla com as informações (nome_produto,
        #       preco_unit, qtd, subtotal).
        vals = item["values"]  # (nome_produto, preco_unit, qtd, subtotal)

        # Percorre o carrinho para encontrar o item correspondente.
        # O carrinho é uma lista de tuplas contendo (nome_produto, preco_unit, qtd, subtotal).
        for i, (pn, pr, qt, st) in enumerate(self.carrinho):

            # Compara os valores do Treeview 'vals' com os do carrinho.
            # A comparação verifica o nome do produto, o preço unitário, a
            #       quantidade e o subtotal.
            # Compara o nome do produto e formata o preço com 2 casas decimais.
            if (pn == vals[0] and f"{pr:.2f}" == vals[1]

                    # Verifica também a quantidade e o subtotal.
                    and f"{qt:.2f}" == vals[2] and f"{st:.2f}" == vals[3]):

                # Remove o item do carrinho na posição 'i' quando
                #       encontra uma correspondência.
                self.carrinho.pop(i)

                # Interrompe o loop após encontrar e remover o item correspondente.
                break

        # Remove o item do Treeview usando o método 'delete()' e o
        #       identificador do item selecionado.
        self.tree_carrinho.delete(selection[0])

        # Atualiza os totais exibidos na interface gráfica após a remoção do item.
        # Os totais são recalculados para refletir a remoção do produto.
        self.atualizar_totais()



# -----------------------------------------------------------
# Comprovante
# -----------------------------------------------------------

class JanelaComprovanteAtendimento(tk.Toplevel):

    # Método inicializador da classe 'JanelaComprovanteAtendimento'.
    # Este método é chamado automaticamente quando um objeto desta classe é criado.
    # Ele configura os elementos iniciais necessários para exibir o comprovante do atendimento.
    def __init__(self, parent, nome_animal, nome_dono, registro_atendimento, *args, **kwargs):

        # Chama o método inicializador da classe pai (tk.Toplevel).
        # Isso configura a janela como uma subjanela ou janela modal derivada da janela principal.
        # 'parent' é a janela principal que criou esta janela, usada como referência hierárquica.
        # '*args' e '**kwargs' permitem passar argumentos extras sem alterar a definição do método.
        super().__init__(parent, *args, **kwargs)

        # Define o título da janela que será exibido na barra de título do sistema operacional.
        # 'Comprovante de Atendimento' é o texto que identifica o propósito desta janela.
        self.title("Comprovante de Atendimento")

        # Configura o tamanho inicial da janela em 600 pixels de largura e 900 pixels de altura.
        # O método 'geometry' define as dimensões e a posição inicial da janela na tela.
        self.geometry("600x900")

        # Define a cor de fundo da janela como branca.
        # 'bg="white"' especifica que o fundo será totalmente branco.
        self.configure(bg="white")

        # Cria um container principal (Frame) para organizar os elementos
        #       visuais dentro da janela.
        # Este frame atua como o contêiner principal onde outros
        #       widgets (como rótulos e botões) serão colocados.
        # 'bg="white"' define a cor de fundo do frame para coincidir com o fundo da janela.
        frame_main = tk.Frame(self, bg="white")

        # Posiciona o frame principal dentro da janela para ocupar todo o espaço disponível.
        # 'fill="both"' faz o frame expandir tanto em largura quanto em altura para preencher o espaço.
        # 'expand=True' permite que o frame se ajuste automaticamente caso a janela seja redimensionada.
        # 'padx=20' adiciona 20 pixels de margem horizontal em ambos os lados do frame.
        # 'pady=20' adiciona 20 pixels de margem vertical acima e abaixo do frame.
        frame_main.pack(fill="both",
                        expand=True,
                        padx=20,
                        pady=20)

        # Obtém a data do atendimento a partir do registro fornecido.
        # Se o campo "data" não existir no registro, usa a data e hora atuais como padrão.
        data_atend = registro_atendimento.get("data", datetime.datetime.now())

        # Converte a data do atendimento para uma string formatada no estilo DD/MM/YYYY HH:MM.
        # Isso facilita a exibição da data de forma legível para o usuário.
        data_str = data_atend.strftime("%d/%m/%Y %H:%M")

        # Obtém o valor da consulta do registro.
        # Se o campo "valor_consulta" não existir, o valor padrão será 0.0.
        valor_consulta = registro_atendimento.get("valor_consulta", 0.0)

        # Obtém a lista de produtos usados no atendimento a partir do registro.
        # Se o campo "produtos_usados" não existir, retorna uma lista vazia como padrão.
        produtos_usados = registro_atendimento.get("produtos_usados", [])

        # Obtém o total de produtos (soma dos subtotais dos
        #       produtos usados) a partir do registro.
        # Se o campo "total_produtos" não existir, retorna 0.0 como padrão.
        total_produtos = registro_atendimento.get("total_produtos", 0.0)

        # Obtém o total geral do atendimento (valor da consulta + total de
        #       produtos) a partir do registro.
        # Se o campo "total_geral" não existir, retorna 0.0 como padrão.
        total_geral = registro_atendimento.get("total_geral", 0.0)

        # Obtém o nome do médico que realizou o atendimento a partir do registro.
        # Se o campo "realizado_por_nome" não existir, retorna uma string vazia como padrão.
        medico_nome = registro_atendimento.get("realizado_por_nome", "")

        # Obtém as observações do atendimento a partir do registro.
        # Se o campo "observacoes" não existir, retorna uma string vazia como padrão.
        obs = registro_atendimento.get("observacoes", "")

        # Cria e posiciona o título do comprovante
        # 'text="Comprovante de Atendimento"' define o texto exibido no rótulo.
        # 'font=("Arial", 20, "bold")' configura a fonte do texto com tamanho 20 e em negrito.
        # 'bg="white"' ajusta o fundo do rótulo para branco, combinando com o fundo da janela.
        tk.Label(frame_main,
                text="Comprovante de Atendimento",
                font=("Arial", 20, "bold"),
                bg="white").pack(pady=20)  # 'pady=20' adiciona 20 pixels de espaço vertical acima e abaixo do rótulo.

        # Cria e posiciona o rótulo para exibir a data e hora do atendimento.
        # 'text=f"Data/Hora: {data_str}"' define o texto exibido,
        #       formatando a data e hora do atendimento.
        # 'font=("Arial", 14)' configura a fonte do texto com tamanho 14.
        # 'bg="white"' ajusta o fundo do rótulo para branco.
        # 'anchor="w"' alinha o texto à esquerda.
        tk.Label(frame_main,
                 text=f"Data/Hora: {data_str}",
                 font=("Arial", 14),
                 bg="white").pack(anchor="w", pady=5)

        # Cria e posiciona o rótulo para exibir o nome do animal.
        # 'frame_main' é o widget pai onde o rótulo será posicionado.
        # 'text=f"Animal: {nome_animal}"' define o texto exibido no rótulo,
        #       mostrando o nome do animal associado ao atendimento.
        # 'font=("Arial", 14)' configura a fonte do texto para Arial, tamanho 14.
        # 'bg="white"' ajusta a cor de fundo do rótulo para branco, harmonizando com o fundo da janela.
        # 'pack(anchor="w", pady=5)' posiciona o rótulo alinhado à
        #       esquerda ('anchor="w"') com 5 pixels de espaçamento vertical ('pady=5').
        tk.Label(frame_main,
                 text=f"Animal: {nome_animal}",
                 font=("Arial", 14),
                 bg="white").pack(anchor="w", pady=5)

        # Cria e posiciona o rótulo para exibir o nome do dono do animal.
        # 'frame_main' é o widget pai onde o rótulo será adicionado.
        # 'text=f"Dono: {nome_dono}"' define o texto exibido no rótulo,
        #       utilizando a variável 'nome_dono'.
        # 'font=("Arial", 14)' configura a fonte como Arial, tamanho 14.
        # 'bg="white"' ajusta a cor de fundo do rótulo para branco,
        #       harmonizando com o fundo da janela.
        # 'pack(anchor="w", pady=5)' posiciona o rótulo alinhado à
        #       esquerda ('anchor="w') e com 5 pixels de espaçamento vertical ('pady=5').
        tk.Label(frame_main,
                 text=f"Dono: {nome_dono}",
                 font=("Arial", 14),
                 bg="white").pack(anchor="w", pady=5)

        # Cria e posiciona o rótulo para exibir o nome do médico responsável pelo atendimento.
        # 'frame_main' é o widget pai do rótulo.
        # 'text=f"Médico Responsável: {medico_nome}"' define o texto exibido no
        #       rótulo, utilizando a variável 'medico_nome'.
        # 'font=("Arial", 14)' configura a fonte como Arial, tamanho 14, para
        #       manter consistência com os outros rótulos.
        # 'bg="white"' ajusta a cor de fundo do rótulo para branco.
        # 'pack(anchor="w", pady=5)' posiciona o rótulo alinhado à esquerda
        #       com 5 pixels de espaçamento vertical.
        tk.Label(frame_main,
                 text=f"Médico Responsável: {medico_nome}",
                 font=("Arial", 14),
                 bg="white").pack(anchor="w", pady=5)

        # Cria e posiciona o rótulo para exibir o valor da consulta.
        # 'frame_main' é o widget pai do rótulo.
        # 'text=f"Valor da Consulta: R$ {valor_consulta:.2f}"' define o texto
        #       exibido no rótulo, formatando 'valor_consulta' com duas casas decimais.
        # 'font=("Arial", 14)' configura a fonte como Arial, tamanho 14, destacando o valor da consulta.
        # 'bg="white"' ajusta a cor de fundo do rótulo para branco.
        # 'pack(anchor="w", pady=5)' posiciona o rótulo alinhado à esquerda e
        #       com 5 pixels de espaçamento vertical.
        tk.Label(frame_main,
                 text=f"Valor da Consulta: R$ {valor_consulta:.2f}",
                 font=("Arial", 14),
                 bg="white").pack(anchor="w", pady=5)

        # Cria e posiciona o rótulo para exibir o total dos produtos usados no atendimento.
        # 'frame_main' é o widget pai onde o rótulo será adicionado.
        # 'text=f"Total Produtos: R$ {total_produtos:.2f}"' formata o valor de 'total_produtos'
        #       com duas casas decimais e o exibe no rótulo.
        # 'font=("Arial", 14)' utiliza a fonte Arial no tamanho 14, mantendo
        #       consistência com os outros rótulos.
        # 'bg="white"' ajusta a cor de fundo do rótulo para branco, harmonizando
        #       com o restante da interface.
        # 'pack(anchor="w", pady=(10, 0))' posiciona o rótulo alinhado à
        #       esquerda, com um espaçamento superior de 10 pixels e sem
        #       espaçamento inferior.
        tk.Label(frame_main,
                 text=f"Total Produtos: R$ {total_produtos:.2f}",
                 font=("Arial", 14),
                 bg="white").pack(anchor="w", pady=(10, 0))

        # Cria e posiciona o rótulo para exibir o valor total geral do atendimento.
        # 'text=f"Total Geral: R$ {total_geral:.2f}"' exibe o valor de 'total_geral',
        #       formatado com duas casas decimais.
        # 'font=("Arial", 14, "bold")' define a fonte Arial no tamanho 14 e aplica o
        #       estilo 'bold' para dar destaque ao total geral.
        # 'bg="white"' ajusta a cor de fundo do rótulo para branco, mantendo a
        #       harmonia visual com os outros elementos.
        # 'pack(anchor="w", pady=5)' posiciona o rótulo alinhado à esquerda e
        #       adiciona um espaçamento vertical de 5 pixels para separação visual.
        tk.Label(frame_main,
                 text=f"Total Geral: R$ {total_geral:.2f}",
                 font=("Arial", 14, "bold"),
                 bg="white").pack(anchor="w", pady=5)

        # Verifica se há observações no registro de atendimento.
        if obs:

            # Cria e posiciona o rótulo para identificar o campo de observações.
            # 'text="Observações:"' exibe o título do campo.
            # 'font=("Arial", 14, "bold")' define a fonte Arial, tamanho 14, com
            #       estilo negrito para destacar o título.
            # 'bg="white"' ajusta a cor de fundo para branco, mantendo a
            #       harmonia com os outros elementos.
            # 'pack(anchor="w", pady=(10, 0))' posiciona o rótulo alinhado à esquerda
            #       com 10 pixels de espaçamento superior e nenhum espaçamento inferior.
            tk.Label(frame_main,
                     text="Observações:",
                     font=("Arial", 14, "bold"),
                     bg="white").pack(anchor="w", pady=(10, 0))

            # Cria um widget de texto (Text) para exibir as observações do atendimento.
            # 'width=80' define a largura do campo de texto em 80 caracteres.
            # 'height=5' define a altura do campo de texto em 5 linhas.
            # 'font=("Arial", 12)' define a fonte Arial no tamanho 12, garantindo legibilidade.
            txt_obs = tk.Text(frame_main,
                              width=80,
                              height=5,
                              font=("Arial", 12))

            # Posiciona o campo de texto na interface.
            # 'anchor="w"' alinha o campo à esquerda.
            # 'fill="x"' permite que o campo preencha toda a largura disponível.
            # 'pady=(5, 10)' adiciona 5 pixels de espaçamento superior e
            #       10 pixels de espaçamento inferior.
            txt_obs.pack(anchor="w",
                         fill="x",
                         pady=(5, 10))

            # Insere o texto de observações no campo.
            # 'insert(tk.END, obs)' adiciona o texto 'obs' ao final do campo de texto.
            txt_obs.insert(tk.END, obs)

            # Configura o campo de texto como somente leitura.
            # 'state="disabled"' impede que o usuário edite as observações exibidas.
            txt_obs.config(state="disabled")

        # Cria o botão para gerar o comprovante em formato PDF.
        btn_pdf = tk.Button(

            # Define o container pai como 'frame_main', onde o botão será posicionado.
            frame_main,

            # Configurações do texto do botão.
            text="Gerar PDF",  # O texto exibido no botão é "Gerar PDF".

            # Define o comando que será executado ao clicar no botão.
            # 'lambda' permite passar argumentos específicos para o método 'gerar_pdf'.
            command=lambda: self.gerar_pdf(

                nome_animal,  # Passa o nome do animal como argumento.
                nome_dono,  # Passa o nome do dono do animal como argumento.
                data_str,  # Passa a data formatada do atendimento como argumento.
                medico_nome,  # Passa o nome do médico responsável como argumento.
                produtos_usados,  # Passa a lista de produtos usados no atendimento.
                total_produtos,  # Passa o valor total dos produtos utilizados.
                total_geral,  # Passa o valor total geral (produtos + consulta).
                obs  # Passa as observações do atendimento, se houver.

            ),

            # Configurações de estilo do botão.
            bg="#4CAF50",  # Define a cor de fundo como verde (#4CAF50), indicando ação positiva.
            fg="white",  # Define a cor do texto como branca para contraste com o fundo.
            font=("Arial", 14, "bold")  # Configura a fonte para Arial, tamanho 14, com estilo negrito.

        )

        # Posiciona o botão na interface.
        # 'pack(pady=10)' adiciona 10 pixels de espaçamento vertical ao redor do botão.
        btn_pdf.pack(pady=10)

        # Cria um container com borda e título para exibir os produtos
        #       utilizados no atendimento.
        frame_prods = tk.LabelFrame(

            frame_main,  # Define o container pai como 'frame_main', o frame principal.
            text="Produtos Utilizados",  # Título exibido no topo do LabelFrame.
            font=("Arial", 12, "bold"),  # Define a fonte do título como Arial, tamanho 12, em negrito.
            bg="white",  # Define a cor de fundo do LabelFrame como branco.
            padx=10,  # Adiciona 10 pixels de espaçamento horizontal interno no LabelFrame.
            pady=10  # Adiciona 10 pixels de espaçamento vertical interno no LabelFrame.

        )

        # Posiciona o LabelFrame na interface.
        # 'fill="both"' permite que o LabelFrame preencha todo o espaço
        #       disponível horizontal e verticalmente.
        # 'expand=True' faz com que o LabelFrame se expanda, ocupando mais espaço se disponível.
        # 'pady=20' adiciona 20 pixels de espaçamento vertical ao redor do LabelFrame.
        frame_prods.pack(fill="both", expand=True, pady=20)

        # Define as colunas da tabela de produtos utilizados.
        cols = ("Produto", "Preço Unitário", "Quantidade", "Subtotal")  # Nomes das colunas.

        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria um widget Treeview para exibir os produtos em formato de tabela.
        tree_comprovante = ttk.Treeview(

            frame_prods,  # Define o container pai como 'frame_prods', onde a tabela será exibida.
            columns=cols,  # Define as colunas da tabela como especificado na lista 'cols'.
            show="headings"  # Exibe apenas os cabeçalhos das colunas, sem coluna extra para ícones.

        )

        # Itera sobre cada coluna definida na tabela 'cols'.
        for col in cols:

            # Configura o cabeçalho da coluna no Treeview.
            # 'heading(col, text=col)' define o texto exibido no cabeçalho como o nome da coluna.
            tree_comprovante.heading(col, text=col)

            # Configura as propriedades visuais de cada coluna.
            # 'column(col, anchor="center")' centraliza o texto exibido na coluna.
            tree_comprovante.column(col, anchor="center")

        # Posiciona o Treeview na interface.
        # 'fill="both"' faz o Treeview preencher o espaço disponível
        #       horizontal e verticalmente.
        # 'expand=True' permite que o Treeview expanda para ocupar todo o
        #       espaço disponível dentro do frame.
        tree_comprovante.pack(fill="both", expand=True)

        # Itera sobre a lista de produtos utilizados no atendimento.
        for p in produtos_usados:

            # Adiciona uma nova linha (item) no Treeview.
            # 'insert("", tk.END, values=...)' insere o item no final da tabela.
            # 'values=...' define os valores exibidos nas colunas da linha.
            tree_comprovante.insert(

                "",  # Define o item como filho da raiz do Treeview.
                tk.END,  # Adiciona o item no final da lista.
                values=(
                    p["nome"],  # Nome do produto.
                    f"R$ {p['preco_unitario']:.2f}",  # Preço unitário formatado como moeda.
                    p["quantidade"],  # Quantidade utilizada.
                    f"R$ {p['subtotal']:.2f}"  # Subtotal formatado como moeda.
                )
            )

    def gerar_pdf(self, nome_animal, nome_dono, data_str, medico_nome, produtos_usados, total_produtos, total_geral,
                  obs):

        # Inicia a criação do PDF do comprovante.
        try:

            # Cria um novo objeto PDF utilizando a biblioteca FPDF.
            pdf = FPDF()

            # Adiciona uma nova página ao documento PDF.
            pdf.add_page()

            # Define a fonte padrão para o texto do PDF.
            # 'Arial' é a fonte, e 'size=12' define o tamanho da fonte.
            pdf.set_font("Arial", size=12)

            # Adiciona o título do comprovante na parte superior da página.
            # 'cell(200, 10, txt="...", ln=True, align="C")' cria uma
            #       célula com largura de 200, altura de 10,
            # o texto especificado e centralizado ('align="C"'). 'ln=True'
            #       move o cursor para a próxima linha.
            pdf.cell(200, 10, txt="Comprovante de Atendimento", ln=True, align="C")

            # Adiciona um espaçamento vertical de 10 unidades após o título.
            pdf.ln(10)

            # Adiciona a data e hora do atendimento no PDF.
            # 'cell(0, 10, ...)' cria uma célula que ocupa toda a linha
            #       disponível ('0' de largura).
            # 'ln=True' move o cursor para a próxima linha após exibir o texto.
            pdf.cell(0, 10, txt=f"Data/Hora: {data_str}", ln=True)

            # Adiciona o nome do animal no PDF.
            # 'txt=f"Animal: {nome_animal}"' exibe o texto formatado com o nome do animal.
            pdf.cell(0, 10, txt=f"Animal: {nome_animal}", ln=True)

            # Adiciona o nome do dono do animal no PDF.
            # 'txt=f"Dono: {nome_dono}"' exibe o texto formatado com o nome do dono.
            pdf.cell(0, 10, txt=f"Dono: {nome_dono}", ln=True)

            # Adiciona o nome do médico responsável no PDF.
            # 'txt=f"Médico Responsável: {medico_nome}"' exibe o texto
            #       formatado com o nome do médico.
            pdf.cell(0, 10, txt=f"Médico Responsável: {medico_nome}", ln=True)

            # Adiciona o valor total da consulta no PDF.
            # 'txt=f"Valor da Consulta: R$ {total_geral:.2f}"' exibe o valor
            #       formatado como moeda com duas casas decimais.
            pdf.cell(0, 10, txt=f"Valor da Consulta: R$ {total_geral:.2f}", ln=True)

            # Adiciona um espaçamento vertical de 10 unidades antes de
            #       continuar com outros dados.
            pdf.ln(10)

            # Adiciona o título para a seção de produtos utilizados.
            # 'txt="Produtos Utilizados:"' indica que a lista de produtos
            #       utilizados será detalhada abaixo.
            # 'ln=True' move o cursor para a próxima linha após exibir o texto.
            pdf.cell(0, 10, txt="Produtos Utilizados:", ln=True)

            # Reduz o tamanho da fonte para 10 para exibir os detalhes dos
            #       produtos em um formato mais compacto.
            pdf.set_font("Arial", size=10)

            # Percorre a lista de produtos utilizados e adiciona cada produto ao PDF.
            for p in produtos_usados:

                # Cria uma linha para cada produto, formatada com nome,
                #       quantidade, preço unitário e subtotal.
                # Exemplo: "- Produto A: 2 x R$ 10.00 = R$ 20.00"
                pdf.cell(0, 10,
                         txt=f"- {p['nome']}: {p['quantidade']} x R$ {p['preco_unitario']:.2f} = R$ {p['subtotal']:.2f}",
                         ln=True)

            # Adiciona um espaço vertical de 10 unidades para separar os produtos da próxima seção.
            pdf.ln(10)

            # Retorna ao tamanho de fonte 12 para as próximas informações.
            pdf.set_font("Arial", size=12)

            # Adiciona o total de produtos ao PDF.
            # 'txt=f"Total Produtos: R$ {total_produtos:.2f}"' exibe o
            #       total de produtos formatado como moeda.
            pdf.cell(0, 10, txt=f"Total Produtos: R$ {total_produtos:.2f}", ln=True)

            # Adiciona o total geral ao PDF.
            # 'txt=f"Total Geral: R$ {total_geral:.2f}"' exibe o total geral formatado como moeda.
            pdf.cell(0, 10, txt=f"Total Geral: R$ {total_geral:.2f}", ln=True)

            # Verifica se há observações no registro do atendimento.
            # Se o campo 'obs' não estiver vazio, as observações serão adicionadas ao PDF.
            if obs:

                # Adiciona um espaço vertical antes das observações.
                pdf.ln(10)

                # Adiciona um título para a seção de observações no PDF.
                # 'txt="Observações:"' indica que o texto a seguir será
                #       detalhado como observações.
                # 'ln=True' move o cursor para a próxima linha após exibir o texto.
                pdf.cell(0, 10, txt="Observações:", ln=True)

                # Adiciona o texto das observações ao PDF.
                # 'multi_cell' permite que o texto seja quebrado automaticamente em várias linhas,
                # respeitando a largura disponível (0 indica largura total).
                pdf.multi_cell(0, 10, txt=obs)

            # Salva o PDF no arquivo 'comprovante_atendimento.pdf'.
            # 'output' é o método que gera e salva o arquivo PDF no diretório atual.
            pdf.output("comprovante_atendimento.pdf")

            # Exibe uma mensagem de sucesso ao usuário informando que o PDF foi gerado.
            messagebox.showinfo("Sucesso", "PDF gerado com sucesso!")

        except Exception as e:

            # Caso ocorra um erro durante a geração do PDF, exibe uma mensagem de erro ao usuário.
            # A mensagem inclui detalhes do erro ('e') para ajudar no diagnóstico do problema.
            messagebox.showerror("Erro", f"Erro ao gerar PDF: {e}")




# ---------------------------------------------------------------------
# TelaPrincipal (Calendário + Agendamentos do dia + Botão "Atender" + Filtro)
# ---------------------------------------------------------------------
class TelaPrincipal(tk.Frame):

    # Método inicializador da classe 'TelaPrincipal'
    # Este método é chamado automaticamente ao criar uma nova instância da classe.
    # Ele configura a estrutura inicial e os elementos da tela principal do sistema.
    def __init__(self, master, medico_logado, *args, **kwargs):

        # 'self' é uma referência à instância atual da classe. Ele é usado para
        #       acessar atributos e métodos da própria instância.
        # 'master' é o widget pai, ou seja, o widget onde este frame será
        #       inserido. Geralmente, será o notebook ou janela principal.

        # 'medico_logado' contém informações sobre o médico autenticado no sistema.
        # Ele será usado para personalizar a interface ou carregar dados relacionados ao usuário.

        # '*args' captura argumentos posicionais adicionais passados para o inicializador da classe.
        # Por exemplo, se o frame precisar de configurações específicas, como uma cor de fundo.

        # '**kwargs' captura argumentos nomeados adicionais passados para o inicializador da classe.
        # Por exemplo, podemos definir 'bg="blue"' ou 'height=300' ao criar a instância.

        # Chama o inicializador da classe pai (tk.Frame), garantindo que o
        #       frame seja configurado corretamente.
        super().__init__(master, *args, **kwargs)

        # Armazena o médico logado como um atributo da instância.
        # Isso permite que outros métodos da classe acessem facilmente as
        #       informações do médico autenticado.
        self.medico_logado = medico_logado

        # Inicializa uma lista vazia que pode ser usada para armazenar
        #       informações do dia, como agendamentos.
        # Será manipulada posteriormente para exibir ou processar dados relevantes.
        self.lista_dia = []

        # Cria um 'LabelFrame' para agrupar elementos relacionados à seleção de data e filtros.
        # 'self' indica que o frame pertence à instância atual de 'TelaPrincipal'.
        # Define o widget pai como o frame atual ('TelaPrincipal').
        frame_cal = tk.LabelFrame(self,

                                  # Define o título do LabelFrame como "Selecione a Data e Filtros".
                                  text="Selecione a Data e Filtros",

                                  # Configura a fonte do título do LabelFrame.
                                  font=("Arial", 14, "bold"),  # Usa a fonte Arial, tamanho 14, em negrito.

                                  # Configura o preenchimento interno (espaço entre as bordas
                                  #         internas do LabelFrame e seu conteúdo).
                                  padx=15,  # Adiciona 15 pixels de espaço horizontal interno.

                                  pady=15,  # Adiciona 15 pixels de espaço vertical interno.

                                  # Define a cor de fundo do LabelFrame.
                                  bg="#F9F9F9",  # Cor de fundo em tom claro (#F9F9F9).

                                  # Configura o estilo das bordas do LabelFrame.
                                  relief="groove",  # Estilo de borda em sulco para dar uma aparência tridimensional.

                                  # Define a largura da borda.
                                  bd=2  # Bordas com 2 pixels de espessura.
                                  )

        # Posiciona o LabelFrame dentro do layout.
        # 'side="left"' posiciona o frame à esquerda do espaço disponível.
        # 'fill="y"' ajusta a altura do frame para preencher verticalmente o espaço disponível.
        # 'padx=20' adiciona 20 pixels de espaço externo horizontal (esquerda e direita).
        # 'pady=20' adiciona 20 pixels de espaço externo vertical (acima e abaixo).
        frame_cal.pack(side="left",
                       fill="y",
                       padx=20,
                       pady=20)

        # Cria um widget de calendário dentro do 'frame_cal'.
        # 'self.cal' armazena a referência ao calendário para uso posterior.
        self.cal = Calendar(frame_cal,  # Define o 'frame_cal' como o widget pai do calendário.

                            # Configura o modo de seleção para 'day' (seleção de um único dia).
                            selectmode='day',

                            # Define o padrão de exibição da data no formato 'dd/mm/yyyy'.
                            date_pattern='dd/mm/yyyy',

                            # Configura a fonte do calendário.
                            font=("Arial", 12))  # Fonte Arial com tamanho 12.


        # Posiciona o calendário no layout.
        # 'pady=10' adiciona 10 pixels de espaço vertical acima e abaixo do calendário.
        self.cal.pack(pady=10)

        # Cria um botão para listar os agendamentos do dia selecionado no calendário.
        # Define o 'frame_cal' como o widget pai do botão.
        btn_ver = tk.Button(frame_cal,

                            # Define o texto exibido no botão.
                            text="Ver Agendamentos do Dia",

                            # Configura a fonte do texto no botão.
                            font=("Arial", 12, "bold"),  # Fonte Arial, tamanho 12, em negrito.

                            # Define a cor de fundo do botão.
                            bg="#4CAF50",  # Verde (#4CAF50) para destacar a ação.

                            # Define a cor do texto do botão.
                            fg="white",  # Texto branco para contraste.

                            # Associa o botão ao método 'listar_agendamentos_dia'.
                            # Quando clicado, o botão executa esta função.
                            command=self.listar_agendamentos_dia)

        # Posiciona o botão no layout.
        # 'pady=10' adiciona 10 pixels de espaço vertical acima e abaixo do botão.
        # 'fill="x"' faz o botão preencher horizontalmente o espaço
        #       disponível dentro do 'frame_cal'.
        btn_ver.pack(pady=10, fill="x")

        # Cria um frame para conter o campo de filtro de agendamentos.
        # O frame é um contêiner que organiza os widgets relacionados ao filtro.
        filtro_frame = tk.Frame(frame_cal,  # Define 'frame_cal' como o widget pai.
                                bg="#F9F9F9")  # Define a cor de fundo do frame como um tom claro de cinza.


        # Posiciona o frame no layout.
        # 'pady=20' adiciona 20 pixels de espaço vertical acima e abaixo do frame.
        # 'fill="x"' faz com que o frame preencha horizontalmente o espaço disponível.
        filtro_frame.pack(pady=20, fill="x")

        # Cria um rótulo para identificar o filtro de agendamentos.
        tk.Label( filtro_frame,  # Define o 'filtro_frame' como o widget pai.
                text="Filtro de Agendamentos:",  # Define o texto exibido no rótulo.
                font=("Arial", 12),  # Configura a fonte para Arial, tamanho 12.
                bg="#F9F9F9"  # Define a cor de fundo do rótulo para combinar com o frame.

        ).grid(row=0,  # Define a linha 0 no grid do 'filtro_frame'.
                column=0,  # Define a coluna 0 no grid.
                sticky="w",  # Alinha o rótulo à esquerda dentro da célula do grid.
                padx=5)  # Adiciona 5 pixels de espaço horizontal à esquerda do rótulo.


        # Cria um campo de entrada de texto para inserir o filtro de agendamentos.
        self.entry_filtro = tk.Entry( filtro_frame,  # Define o 'filtro_frame' como o widget pai.
                                    font=("Arial", 12),  # Configura a fonte para Arial, tamanho 12.
                                    width=25,  # Define a largura do campo de entrada como 25 caracteres.
                                    relief="solid",  # Configura o estilo da borda do campo como sólida.
                                    bd=1)  # Define a espessura da borda como 1 pixel.

        # Posiciona o campo de entrada no layout.
        # 'row=1' posiciona o campo de entrada na linha 1 do grid do 'filtro_frame'.
        # 'column=0' posiciona o campo na primeira coluna.
        # 'pady=5' adiciona 5 pixels de espaço vertical acima e abaixo do campo.
        # 'padx=5' adiciona 5 pixels de espaço horizontal à esquerda e à direita do campo.
        # 'sticky="w"' alinha o campo de entrada à esquerda dentro da célula do grid.
        self.entry_filtro.grid(row=1,
                               column=0,
                               pady=5,
                               padx=5,
                               sticky="w")

        # Cria um botão para aplicar o filtro de agendamentos.
        btn_filtrar = tk.Button(filtro_frame,  # Define o 'filtro_frame' como o widget pai.
                                text="Aplicar Filtro",  # Texto exibido no botão.
                                font=("Arial", 11, "bold"),  # Configura a fonte para Arial, tamanho 11, em negrito.
                                bg="#2196F3",  # Define a cor de fundo do botão como azul (código hexadecimal).
                                fg="white",  # Define a cor do texto no botão como branco.
                                command=self.filtrar_local) # Define a função 'self.filtrar_local' que será chamada quando o botão for clicado.


        # Posiciona o botão no layout do 'filtro_frame'.
        # 'row=1' coloca o botão na mesma linha do campo de entrada.
        # 'column=1' coloca o botão na segunda coluna do grid.
        # 'padx=5' adiciona 5 pixels de espaço horizontal à esquerda e à direita do botão.
        # 'pady=5' adiciona 5 pixels de espaço vertical acima e abaixo do botão.
        btn_filtrar.grid(row=1, column=1, padx=5, pady=5)

        # Cria um botão para iniciar o atendimento de um animal.
        btn_atender = tk.Button(frame_cal,  # Define o 'frame_cal' como o widget pai onde o botão será posicionado.
                                text="Atender Animal",  # Texto exibido no botão.
                                font=("Arial", 12, "bold"),  # Define a fonte do texto como Arial, tamanho 12, em negrito.
                                bg="#FF9800",  # Define a cor de fundo do botão como laranja (código hexadecimal #FF9800).
                                fg="white",  # Define a cor do texto no botão como branco.
                                command=self.atender_animal)  # Vincula a ação de clicar no botão à função 'self.atender_animal'.


        # Posiciona o botão no layout do 'frame_cal'.
        # 'pady=30' adiciona 30 pixels de espaço vertical acima e abaixo do botão.
        # 'fill="x"' faz o botão preencher toda a largura disponível no widget pai ('frame_cal').
        btn_atender.pack(pady=30, fill="x")

        # Cria um frame rotulado para exibir a lista de agendamentos do dia.
        frame_listagem = tk.LabelFrame(self,  # Define o frame principal como widget pai deste LabelFrame.
                                        text="Agendamentos do Dia",  # Define o texto do rótulo exibido na borda superior do frame.
                                        font=("Arial", 14, "bold"),  # Configura o estilo do texto do rótulo para Arial, tamanho 14, em negrito.
                                        padx=15,  # Adiciona 15 pixels de preenchimento interno horizontal ao conteúdo dentro do frame.
                                        pady=15,  # Adiciona 15 pixels de preenchimento interno vertical ao conteúdo dentro do frame.
                                        bg="#F9F9F9",  # Define a cor de fundo do frame como um cinza claro.
                                        relief="groove",  # Configura o estilo da borda como "groove" (entalhado).
                                        bd=2)  # Define a largura da borda do frame como 2 pixels.


        # Posiciona o frame no layout geral da janela.
        # 'side="right"' posiciona o frame no lado direito da janela.
        # 'fill="both"' faz com que o frame preencha todo o espaço disponível em
        #       ambas as direções (horizontal e vertical).
        # 'expand=True' permite que o frame expanda para ocupar espaço extra disponível.
        # 'padx=20' adiciona 20 pixels de margem horizontal externa ao redor do frame.
        # 'pady=20' adiciona 20 pixels de margem vertical externa ao redor do frame.
        frame_listagem.pack(side="right",
                            fill="both",
                            expand=True,
                            padx=20,
                            pady=20)

        # Define as colunas para a Treeview, que exibe os dados dos agendamentos.
        # Cada item na tupla 'colunas' representa uma coluna com um nome interno para identificação.
        colunas = ("_id", "animal", "medico", "data", "tipo", "status", "produtos")


        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria uma Treeview dentro do 'frame_listagem'.
        # 'columns=colunas' define as colunas da Treeview com base na tupla 'colunas'.
        # 'show="headings"' exibe apenas os cabeçalhos das colunas, sem
        #       uma coluna extra para ícones.
        # 'height=20' define o número de linhas visíveis na Treeview
        #       antes de precisar de uma barra de rolagem.
        self.tree = ttk.Treeview(frame_listagem,
                                columns=colunas,
                                show="headings",
                                height=20)

        # Configurações de estilo para a Treeview.

        # Cria um estilo para os cabeçalhos da Treeview.
        # 'Treeview.Heading' é o estilo usado para os cabeçalhos.
        # 'font=("Arial", 12, "bold")' define a fonte dos cabeçalhos
        #       como Arial, tamanho 12, em negrito.
        estilo_treeview = ttk.Style()
        estilo_treeview.configure("Treeview.Heading",
                                  font=("Arial", 12, "bold"))

        # Cria um estilo para as linhas de dados na Treeview.
        # 'Treeview' é o estilo usado para as células da tabela.
        # 'font=("Arial", 11)' define a fonte das células como Arial, tamanho 11.
        estilo_treeview.configure("Treeview",
                                  font=("Arial", 11))

        # Define os cabeçalhos para cada coluna na Treeview.
        # O loop percorre a lista de colunas definidas anteriormente.
        for c in colunas:

            # Configura o cabeçalho de cada coluna.
            # 'c' é o identificador interno da coluna.
            # 'text=c' define o texto visível no cabeçalho da coluna.
            self.tree.heading(c, text=c)

            # Define o alinhamento do texto na coluna.
            # 'anchor="center"' centraliza o texto dentro das células da coluna.
            self.tree.column(c, anchor="center")

        # Configurações de largura para colunas específicas.
        # A largura é ajustada para permitir que os dados sejam exibidos de
        #       forma clara e legível.
        # '_id': Coluna de identificador único, normalmente não
        #       precisa de muita largura.
        self.tree.column("_id", width=100)

        # 'animal': Exibe o nome do animal, pode conter nomes longos,
        #       então precisa de largura extra.
        self.tree.column("animal", width=150)

        # 'medico': Exibe o nome do médico responsável, largura
        #       semelhante ao nome do animal.
        self.tree.column("medico", width=150)

        # 'data': Exibe a data do agendamento, largura ajustada
        #       para formato de data/hora.
        self.tree.column("data", width=150)

        # 'tipo': Exibe o tipo de atendimento, precisa de largura
        #       para termos descritivos.
        self.tree.column("tipo", width=150)

        # 'status': Exibe o status do agendamento (Ex: Realizado, Pendente),
        #       largura suficiente para textos curtos.
        self.tree.column("status", width=100)

        # 'produtos': Exibe a lista de produtos utilizados no atendimento,
        #       requer largura maior para detalhes.
        self.tree.column("produtos", width=200)

        # Posiciona e dimensiona a Treeview no frame de listagem.
        # 'fill="both"' faz com que a Treeview preencha todo o espaço disponível no frame.
        # 'expand=True' permite que a Treeview se ajuste dinamicamente ao
        #       redimensionar a janela.
        self.tree.pack(fill="both", expand=True)


    # Define o método atender_animal, que é chamado quando o
    #       usuário deseja iniciar o atendimento de um animal.
    def atender_animal(self):

        # Obtém o item selecionado na Treeview.
        # 'self.tree.selection()' retorna uma tupla com os
        #       identificadores dos itens selecionados.
        selection = self.tree.selection()

        # Verifica se nenhum item foi selecionado.
        if not selection:

            # Exibe uma mensagem de aviso ao usuário informando que é
            #       necessário selecionar um agendamento.
            messagebox.showwarning("Aviso",
                                   "Selecione um agendamento.")

            # Sai do método, pois não há um item para processar.
            return

        # Obtém os dados do item selecionado na Treeview.
        # 'self.tree.item(selection[0])' retorna um dicionário
        #       com informações sobre o item.
        item = self.tree.item(selection[0])

        # Acessa os valores associados ao item selecionado.
        # Os valores correspondem às colunas definidas na
        #       Treeview (ex.: _id, animal, médico, etc.).
        vals = item["values"]

        # O primeiro valor da linha (índice 0) é o ID do
        #       agendamento no banco de dados.
        agendamento_id_str = vals[0]

        # O segundo valor da linha (índice 1) é o nome do
        #       animal associado ao agendamento.
        nome_animal = vals[1]

        # Busca o agendamento no banco de dados com base no ID selecionado.
        # 'ObjectId(agendamento_id_str)' converte a string do ID em um
        #       formato reconhecido pelo MongoDB.
        agendamento = db.agenda.find_one({"_id": ObjectId(agendamento_id_str)})

        # Verifica se o status do agendamento não permite o atendimento.
        # Somente agendamentos com status "Pendente" ou "Confirmado" podem ser atendidos.
        if agendamento["status"] not in ["Pendente", "Confirmado"]:

            # Exibe um aviso ao usuário informando que o status atual do
            #       agendamento não permite atendimento.
            messagebox.showwarning("Aviso",
                                   f"Agendamento com status '{agendamento['status']}' não pode ser atendido.")

            # Interrompe o método, já que o atendimento não pode continuar.
            return

        # Prepara a referência da aba de estoque para possibilitar atualizações em tempo real.
        # Inicialmente, definimos a referência como 'None', pois ela será
        #       atribuída posteriormente se necessário.
        referencia_aba_estoque = None

        try:

            # Acessa a aba de estoque para obter sua referência.
            # 'self.master' se refere ao 'Notebook' que contém as abas.
            # 'self.master.master' se refere à janela principal (JanelaPrincipal) que contém o 'Notebook'.
            # 'aba_estoque' é a aba de estoque que será referenciada
            #       para atualizações em tempo real.
            referencia_aba_estoque = self.master.master.aba_estoque

        except:

            # Em caso de erro (por exemplo, se a aba de estoque não estiver
            #       acessível), continua sem referência.
            pass

        # Cria uma nova instância da janela de atendimento.
        # Passa os seguintes argumentos:
        # - 'self': a instância da classe atual, usada como pai da nova janela.
        # - 'agendamento_id_str': o ID do agendamento selecionado.
        # - 'nome_animal': o nome do animal selecionado para atendimento.
        # - 'self.medico_logado': o médico atualmente logado.
        # - 'referencia_aba_estoque': a referência à aba de estoque para
        #       atualizações em tempo real (se disponível).
        JanelaAtendimento(self,
                          agendamento_id_str,
                          nome_animal,
                          self.medico_logado,
                          referencia_aba_estoque)


    # Define o método filtrar_local para aplicar filtros nos
    #       agendamentos exibidos na Treeview.
    def filtrar_local(self):

        # Obtém o texto do filtro digitado no campo de entrada 'entry_filtro'.
        # 'strip()' remove espaços em branco no início e no fim, e 'lower()'
        #       converte o texto para minúsculas para comparação insensível a
        #       maiúsculas/minúsculas.
        filtro = self.entry_filtro.get().strip().lower()

        # Verifica se o campo de filtro está vazio.
        if not filtro:

            # Se o filtro estiver vazio, exibe a lista completa de agendamentos do dia.
            self._exibir_lista_dia(self.lista_dia)

            # Sai do método, pois nenhuma filtragem adicional é necessária.
            return

        # Lista para armazenar os registros que atendem ao critério de filtro.
        filtrados = []

        # Itera sobre os registros de agendamentos do dia.
        for reg in self.lista_dia:

            # Combina todos os valores do registro em uma única string,
            #       convertendo-os para minúsculas.
            # 'str(v)' garante que todos os valores sejam convertidos para string,
            #       caso sejam números ou outros tipos.
            row_text = " ".join(str(v).lower() for v in reg.values())

            # Verifica se o texto do filtro está presente em qualquer
            #       parte do texto combinado da linha.
            if filtro in row_text:

                # Se o filtro corresponder, adiciona o registro à
                #       lista de registros filtrados.
                filtrados.append(reg)

        # Exibe a lista de registros filtrados na Treeview.
        self._exibir_lista_dia(filtrados)


    def listar_agendamentos_dia(self):

        # Limpa todos os itens atualmente exibidos na Treeview.
        # 'self.tree.get_children()' retorna todos os itens presentes na Treeview.
        for item in self.tree.get_children():
            self.tree.delete(item)  # Remove cada item da Treeview.

        # Limpa a lista que armazena os agendamentos do dia.
        # Garante que os dados anteriores não interfiram nos novos resultados.
        self.lista_dia.clear()

        # Obtém a data selecionada no calendário.
        # O método 'self.cal.get_date()' retorna a data no formato 'dd/mm/yyyy'.
        data_str = self.cal.get_date()

        try:

            # Divide a data selecionada em dia, mês e ano.
            # O método 'split("/")' separa a string 'data_str' em partes
            #       usando o caractere '/' como delimitador.
            dia, mes, ano = data_str.split("/")

            # Constrói o intervalo de tempo para o dia selecionado.
            # 'dt_ini' representa o início do dia (00:00:00).
            dt_ini = datetime.datetime(int(ano), int(mes), int(dia), 0, 0, 0)

            # 'dt_fim' representa o final do dia (23:59:59).
            # Isso é feito somando 1 dia a 'dt_ini'.
            dt_fim = dt_ini + datetime.timedelta(days=1)

        except:

            # Exibe uma mensagem de erro caso ocorra algum problema ao processar a data.
            # Isso pode acontecer se a data estiver em um formato inesperado.
            messagebox.showerror("Erro", "Data inválida.")
            return

        # Busca os agendamentos no banco de dados com base no intervalo de data selecionado.
        # 'db.agenda.find()' realiza uma consulta no banco MongoDB na coleção 'agenda'.
        # O filtro 'data_agendamento' seleciona registros com datas
        #       entre 'dt_ini' (início do dia) e 'dt_fim' (fim do dia).
        # '.sort("data_agendamento", 1)' organiza os resultados em
        #       ordem crescente pela data do agendamento.
        agendamentos = db.agenda.find({
            "data_agendamento": {"$gte": dt_ini, "$lt": dt_fim}
        }).sort("data_agendamento", 1)

        # Itera sobre cada agendamento encontrado na consulta.
        for ag in agendamentos:

            # Busca o documento do animal associado ao agendamento.
            # 'ag["id_animal"]' é o identificador do animal armazenado no agendamento.
            animal_doc = db.animais.find_one({"_id": ag["id_animal"]})

            # Obtém o nome do animal do documento encontrado.
            # Se o documento não for encontrado ou o campo 'nome_animal'
            #       estiver ausente, retorna '??'.
            nome_animal = animal_doc.get("nome_animal", "??") if animal_doc else "??"

            # Inicializa o nome do médico como uma string vazia.
            # Esse valor será preenchido caso o agendamento tenha um médico associado.
            nome_medico = ""

            # Verifica se o agendamento possui um médico associado ('id_medico').
            if ag.get("id_medico"):

                # Busca o documento do médico no banco de dados.
                # 'ag["id_medico"]' é o identificador do médico armazenado no agendamento.
                medico_doc = db.medicos.find_one({"_id": ag["id_medico"]})

                # Obtém o nome do médico do documento encontrado.
                # Se o documento não for encontrado ou o campo 'nome' estiver
                #       ausente, mantém o valor vazio.
                nome_medico = medico_doc.get("nome", "") if medico_doc else ""

            # Formata a data e hora do agendamento para exibição.
            # 'formatar_data_hora' é uma função auxiliar que converte
            #       objetos de data e hora em strings legíveis.
            data_fmt = formatar_data_hora(ag.get("data_agendamento"))

            # Combina os produtos usados no agendamento em uma única string separada por vírgulas.
            # Se o campo 'produtos_usados' não existir ou estiver vazio, retorna uma string vazia.
            produtos = ", ".join(ag.get("produtos_usados", []))

            # Cria um dicionário para armazenar as informações de cada
            #       agendamento encontrado.
            registro = {

                "_id": str(ag["_id"]),  # Converte o ID do agendamento para string para facilitar a manipulação.
                "animal": nome_animal,  # Adiciona o nome do animal associado ao agendamento.
                "medico": nome_medico,  # Adiciona o nome do médico responsável pelo agendamento.
                "data": data_fmt,  # Adiciona a data do agendamento formatada como string.
                "tipo": ag.get("tipo_consulta", ""),  # Obtém o tipo de consulta ou usa uma string vazia como padrão.
                "status": ag.get("status", ""),  # Obtém o status do agendamento ou usa uma string vazia como padrão.
                "produtos": produtos  # Adiciona a lista de produtos utilizados no agendamento como uma string.

            }

            # Adiciona o dicionário criado à lista de agendamentos do dia.
            # Essa lista será usada para exibir os agendamentos na interface.
            self.lista_dia.append(registro)

        # Após processar todos os agendamentos, exibe a lista completa na interface.
        # 'self._exibir_lista_dia()' é um método que popula a
        #       interface com os dados processados.
        self._exibir_lista_dia(self.lista_dia)

    # Define o método _exibir_lista_dia para exibir os
    #       agendamentos do dia na interface.
    def _exibir_lista_dia(self, lista):

        # Limpa todos os itens existentes na Treeview.
        # 'self.tree.get_children()' retorna os identificadores de
        #       todos os itens na Treeview.
        for item in self.tree.get_children():

            # Remove cada item da Treeview usando o método 'delete'.
            self.tree.delete(item)

        # Itera sobre a lista de registros de agendamentos para exibi-los.
        for reg in lista:

            # Insere cada registro na Treeview.
            # Os valores são extraídos do dicionário 'reg' e associados às colunas configuradas.
            self.tree.insert(

                "",
                # O primeiro argumento é o identificador pai. Aqui, usamos uma
                #       string vazia para indicar que o item não tem pai (nível raiz).
                tk.END,  # O segundo argumento indica onde o item será adicionado. 'tk.END' significa no final da lista.
                values=(  # Os valores a serem exibidos nas colunas da Treeview.
                    reg["_id"],  # Identificador único do agendamento.
                    reg["animal"],  # Nome do animal associado ao agendamento.
                    reg["medico"],  # Nome do médico responsável pelo agendamento.
                    reg["data"],  # Data formatada do agendamento.
                    reg["tipo"],  # Tipo de consulta (ex.: "Rotina", "Emergência").
                    reg["status"],  # Status do agendamento (ex.: "Pendente", "Concluído").
                    reg["produtos"]  # Produtos utilizados durante o atendimento, se aplicável.
                )

            )



# ---------------------------------------------------------------------
# TelaClientes (CRUD)
# ---------------------------------------------------------------------
# Define a classe 'TelaClientes', que herda de 'tk.Frame'. Essa classe
#       representa a interface para gerenciamento de clientes.
class TelaClientes(tk.Frame):

    # Método inicializador da classe 'TelaClientes', chamado
    #       automaticamente quando um objeto desta classe é criado.
    def __init__(self, master, medico_logado, *args, **kwargs):

        # Chama o método inicializador da classe pai ('tk.Frame').
        # O 'super()' permite acessar métodos da classe pai (no caso, 'tk.Frame').
        # Isso configura o frame base para que ele possa ser usado como parte da interface.
        # 'master' é o widget pai onde este frame será exibido, como um notebook ou janela principal.
        # '*args' e '**kwargs' permitem passar argumentos adicionais que o 'tk.Frame' possa aceitar.
        super().__init__(master, *args, **kwargs)

        # Armazena as informações do médico logado para possíveis
        #       interações ou permissões específicas.
        self.medico_logado = medico_logado

        # Define a cor de fundo da tela (frame) como um cinza claro.
        self.configure(bg="#F7F7F7")

        # Cria um rótulo que serve como título da tela.
        # 'text="Gerenciamento de Clientes"' define o texto exibido.
        # 'font=("Arial", 18, "bold")' configura a fonte, tamanho e estilo (negrito).
        # 'bg="#F7F7F7"' ajusta a cor de fundo do rótulo para coincidir com a do frame.
        # 'fg="#333333"' define a cor do texto como cinza escuro.
        titulo = tk.Label(
            self,
            text="Gerenciamento de Clientes",
            font=("Arial", 18, "bold"),
            bg="#F7F7F7",
            fg="#333333"
        )

        # Posiciona o rótulo na tela com um espaçamento
        #       vertical de 10 pixels ('pady=10').
        titulo.pack(pady=10)

        # Cria um frame rotulado ('LabelFrame') para o formulário de cadastro de cliente.
        frame_form = tk.LabelFrame(

            self,  # O widget pai deste frame é o próprio frame da tela de clientes ('self').
            text="Cadastro de Cliente",  # Define o texto exibido como título do frame.
            font=("Arial", 12, "bold"),  # Define a fonte do título como Arial, tamanho 12, em negrito.
            bg="#FFFFFF",  # Define a cor de fundo do frame como branco.
            fg="#333333",  # Define a cor do texto do título como cinza escuro.
            padx=10,  # Define um espaçamento horizontal interno (padding) de 10 pixels dentro do frame.
            pady=10  # Define um espaçamento vertical interno (padding) de 10 pixels dentro do frame.

        )

        # Posiciona o frame no topo do widget pai ('self').
        # 'side="top"' alinha o frame no topo.
        # 'fill="x"' permite que o frame ocupe toda a largura disponível no eixo horizontal.
        # 'padx=20' adiciona 20 pixels de espaçamento horizontal
        #       externo (margem) em ambos os lados do frame.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical
        #       externo (margem) acima e abaixo do frame.
        frame_form.pack(side="top", fill="x", padx=20, pady=10)

        # Define uma lista de campos para o formulário de cadastro de cliente.
        # Cada elemento da lista é uma tupla com:
        # 1. O texto do rótulo a ser exibido ao lado do campo.
        # 2. O nome da variável onde será armazenado o widget 'Entry' correspondente.
        campos = [

            ("Nome:", "entry_nome"),  # Rótulo "Nome:" e variável associada 'entry_nome'.
            ("CPF:", "entry_cpf"),  # Rótulo "CPF:" e variável associada 'entry_cpf'.
            ("Telefone:", "entry_telefone"),  # Rótulo "Telefone:" e variável associada 'entry_telefone'.
            ("Endereço:", "entry_endereco"),  # Rótulo "Endereço:" e variável associada 'entry_endereco'.
            ("E-mail:", "entry_email")  # Rótulo "E-mail:" e variável associada 'entry_email'.

        ]

        # Inicializa a variável que armazenará o ID do cliente
        #       atualmente selecionado ou em edição.
        # O valor padrão é 'None', indicando que nenhum cliente está selecionado.
        self.cliente_id_atual = None

        # Loop para criar dinamicamente os campos do formulário com base na lista 'campos'.
        for i, (label_text, var_name) in enumerate(campos):

            # Cria um rótulo ('Label') para o campo atual.
            # 'text=label_text' exibe o texto definido na lista 'campos' para o rótulo.
            # 'font=("Arial", 12)' define a fonte do texto no rótulo.
            # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco.
            # 'anchor="w"' alinha o texto do rótulo à esquerda.
            tk.Label(frame_form,
                     text=label_text,
                     font=("Arial", 12),
                     bg="#FFFFFF",
                     anchor="w"
                     ).grid(row=i,  # Define a linha do grid correspondente ao índice atual do loop.
                            column=0,  # Coloca o rótulo na primeira coluna (coluna 0).
                            sticky="e",  # Alinha o rótulo à direita na célula.
                            padx=5,  # Adiciona 5 pixels de espaço horizontal ao redor do rótulo.
                            pady=5  # Adiciona 5 pixels de espaço vertical ao redor do rótulo.
                            )

            # Cria um campo de entrada ('Entry') associado ao rótulo.
            # 'font=("Arial", 12)' define a fonte para o texto digitado no campo.
            # 'width=40' define a largura do campo em caracteres.
            setattr(

                self,
                var_name,  # Atributo da instância associado ao nome da variável definido na lista 'campos'.
                tk.Entry(frame_form, font=("Arial", 12), width=40)

            )

            # Posiciona o campo de entrada no grid, ao lado do rótulo correspondente.
            getattr(self,
                    var_name  # Recupera o widget 'Entry' associado ao atributo criado acima.
                    ).grid( row=i,  # Define a linha do grid correspondente ao índice atual do loop.
                            column=1,  # Coloca o campo de entrada na segunda coluna (coluna 1).
                            padx=5,  # Adiciona 5 pixels de espaço horizontal ao redor do campo.
                            pady=5)  # Adiciona 5 pixels de espaço vertical ao redor do campo.


        # Cria um frame para organizar os botões relacionados às ações do formulário.
        # 'bg="#FFFFFF"' define a cor de fundo do frame como branco,
        #       para coincidir com o restante do formulário.
        frame_botoes = tk.Frame(frame_form, bg="#FFFFFF")

        # Posiciona o frame no grid do formulário.
        # 'row=len(campos)' coloca o frame logo abaixo dos campos de entrada,
        #       usando o comprimento da lista 'campos' para determinar a linha.
        # 'column=0, columnspan=2' faz com que o frame ocupe ambas as colunas do grid.
        # 'pady=10' adiciona 10 pixels de espaço vertical acima e abaixo do frame.
        frame_botoes.grid(row=len(campos),
                          column=0,
                          columnspan=2,
                          pady=10)

        # Define uma lista de botões, onde cada botão é uma tupla contendo:
        # - O texto exibido no botão.
        # - O comando associado ao botão (uma função que será chamada ao clicar).
        # - A cor de fundo do botão.
        botoes = [
            ("Salvar/Atualizar", self.salvar, "#4CAF50"),  # Botão para salvar ou atualizar um cliente.
            ("Excluir", self.excluir, "#F44336"),  # Botão para excluir o cliente selecionado.
            ("Limpar", self.limpar_form, "#2196F3"),  # Botão para limpar os campos do formulário.
        ]

        # Itera sobre a lista de botões definida anteriormente.
        # Para cada botão, são extraídos:
        # - 'text': O texto a ser exibido no botão.
        # - 'command': A função que será chamada ao clicar no botão.
        # - 'color': A cor de fundo do botão.
        for text, command, color in botoes:

            # Cria um botão dentro do frame_botoes.
            tk.Button(
                frame_botoes,  # O botão será filho do frame_botoes.
                text=text,  # Define o texto exibido no botão.
                font=("Arial", 12, "bold"),  # Usa fonte Arial, tamanho 12, em negrito.
                bg=color,  # Define a cor de fundo do botão, de acordo com a lista de botões.
                fg="#FFFFFF",  # Define a cor do texto do botão como branco.
                activebackground=color,  # Define a cor de fundo quando o botão é clicado.
                activeforeground="#FFFFFF",  # Define a cor do texto quando o botão é clicado como branco.
                command=command,  # Associa o comando (função) que será executado ao clicar no botão.
                width=15  # Define a largura do botão em 15 caracteres.
            ).pack(
                side="left",  # Posiciona os botões lado a lado horizontalmente dentro do frame.
                padx=10  # Adiciona 10 pixels de espaço horizontal entre os botões.
            )

        # Cria um LabelFrame para agrupar e exibir a lista de clientes.
        # 'LabelFrame' é um widget do Tkinter que permite
        #       adicionar um título ao frame.
        frame_lista = tk.LabelFrame(self,  # O frame é filho do widget atual (neste caso, a tela de clientes).
                                    text="Lista de Clientes",  # Define o título do LabelFrame como "Lista de Clientes".
                                    font=("Arial", 12, "bold"),
                                    # Define a fonte do título como Arial, tamanho 12, em negrito.
                                    bg="#FFFFFF",  # Define a cor de fundo do LabelFrame como branco.
                                    fg="#333333",  # Define a cor do texto do título como cinza escuro (#333333).
                                    padx=10,  # Adiciona 10 pixels de preenchimento horizontal interno.
                                    pady=10)  # Adiciona 10 pixels de preenchimento vertical interno.


        # Posiciona o LabelFrame na parte inferior da janela.
        # 'side="bottom"' alinha o frame na parte inferior.
        # 'fill="both"' faz com que o frame preencha tanto a largura
        #       quanto a altura disponíveis.
        # 'expand=True' permite que o frame expanda proporcionalmente
        #       quando a janela é redimensionada.
        # 'padx=20' adiciona 20 pixels de margem horizontal ao redor do frame.
        # 'pady=10' adiciona 10 pixels de margem vertical ao redor do frame.
        frame_lista.pack(side="bottom",
                         fill="both",
                         expand=True,
                         padx=20,
                         pady=10)

        # Define as colunas que serão exibidas na Treeview.
        # Cada item na tupla 'colunas' corresponde a uma coluna na tabela.
        colunas = ("_id", "nome", "cpf", "telefone", "email")

        # Cria e configura o estilo da Treeview (tabela).
        # 'ttk.Style()' é usado para personalizar widgets no Tkinter.
        estilo = ttk.Style()

        # Configura o estilo das linhas da Treeview.
        # 'font=("Arial", 10)' define a fonte das linhas como Arial, tamanho 10.
        # 'rowheight=25' define a altura de cada linha como 25 pixels.
        # 'background="#FFFFFF"' define o fundo das linhas como branco.
        # 'fieldbackground="#FFFFFF"' define o fundo do campo de texto como branco.
        estilo.configure("Treeview",
                        font=("Arial", 10),
                        rowheight=25,
                        background="#FFFFFF",
                        fieldbackground="#FFFFFF")

        # Configura o estilo do cabeçalho das colunas.
        # 'font=("Arial", 12, "bold")' define o texto do cabeçalho
        #       como Arial, tamanho 12, em negrito.
        # 'background="#F0F0F0"' define a cor de fundo do cabeçalho como cinza claro.
        # 'foreground="#333333"' define a cor do texto do cabeçalho como cinza escuro.
        estilo.configure("Treeview.Heading",
                        font=("Arial", 12, "bold"),
                        background="#F0F0F0",
                        foreground="#333333")

        # Configura o comportamento visual quando uma linha é selecionada.
        # 'background=[("selected", "#D3D3D3")]' define a cor de
        #       fundo como cinza claro para a linha selecionada.
        estilo.map("Treeview",
                        background=[("selected", "#D3D3D3")])


        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado

        # Cria o widget Treeview para exibir a tabela de clientes.
        # 'frame_lista' é o pai do Treeview, onde ele será exibido.
        # 'columns=colunas' define as colunas que a tabela terá.
        # 'show="headings"' oculta a coluna padrão e exibe apenas os cabeçalhos definidos.
        # 'style="Treeview"' aplica o estilo personalizado configurado anteriormente.
        self.tree = ttk.Treeview(frame_lista,
                                columns=colunas,
                                show="headings",
                                style="Treeview")

        # Itera sobre cada coluna definida na variável 'colunas'
        #       para configurar os cabeçalhos e colunas.
        for c in colunas:

            # Define o texto do cabeçalho de cada coluna.
            # 'text=c.title()' transforma o nome da coluna em título (primeira letra maiúscula).
            self.tree.heading(c, text=c.title())

            # Configura a largura e o alinhamento de cada coluna.
            # 'anchor="center"' centraliza o texto dentro da coluna.
            # 'width=150' define a largura padrão das colunas, exceto
            #       para a coluna '_id', que tem largura menor (120 pixels).
            self.tree.column(c,
                             anchor="center",
                             width=150 if c != "_id" else 120)

        # Empacota o Treeview para que ele ocupe o espaço disponível no widget pai.
        # 'fill="both"' faz com que o Treeview se expanda tanto horizontal quanto verticalmente.
        # 'expand=True' permite que o Treeview ocupe o espaço extra quando a janela é redimensionada.
        # 'pady=5' adiciona 5 pixels de espaçamento vertical acima e abaixo do Treeview.
        self.tree.pack(fill="both",
                       expand=True,
                       pady=5)

        # Cria uma barra de rolagem vertical para o Treeview.
        # 'frame_lista' é o contêiner onde a barra de rolagem será posicionada.
        # 'orient="vertical"' especifica que a barra será vertical.
        # 'command=self.tree.yview' vincula a barra de rolagem ao movimento vertical do Treeview.
        scrollbar = ttk.Scrollbar(frame_lista,
                                  orient="vertical",
                                  command=self.tree.yview)

        # Configura o Treeview para utilizar a barra de rolagem.
        # 'yscrollcommand=scrollbar.set' sincroniza a barra de rolagem
        #       com o movimento vertical do Treeview.
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do frame que contém o Treeview.
        # 'side="right"' alinha a barra de rolagem à direita.
        # 'fill="y"' faz com que a barra de rolagem preencha todo o
        #       espaço vertical disponível.
        scrollbar.pack(side="right", fill="y")

        # Vincula o evento de seleção de linha no Treeview a um método de callback.
        # '<<TreeviewSelect>>' é o evento disparado ao selecionar uma linha no Treeview.
        # 'self.ao_selecionar_linha' é o método que será chamado quando o evento ocorrer.
        self.tree.bind("<<TreeviewSelect>>", self.ao_selecionar_linha)

        # Chama o método para carregar os dados dos clientes no Treeview.
        # 'self.carregar_clientes()' popula o Treeview com os registros existentes.
        self.carregar_clientes()


    # Define o método para carregar os dados dos clientes no Treeview.
    def carregar_clientes(self):

        # Itera sobre todos os itens do Treeview e os remove.
        # 'self.tree.get_children()' retorna uma lista de todos os itens do Treeview.
        # 'self.tree.delete(item)' remove cada item pelo seu identificador.
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Recupera todos os registros de clientes do banco de dados.
        # 'db.clientes.find()' retorna um cursor contendo
        #       todos os documentos da coleção 'clientes'.
        for cli in db.clientes.find():

            # Insere os dados de cada cliente no Treeview.
            # 'str(cli["_id"])' converte o identificador do cliente (ObjectId) para string.
            # 'cli.get("nome", "")' tenta obter o valor do campo 'nome', ou
            #       retorna uma string vazia se o campo não existir.
            # Outros campos ('cpf', 'telefone', 'email') seguem a mesma lógica.
            self.tree.insert("", tk.END, values=(str(cli["_id"]),
                                                 cli.get("nome", ""),
                                                 cli.get("cpf", ""),
                                                 cli.get("telefone", ""),
                                                 cli.get("email", "")))


    # Define o método chamado quando uma linha do Treeview é selecionada.
    def ao_selecionar_linha(self, event):

        # Obtém a seleção atual do Treeview.
        # 'self.tree.selection()' retorna uma lista de identificadores
        #       das linhas selecionadas.
        selection = self.tree.selection()

        # Verifica se nenhuma linha foi selecionada.
        # Se a lista estiver vazia, simplesmente retorna sem fazer nada.
        if not selection:
            return

        # Obtém os valores da linha selecionada.
        # 'self.tree.item(selection[0])' recupera os dados da primeira linha selecionada.
        # 'item["values"]' retorna os valores de todas as colunas da
        #       linha selecionada como uma tupla.
        item = self.tree.item(selection[0])
        vals = item["values"]

        # Armazena o ID do cliente da linha selecionada em um atributo da classe.
        # Isso será usado posteriormente para identificar qual
        #       cliente está sendo manipulado.
        self.cliente_id_atual = vals[0]

        # Busca o documento completo do cliente no banco de dados.
        # Usa o ID armazenado anteriormente como filtro para localizar o cliente.
        # 'ObjectId(self.cliente_id_atual)' converte a string ID de
        #       volta para o formato ObjectId do MongoDB.
        doc = db.clientes.find_one({"_id": ObjectId(self.cliente_id_atual)})

        # Se o cliente for encontrado no banco de dados:
        if doc:

            # Limpa e preenche o campo de entrada para o nome do cliente.
            self.entry_nome.delete(0, tk.END)  # Remove qualquer texto existente.
            self.entry_nome.insert(0, doc.get("nome", ""))  # Insere o nome do cliente ou uma string vazia.

            # Limpa e preenche o campo de entrada para o CPF do cliente.
            self.entry_cpf.delete(0, tk.END)
            self.entry_cpf.insert(0, doc.get("cpf", ""))

            # Limpa e preenche o campo de entrada para o telefone do cliente.
            self.entry_telefone.delete(0, tk.END)
            self.entry_telefone.insert(0, doc.get("telefone", ""))

            # Limpa e preenche o campo de entrada para o endereço do cliente.
            self.entry_endereco.delete(0, tk.END)
            self.entry_endereco.insert(0, doc.get("endereco", ""))

            # Limpa e preenche o campo de entrada para o e-mail do cliente.
            self.entry_email.delete(0, tk.END)
            self.entry_email.insert(0, doc.get("email", ""))


    # Define o método para salvar ou atualizar os dados do cliente.
    def salvar(self):

        # Obtém o valor do campo de entrada para o nome do cliente.
        # O método 'get()' retorna o texto digitado, enquanto 'strip()'
        #       remove espaços extras no início e no final.
        nome = self.entry_nome.get().strip()

        # Verifica se o campo de nome está vazio.
        # Se estiver, exibe uma mensagem de aviso e encerra a execução da função.
        if not nome:
            messagebox.showwarning("Aviso", "O campo 'Nome' é obrigatório.")
            return

        # Cria um dicionário chamado 'doc' com os dados do cliente.
        # Cada chave representa um campo do cliente, e os valores
        #       são extraídos dos campos de entrada.
        doc = {

            # Armazena o nome do cliente.
            "nome": nome,

            # Armazena o CPF digitado, removendo espaços extras.
            "cpf": self.entry_cpf.get().strip(),

            # Armazena o telefone digitado.
            "telefone": self.entry_telefone.get().strip(),

            # Armazena o endereço digitado.
            "endereco": self.entry_endereco.get().strip(),

            # Armazena o e-mail digitado.
            "email": self.entry_email.get().strip(),

            # Armazena a data e hora do cadastro, utilizando a data atual do sistema.
            "data_cadastro": datetime.datetime.now(),

            # Armazena o ID do médico que realizou a última alteração.
            "ultima_alteracao_por_id": self.medico_logado["_id"],

            # Armazena o nome do médico que realizou a última alteração.
            "ultima_alteracao_por_nome": self.medico_logado["nome"],

            # Armazena a data e hora da última alteração.
            "ultima_alteracao_em": datetime.datetime.now()

        }

        try:

            # Verifica se 'self.cliente_id_atual' é None, ou seja, se
            #       estamos adicionando um novo cliente.
            if self.cliente_id_atual is None:

                # Insere o novo cliente no banco de dados, usando os dados do dicionário 'doc'.
                db.clientes.insert_one(doc)

            else:

                # Atualiza o cliente existente no banco de dados.
                # O cliente é identificado pelo seu '_id', armazenado em 'self.cliente_id_atual'.
                # Apenas os campos no dicionário 'doc' são atualizados.
                db.clientes.update_one({"_id": ObjectId(self.cliente_id_atual)}, {"$set": doc})

            # Exibe uma mensagem de sucesso informando que o cliente foi salvo ou atualizado.
            messagebox.showinfo("Sucesso",
                                "Cliente salvo/atualizado!")

            # Recarrega a lista de clientes para refletir as alterações feitas.
            self.carregar_clientes()

            # Limpa os campos do formulário para preparar para um
            #       novo cadastro ou edição.
            self.limpar_form()

        # Captura qualquer erro que possa ocorrer durante a operação.
        except Exception as e:

            # Exibe uma mensagem de erro detalhando o problema encontrado.
            messagebox.showerror("Erro",
                                 f"Não foi possível salvar:\n{e}")


    def excluir(self):

        # Verifica se há um cliente selecionado para exclusão.
        if not self.cliente_id_atual:

            # Exibe um aviso caso nenhum cliente esteja selecionado.
            messagebox.showwarning("Aviso",
                                   "Selecione um cliente para excluir.")
            return

        # Exibe uma caixa de confirmação perguntando ao usuário se deseja excluir o cliente.
        resp = messagebox.askyesno("Confirmação",
                                   "Deseja excluir este cliente?")

        # Se o usuário confirmar (clicar em 'Sim').
        if resp:

            try:

                # Exclui o cliente do banco de dados, identificado pelo '_id'
                #       armazenado em 'self.cliente_id_atual'.
                db.clientes.delete_one({"_id": ObjectId(self.cliente_id_atual)})

                # Exibe uma mensagem de sucesso indicando que o cliente foi excluído.
                messagebox.showinfo("Sucesso", "Cliente excluído!")

                # Recarrega a lista de clientes para refletir a exclusão feita.
                self.carregar_clientes()

                # Limpa os campos do formulário para preparar para novas operações.
                self.limpar_form()

            # Captura qualquer erro que possa ocorrer durante a exclusão do cliente.
            except Exception as e:

                # Exibe uma mensagem de erro com detalhes sobre o problema.
                messagebox.showerror("Erro",
                                     f"Não foi possível excluir:\n{e}")



    def limpar_form(self):

        # Reseta a variável 'cliente_id_atual' para None, indicando
        #       que nenhum cliente está sendo editado.
        self.cliente_id_atual = None

        # Limpa o campo de entrada do nome do cliente.
        self.entry_nome.delete(0, tk.END)

        # Limpa o campo de entrada do CPF do cliente.
        self.entry_cpf.delete(0, tk.END)

        # Limpa o campo de entrada do telefone do cliente.
        self.entry_telefone.delete(0, tk.END)

        # Limpa o campo de entrada do endereço do cliente.
        self.entry_endereco.delete(0, tk.END)

        # Limpa o campo de entrada do e-mail do cliente.
        self.entry_email.delete(0, tk.END)



# ---------------------------------------------------------------------
# TelaAnimais (CRUD)
# ---------------------------------------------------------------------
class TelaAnimais(tk.Frame):

    # Método inicializador da classe 'TelaAnimais'. Este método configura o
    #       frame principal da tela de gerenciamento de animais.
    def __init__(self, master, medico_logado, *args, **kwargs):

        # Chama o inicializador da classe pai ('tk.Frame') para
        #       aplicar configurações padrão.
        # 'master' é o widget pai ao qual este frame será anexado (normalmente o
        #       notebook ou janela principal).
        # 'medico_logado' é um dicionário que contém os detalhes do
        #       médico atualmente logado no sistema.
        # '*args' é uma lista de argumentos adicionais que podem ser passados à classe pai.
        # '**kwargs' é um dicionário de argumentos nomeados que podem ser
        #       usados para configurar propriedades adicionais.
        super().__init__(master, *args, **kwargs)

        # Armazena as informações do médico logado em um atributo da classe.
        # Este atributo pode ser usado em métodos futuros para personalizar ou
        #       restringir ações com base no usuário logado.
        self.medico_logado = medico_logado

        # Inicializa o atributo 'animal_id_atual' como None.
        # Este atributo será usado para armazenar o ID do animal selecionado,
        #       permitindo operações como edição ou exclusão.
        self.animal_id_atual = None

        # Configura a cor de fundo do frame para '#F7F7F7'.
        # '#F7F7F7' é um tom claro de cinza usado para criar um contraste
        #       suave e dar um visual profissional e limpo à interface.
        self.configure(bg="#F7F7F7")

        # Cria um rótulo de título para a tela de gerenciamento de animais.
        # 'self' é o frame principal ao qual o rótulo será anexado.
        # 'text="Gerenciamento de Animais"' define o texto exibido no rótulo.
        # 'font=("Arial", 18, "bold")' configura a fonte como Arial,
        #       tamanho 18, em negrito, para destaque.
        # 'bg="#F7F7F7"' define a cor de fundo do rótulo, que combina com a cor do frame.
        # 'fg="#333333"' define a cor do texto como um tom escuro de
        #       cinza, criando um bom contraste com o fundo.
        titulo = tk.Label(self,
                          text="Gerenciamento de Animais",
                          font=("Arial", 18, "bold"),
                          bg="#F7F7F7",
                          fg="#333333")

        # Posiciona o rótulo na tela.
        # 'pady=10' adiciona 10 pixels de margem vertical acima e
        #       abaixo do rótulo para espaçamento visual.
        titulo.pack(pady=10)

        # Cria um LabelFrame para agrupar os campos de cadastro de animais.
        # 'self' é o frame principal ao qual o LabelFrame será anexado.
        # 'text="Cadastro de Animais"' define o título do LabelFrame.
        # 'font=("Arial", 12, "bold")' configura a fonte do título do LabelFrame.
        # 'bg="#FFFFFF"' define a cor de fundo do LabelFrame como
        #       branco, destacando-o do fundo geral.
        # 'fg="#333333"' define a cor do texto do título como cinza escuro.
        # 'padx=10, pady=10' adiciona 10 pixels de margem interna (padding) nas
        #       direções horizontal e vertical.
        frame_form = tk.LabelFrame(self,
                                   text="Cadastro de Animais",
                                   font=("Arial", 12, "bold"),
                                   bg="#FFFFFF",
                                   fg="#333333",
                                   padx=10,
                                   pady=10)

        # Posiciona o LabelFrame na parte superior da tela.
        # 'side="top"' alinha o LabelFrame na parte superior do frame.
        # 'fill="x"' faz com que o LabelFrame ocupe toda a largura disponível.
        # 'padx=20' adiciona 20 pixels de margem externa horizontal.
        # 'pady=10' adiciona 10 pixels de margem externa vertical.
        frame_form.pack(side="top",
                        fill="x",
                        padx=20,
                        pady=10)

        # Define uma lista de tuplas para os campos do formulário de cadastro de animais.
        # Cada tupla contém dois elementos:
        # 1. 'label_text': O texto descritivo que será exibido ao lado do
        #       campo de entrada (ex.: "Nome do Animal:").
        # 2. 'var_name': O nome do atributo que será criado dinamicamente
        #       para armazenar a entrada correspondente.
        campos = [

            # O campo para o nome do animal.
            # 'entry_nome_animal' será o nome do atributo criado
            #       para armazenar a entrada deste campo.
            ("Nome do Animal:", "entry_nome_animal"),

            # O campo para a espécie do animal.
            # 'entry_especie' será o nome do atributo criado para
            #       armazenar a entrada deste campo.
            ("Espécie:", "entry_especie"),

            # O campo para a raça do animal.
            # 'entry_raca' será o nome do atributo criado para
            #       armazenar a entrada deste campo.
            ("Raça:", "entry_raca"),

            # O campo para a idade do animal.
            # 'entry_idade' será o nome do atributo criado para
            #       armazenar a entrada deste campo.
            ("Idade:", "entry_idade"),

            # O campo para o peso do animal em quilogramas.
            # 'entry_peso' será o nome do atributo criado para
            #       armazenar a entrada deste campo.
            ("Peso (kg):", "entry_peso")

        ]

        # Itera sobre a lista de campos para criar dinamicamente os
        #       rótulos e os campos de entrada no formulário.
        # 'i' é o índice da iteração (usado para posicionar os elementos na grade).
        # '(label, attr_name)' descompacta cada tupla da lista 'campos', onde:
        #   'label' é o texto do rótulo (ex.: "Nome do Animal:").
        #   'attr_name' é o nome do atributo que será criado dinamicamente.
        for i, (label, attr_name) in enumerate(campos):

            # Cria um rótulo para o campo no formulário.
            # 'text=label' define o texto exibido no rótulo.
            # 'font=("Arial", 12)' define a fonte e o tamanho do texto.
            # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco.
            # 'grid(row=i, column=0)' posiciona o rótulo na grade do formulário, com:
            #    - 'row=i': na linha correspondente ao índice atual.
            #    - 'column=0': na primeira coluna.
            # 'sticky="e"' alinha o rótulo à direita dentro da célula.
            # 'padx=5' e 'pady=5' adicionam margens horizontal e vertical para espaçamento.
            tk.Label(frame_form,
                     text=label,
                     font=("Arial", 12),
                     bg="#FFFFFF").grid(row=i,
                                        column=0,
                                        sticky="e",
                                        padx=5,
                                        pady=5)

            # Cria um campo de entrada associado ao rótulo.
            # 'font=("Arial", 12)' define a fonte e o tamanho do
            #       texto digitado no campo.
            # 'width=40' define a largura do campo em caracteres.
            setattr(self, attr_name, tk.Entry(frame_form,
                                              font=("Arial", 12),
                                              width=40))

            # Posiciona o campo de entrada na grade do formulário.
            # 'grid(row=i, column=1)' coloca o campo na linha do índice
            #       atual (mesma linha do rótulo) e na segunda coluna.
            # 'padx=5' e 'pady=5' adicionam margens horizontal e vertical para espaçamento.
            getattr(self, attr_name).grid(row=i,
                                          column=1,
                                          padx=5,
                                          pady=5)

        # Cria um rótulo para o campo "Sexo (M/F):".
        # 'text="Sexo (M/F):"' define o texto exibido no rótulo.
        # 'font=("Arial", 12)' define a fonte do texto como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco para
        #       combinar com o frame.
        # 'grid(row=len(campos), column=0)' posiciona o rótulo na próxima
        #       linha (após os campos definidos na lista 'campos') e na primeira coluna.
        # 'sticky="e"' alinha o rótulo à direita dentro da célula.
        # 'padx=5' e 'pady=5' adicionam espaçamento horizontal e vertical.
        tk.Label(frame_form,
                 text="Sexo (M/F):",
                 font=("Arial", 12),
                 bg="#FFFFFF").grid(row=len(campos),
                                    column=0,
                                    sticky="e",
                                    padx=5,
                                    pady=5)

        # Cria um ComboBox (menu suspenso) para o campo "Sexo", permitindo
        #       selecionar entre "M" (Masculino) e "F" (Feminino).
        # 'values=["M", "F"]' define as opções disponíveis no menu suspenso.
        # 'font=("Arial", 12)' define a fonte do texto exibido no ComboBox.
        # 'width=38' ajusta a largura do ComboBox para se alinhar com os campos de entrada.
        self.combo_sexo = ttk.Combobox(frame_form,
                                       values=["M", "F"],
                                       font=("Arial", 12),
                                       width=38)

        # Posiciona o ComboBox na próxima linha (mesma linha do rótulo) e na segunda coluna.
        # 'grid(row=len(campos), column=1)' coloca o ComboBox na próxima
        #       linha calculada com base no comprimento da lista 'campos'.
        # 'padx=5' e 'pady=5' adicionam espaçamento horizontal e vertical.
        self.combo_sexo.grid(row=len(campos),
                             column=1,
                             padx=5,
                             pady=5)

        # Define a seleção inicial do ComboBox para a primeira
        #       opção (índice 0), que é "M".
        self.combo_sexo.current(0)

        # Cria um rótulo para o campo "Dono (Cliente):".
        # 'text="Dono (Cliente):"' define o texto exibido no rótulo.
        # 'font=("Arial", 12)' define a fonte como Arial com tamanho 12.
        # 'bg="#FFFFFF"' define o fundo do rótulo como branco, harmonizando
        #       com o restante do frame.
        # 'grid(row=len(campos)+1, column=0)' posiciona o rótulo na linha
        #       seguinte ao campo "Sexo (M/F)" e na primeira coluna.
        # 'sticky="e"' alinha o rótulo à direita dentro da célula.
        # 'padx=5' e 'pady=5' adicionam margens horizontais e verticais para espaçamento.
        tk.Label(frame_form,
                 text="Dono (Cliente):",
                 font=("Arial", 12),
                 bg="#FFFFFF").grid(row=len(campos) + 1,
                                    column=0,
                                    sticky="e",
                                    padx=5,
                                    pady=5)

        # Cria um ComboBox (menu suspenso) para selecionar o dono (cliente) associado ao animal.
        # 'font=("Arial", 12)' define a fonte como Arial com tamanho 12.
        # 'width=38' ajusta a largura do ComboBox para alinhar com outros campos de entrada.
        self.combo_dono = ttk.Combobox(frame_form,
                                       font=("Arial", 12),
                                       width=38)

        # Posiciona o ComboBox na linha seguinte ao campo "Sexo (M/F)" e na segunda coluna.
        # 'grid(row=len(campos)+1, column=1)' calcula a posição com base no
        #       comprimento da lista 'campos' e adiciona uma linha extra.
        # 'padx=5' e 'pady=5' adicionam margens horizontais e verticais para espaçamento.
        self.combo_dono.grid(row=len(campos) + 1,
                             column=1,
                             padx=5,
                             pady=5)

        # Cria um frame para os botões de ação dentro do formulário de cadastro de animais.
        # 'bg="#FFFFFF"' define o fundo do frame como branco para
        #       combinar com o restante do formulário.
        frame_botoes = tk.Frame(frame_form, bg="#FFFFFF")

        # Posiciona o frame na grade do formulário.
        # 'row=len(campos)+2' posiciona o frame na linha duas abaixo do
        #       último campo (campos + "Sexo" e "Dono").
        # 'column=0' e 'columnspan=2' fazem o frame ocupar as duas colunas do formulário.
        # 'pady=10' adiciona 10 pixels de margem vertical
        #       abaixo do frame para espaçamento.
        frame_botoes.grid(row=len(campos) + 2,
                          column=0,
                          columnspan=2,
                          pady=10)

        # Lista de botões, cada um representado por uma tupla:
        # O primeiro item da tupla é o texto exibido no botão.
        # O segundo item é o método que será chamado quando o botão for clicado.
        # O terceiro item é a cor de fundo do botão, usada para indicar sua função.
        botoes = [

            # Botão verde para salvar ou atualizar o cadastro.
            ("Salvar/Atualizar", self.salvar, "#4CAF50"),
            ("Excluir", self.excluir, "#F44336"),  # Botão vermelho para excluir o cadastro.
            ("Limpar", self.limpar_form, "#2196F3"),  # Botão azul para limpar o formulário.
            ("Histórico", self.ver_historico, "#9C27B0"),
            # Botão roxo para acessar o histórico de atendimentos do animal.
        ]

        # Itera sobre a lista de botões definida anteriormente.
        # Cada botão é representado por uma tupla com o texto exibido, o
        #       comando associado e a cor de fundo.
        for text, command, color in botoes:

            # Cria um botão dentro do frame de botões.
            # 'text=text' define o texto exibido no botão, como "Salvar/Atualizar".
            # 'font=("Arial", 12, "bold")' define a fonte do texto, tamanho 12 e em negrito.
            # 'bg=color' define a cor de fundo do botão, conforme especificado na lista de botões.
            # 'fg="#FFFFFF"' define a cor do texto como branco para contraste com o fundo.
            # 'activebackground=color' mantém a mesma cor de fundo ao clicar no botão.
            # 'activeforeground="#FFFFFF"' mantém a cor do texto como branco ao clicar no botão.
            # 'command=command' associa o botão ao método correspondente, como `self.salvar`.
            # 'width=15' define uma largura fixa para os botões, garantindo consistência visual.
            tk.Button( frame_botoes,
                        text=text,
                        font=("Arial", 12, "bold"),
                        bg=color,
                        fg="#FFFFFF",
                        activebackground=color,
                        activeforeground="#FFFFFF",
                        command=command,
                        width=15).pack(

                                    # Posiciona o botão dentro do frame de botões.
                                    # 'side="left"' alinha os botões horizontalmente à esquerda.
                                    # 'padx=10' adiciona 10 pixels de margem horizontal entre os botões.
                                    side="left",
                                    padx=10)

        # Cria um LabelFrame para a seção "Lista de Animais".
        # 'self' indica que o frame pertence à classe atual.
        # 'text="Lista de Animais"' define o título exibido na borda superior do frame.
        # 'font=("Arial", 12, "bold")' configura o texto do título
        #       com fonte Arial, tamanho 12, em negrito.
        # 'bg="#FFFFFF"' define a cor de fundo do frame como branco.
        # 'fg="#333333"' define a cor do texto do título como cinza escuro.
        # 'padx=10' e 'pady=10' adicionam 10 pixels de preenchimento interno
        #       horizontal e vertical dentro do frame.
        frame_lista = tk.LabelFrame(self,
                                    text="Lista de Animais",
                                    font=("Arial", 12, "bold"),
                                    bg="#FFFFFF",
                                    fg="#333333",
                                    padx=10,
                                    pady=10)

        # Posiciona o frame na parte inferior da janela.
        # 'side="bottom"' alinha o frame na parte inferior.
        # 'fill="both"' permite que o frame se expanda para preencher toda a
        #       largura e altura disponíveis.
        # 'expand=True' permite que o frame ajuste seu tamanho proporcionalmente
        #       quando a janela for redimensionada.
        # 'padx=20' adiciona 20 pixels de margem horizontal externa ao frame.
        # 'pady=10' adiciona 10 pixels de margem vertical externa ao frame.
        frame_lista.pack(side="bottom",
                        fill="both",
                        expand=True,
                        padx=20,
                        pady=10)

        # Define as colunas para a Treeview, representando os dados dos animais.
        # Cada item na tupla 'colunas' corresponde a um atributo que será exibido na tabela.
        colunas = (

            "_id",  # Identificador único do animal (usado internamente).
            "nome_animal",  # Nome do animal.
            "especie",  # Espécie do animal (ex.: cão, gato).
            "raca",  # Raça do animal.
            "idade",  # Idade do animal.
            "sexo",  # Sexo do animal (M/F).
            "peso",  # Peso do animal.
            "dono_nome"  # Nome do dono do animal.

        )

        # Cria e configura o estilo da Treeview para personalizar a aparência da tabela.
        estilo = ttk.Style()

        # Configura o estilo geral da Treeview:
        # 'font=("Arial", 10)' define a fonte dos dados com tamanho 10.
        # 'rowheight=25' define a altura de cada linha na tabela.
        # 'background="#FFFFFF"' e 'fieldbackground="#FFFFFF"' definem a
        #       cor de fundo como branco.
        estilo.configure("Treeview",
                         font=("Arial", 10),
                         rowheight=25,
                         background="#FFFFFF",
                         fieldbackground="#FFFFFF")

        # Configura o estilo do cabeçalho da Treeview:
        # 'font=("Arial", 12, "bold")' define a fonte do cabeçalho com tamanho 12 em negrito.
        # 'background="#F0F0F0"' define a cor de fundo do cabeçalho como cinza claro.
        # 'foreground="#333333"' define a cor do texto do cabeçalho como cinza escuro.
        estilo.configure("Treeview.Heading",
                         font=("Arial", 12, "bold"),
                         background="#F0F0F0",
                         foreground="#333333")

        # Configura a aparência da linha selecionada na Treeview:
        # 'background=[("selected", "#D3D3D3")]' define a cor de fundo como
        #       cinza claro quando uma linha é selecionada.
        estilo.map("Treeview",
                   background=[("selected", "#D3D3D3")])

        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria a Treeview para exibir a lista de animais.
        # 'frame_lista' é o frame onde a Treeview será posicionada.
        # 'columns=colunas' define as colunas que a tabela terá, conforme
        #       especificado anteriormente.
        # 'show="headings"' oculta a coluna de índice padrão e exibe apenas os
        #       cabeçalhos das colunas definidas.
        # 'style="Treeview"' aplica o estilo configurado anteriormente.
        self.tree = ttk.Treeview(frame_lista,
                                 columns=colunas,
                                 show="headings",
                                 style="Treeview")

        # Configura os cabeçalhos e colunas da Treeview.
        for c in colunas:

            # Define o texto do cabeçalho para cada coluna:
            # 'c.replace("_", " ").title()' converte o nome da coluna em texto legível.
            # Exemplo: "nome_animal" se torna "Nome Animal".
            self.tree.heading(c, text=c.replace("_", " ").title())

            # Configura as propriedades de cada coluna:
            # 'anchor="center"' centraliza o texto na coluna.
            # 'width=150' define a largura padrão para a maioria das colunas.
            # 'width=120' é usado especificamente para a coluna '_id', pois requer menos espaço.
            self.tree.column(c, anchor="center", width=150 if c != "_id" else 120)

        # Posiciona a Treeview no frame.
        # 'fill="both"' faz com que a tabela preencha toda a largura e altura disponíveis.
        # 'expand=True' permite que a tabela seja redimensionada junto com o frame.
        # 'pady=5' adiciona 5 pixels de margem vertical acima e abaixo da tabela.
        self.tree.pack(fill="both", expand=True, pady=5)

        # Cria uma barra de rolagem vertical associada à Treeview.
        # 'frame_lista' é o frame onde a barra de rolagem será posicionada.
        # 'orient="vertical"' define a orientação da barra de rolagem como vertical.
        # 'command=self.tree.yview' vincula a barra de rolagem à
        #       visualização vertical da Treeview.
        scrollbar = ttk.Scrollbar(frame_lista,
                                  orient="vertical",
                                  command=self.tree.yview)

        # Configura a Treeview para que a barra de rolagem seja atualizada
        # automaticamente conforme o conteúdo da tabela é visualizado.
        # 'yscrollcommand=scrollbar.set' assegura que a barra de rolagem
        # se sincronize com a visualização vertical da Treeview.
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do frame.
        # 'side="right"' alinha a barra no lado direito do frame.
        # 'fill="y"' faz com que a barra preencha toda a altura do frame.
        scrollbar.pack(side="right", fill="y")

        # Vincula um evento à seleção de uma linha na Treeview.
        # '<<TreeviewSelect>>' é o evento gerado quando o usuário seleciona uma linha.
        # 'self.ao_selecionar_linha' é o método que será chamado ao selecionar uma linha.
        self.tree.bind("<<TreeviewSelect>>", self.ao_selecionar_linha)

        # Vincula um evento para detectar quando a aba onde este frame está se torna visível.
        # '<Visibility>' é o evento gerado quando o frame se torna visível.
        # 'self.on_tab_visible' é o método que será chamado ao tornar a aba visível.
        self.bind("<Visibility>", self.on_tab_visible)


    # Método chamado automaticamente quando a aba contendo este
    #       frame se torna visível.
    def on_tab_visible(self, event):

        # Atualiza a lista de donos no combo box.
        # O método '_carregar_donos()' é responsável por buscar os dados
        #       dos donos de animais no banco de dados
        #       e preencher o combo box com essas informações.
        self._carregar_donos()

        # Carrega a lista de animais na tabela (Treeview).
        # O método 'carregar_animais()' busca os registros de animais no banco de dados
        #       e os exibe na tabela, garantindo que os dados estejam atualizados.
        self.carregar_animais()

    # Método privado para carregar os dados dos donos dos animais.
    def _carregar_donos(self):

        # Inicializa uma lista vazia para armazenar os dados dos donos.
        self.lista_donos = []

        # Realiza uma consulta no banco de dados 'clientes' para obter todos os registros.
        for cli in db.clientes.find():

            # Adiciona cada cliente na lista no formato (ID, Nome).
            # Se o cliente não tiver nome registrado, atribui "Sem Nome" como valor padrão.
            self.lista_donos.append((cli["_id"], cli.get("nome", "Sem Nome")))

        # Cria uma lista apenas com os nomes dos donos para exibir no combobox.
        nomes_donos = [n for (_id, n) in self.lista_donos]

        # Define os valores do combobox com os nomes dos donos carregados.
        self.combo_dono['values'] = nomes_donos

        # Define a seleção inicial do combo box como o primeiro item
        #       da lista, caso existam donos.
        if nomes_donos:
            self.combo_dono.current(0)

    # Método para carregar os dados dos animais no Treeview.
    def carregar_animais(self):

        # Remove todos os itens atualmente exibidos no Treeview.
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Realiza uma consulta no banco de dados 'animais' para obter todos os registros.
        for ani in db.animais.find():

            # Obtém o ID do dono do animal.
            dono_id = ani.get("id_dono")

            # Inicializa o nome do dono como uma string vazia.
            dono_nome = ""

            # Verifica se há um ID de dono associado ao animal.
            if dono_id:

                # Realiza uma busca no banco de dados 'clientes' para
                #       encontrar o dono correspondente.
                dono_doc = db.clientes.find_one({"_id": dono_id})

                # Caso o dono seja encontrado, extrai o nome. Caso contrário, o nome permanece vazio.
                dono_nome = dono_doc.get("nome", "") if dono_doc else ""

            # Insere os dados do animal no Treeview.
            self.tree.insert("",  # Insere o item na raiz do Treeview, sem um pai específico.
                             tk.END,  # Adiciona o item ao final da lista de itens existentes.
                             values=(
                                 str(ani["_id"]),  # Converte o ID do animal para uma string e exibe no Treeview.
                                 ani.get("nome_animal", ""),
                                 # Obtém o nome do animal ou exibe uma string vazia se não existir.
                                 ani.get("especie", ""),  # Obtém a espécie do animal ou exibe uma string vazia.
                                 ani.get("raca", ""),  # Obtém a raça do animal ou exibe uma string vazia.
                                 ani.get("idade", ""),  # Obtém a idade do animal ou exibe uma string vazia.
                                 ani.get("sexo", ""),  # Obtém o sexo do animal ou exibe uma string vazia.
                                 ani.get("peso", ""),  # Obtém o peso do animal ou exibe uma string vazia.
                                 dono_nome  # Exibe o nome do dono associado ao animal.
                             ))


    # Método para manipular o evento de seleção de linha no Treeview.
    def ao_selecionar_linha(self, event):

        # Obtém os itens selecionados no Treeview.
        selection = self.tree.selection()

        if not selection:

            # Se nenhuma linha estiver selecionada, interrompe o processo.
            return

        # Obtém os dados da linha selecionada no Treeview.
        item = self.tree.item(selection[0])  # Pega o primeiro item selecionado.
        vals = item["values"]  # Extrai os valores da linha selecionada.

        # Atualiza o ID do animal atual com o ID selecionado.
        self.animal_id_atual = vals[0]

        # Preenche o campo de entrada para o nome do animal.
        self.entry_nome_animal.delete(0, tk.END)  # Limpa o conteúdo atual do campo.
        self.entry_nome_animal.insert(0, vals[1])  # Insere o valor do nome do animal.

        # Preenche o campo de entrada para a espécie do animal.
        self.entry_especie.delete(0, tk.END)  # Limpa o conteúdo atual do campo.
        self.entry_especie.insert(0, vals[2])  # Insere o valor da espécie.

        # Preenche o campo de entrada para a raça do animal.
        self.entry_raca.delete(0, tk.END)  # Limpa o conteúdo atual do campo.
        self.entry_raca.insert(0, vals[3])  # Insere o valor da raça.

        # Preenche o campo de entrada para a idade do animal.
        self.entry_idade.delete(0, tk.END)  # Limpa o conteúdo atual do campo.
        self.entry_idade.insert(0, vals[4])  # Insere o valor da idade.

        # Define o valor do ComboBox para o sexo do animal.
        self.combo_sexo.set(vals[5])  # Define o valor do ComboBox como o sexo do animal.

        # Preenche o campo de entrada para o peso do animal.
        self.entry_peso.delete(0, tk.END)  # Limpa o conteúdo atual do campo.
        self.entry_peso.insert(0, vals[6])  # Insere o valor do peso.

        # Define o valor do ComboBox para o dono do animal.
        self.combo_dono.set(vals[7])  # Define o valor do ComboBox como o nome do dono.


    # Método para salvar ou atualizar os dados de um animal.
    def salvar(self):

        # Obtém o valor do campo de entrada 'Nome do Animal', removendo
        #       espaços desnecessários.
        nome_animal = self.entry_nome_animal.get().strip()

        if not nome_animal:

            # Se o campo estiver vazio, exibe um aviso ao usuário.
            messagebox.showwarning("Aviso", "Nome do Animal é obrigatório.")
            return  # Interrompe o processo de salvamento.

        # Obtém o valor do campo de entrada 'Espécie', removendo espaços desnecessários.
        especie = self.entry_especie.get().strip()

        # Obtém o valor do campo de entrada 'Raça', removendo espaços desnecessários.
        raca = self.entry_raca.get().strip()

        # Obtém o valor do campo de entrada 'Idade' como texto,
        #       removendo espaços desnecessários.
        idade_txt = self.entry_idade.get().strip()

        # Obtém o valor selecionado no ComboBox 'Sexo' e remove espaços desnecessários.
        sexo = self.combo_sexo.get().strip()

        # Obtém o valor do campo de entrada 'Peso' como texto,
        #       removendo espaços desnecessários.
        peso_txt = self.entry_peso.get().strip()

        # Obtém o valor selecionado no ComboBox 'Dono'.
        dono_nome = self.combo_dono.get()

        # Tenta converter o texto do campo 'Idade' para um número inteiro.
        try:

            # Verifica se o campo 'idade_txt' possui algum valor.
            # Se sim, converte o valor para inteiro; caso contrário,
            #       atribui 'None' (sem valor).
            idade = int(idade_txt) if idade_txt else None

        except:

            # Caso ocorra um erro durante a conversão (ex.: texto não é um número válido),
            # define 'idade' como 'None'.
            idade = None

        # Tenta converter o texto do campo 'Peso' para um número
        #       de ponto flutuante (float).
        try:

            # Verifica se o campo 'peso_txt' possui algum valor.
            # Se sim, converte o valor para float; caso contrário, atribui 'None' (sem valor).
            peso = float(peso_txt) if peso_txt else None

        except:

            # Caso ocorra um erro durante a conversão (ex.: texto não é um número válido),
            # define 'peso' como 'None'.
            peso = None

        # Inicializa a variável 'dono_id' com 'None'.
        # Esta variável será usada posteriormente para armazenar o ID do dono do animal,
        # que será obtido a partir da seleção do combo box de donos.
        dono_id = None

        # Percorre a lista de donos ('self.lista_donos'), que contém pares de ID e nome.
        for (_id, nm) in self.lista_donos:

            # Verifica se o nome do dono selecionado no combo box ('dono_nome')
            #       corresponde a algum nome na lista de donos.
            if nm == dono_nome:

                # Se encontrar uma correspondência, armazena o ID do dono
                #       correspondente em 'dono_id'.
                dono_id = _id

                # Encerra o loop, pois o ID já foi encontrado.
                break

        # Cria um dicionário chamado 'doc' para representar os dados do animal.
        doc = {

            # Adiciona o nome do animal.
            "nome_animal": nome_animal,

            # Adiciona a espécie do animal.
            "especie": especie,

            # Adiciona a raça do animal.
            "raca": raca,

            # Adiciona a idade do animal. Se for inválida, estará como 'None'.
            "idade": idade,

            # Adiciona o sexo do animal, selecionado no combo box.
            "sexo": sexo,

            # Adiciona o peso do animal. Se for inválido, estará como 'None'.
            "peso": peso,

            # Adiciona o ID do dono do animal, obtido anteriormente.
            #       Se não houver dono, será 'None'.
            "id_dono": dono_id,

            # Cria uma lista vazia para armazenar os históricos do animal.
            "historicos": [],

            # Adiciona o ID do médico que está realizando a alteração.
            "ultima_alteracao_por_id": self.medico_logado["_id"],

            # Adiciona o nome do médico que está realizando a alteração.
            "ultima_alteracao_por_nome": self.medico_logado["nome"],

            # Adiciona a data e hora da última alteração, usando o horário atual.
            "ultima_alteracao_em": datetime.datetime.now()

        }


        # Inicia um bloco try-except para tratar possíveis erros durante a
        #       operação de salvar/atualizar o animal.
        try:

            # Verifica se há um registro selecionado no sistema.
            # O atributo `self.animal_id_atual` representa o ID do registro atualmente em foco.
            # Se ele for diferente de `None`, significa que existe um registro ativo que pode ser alterado.
            if self.animal_id_atual is not None:

                # Cria uma nova janela modal do tipo `Toplevel`, que será usada para
                #       exibir as opções de ação ao usuário.
                # A janela modal será uma "filha" da janela principal (`self.master`), ou
                #       seja, estará hierarquicamente subordinada a ela.
                janela = Toplevel(self.master)

                # Define o título da nova janela modal.
                # Este título aparecerá na barra superior da janela, informando o objetivo dela.
                janela.title("Ação Necessária")

                # Define as dimensões da nova janela modal.
                # Aqui, ela terá 400 pixels de largura e 200 pixels de altura, garantindo um
                #       tamanho adequado para o conteúdo.
                janela.geometry("400x200")

                # Centraliza a janela na tela.
                # A função `centralizar_janela` calcula a posição exata para que a janela seja
                #       exibida no centro da tela.
                # É necessário passar a referência da janela e suas dimensões (largura e altura) como parâmetros.
                centralizar_janela(janela, 400, 200)

                # Configura a nova janela modal como "transiente" da janela principal.
                # Isso significa que a janela modal será vinculada à janela principal,
                #       garantindo que a modal seja exibida sempre no topo
                #       e que ela seja fechada automaticamente caso a janela principal seja encerrada.
                janela.transient(self.master)

                # Torna a janela modal exclusiva (bloqueia a interação com a janela principal
                #       enquanto a modal estiver aberta).
                # Isso impede que o usuário interaja com a janela principal até que uma decisão
                #       seja tomada e a modal seja fechada.
                janela.grab_set()

                # Cria um rótulo (texto) dentro da janela modal para exibir informações ao usuário.
                # O rótulo serve para orientar o usuário sobre o que ele precisa fazer,
                #       neste caso, decidir entre alterar o registro ou cadastrar um novo.
                label = Label(

                    janela,  # Define que o rótulo será exibido dentro da janela modal recém-criada.
                    text="Um registro já está selecionado.\nEscolha uma das opções abaixo:",

                    # Texto que será exibido no rótulo.
                    font=("Arial", 12),  # Define a fonte e o tamanho do texto do rótulo (fonte Arial, tamanho 12).
                    wraplength=350,
                    # Configura o comprimento máximo para o texto antes de ser quebrado em uma nova linha (350px de largura).
                    justify="center"  # Centraliza o texto dentro do rótulo para melhorar a legibilidade e estética.

                )

                # Posiciona o rótulo dentro da janela modal.
                # A função `pack` é usada para gerenciar o layout do rótulo, garantindo
                #       que ele seja exibido corretamente.
                # O parâmetro `pady=20` adiciona um espaçamento vertical (20 pixels) acima e
                #       abaixo do rótulo, deixando o conteúdo mais organizado.
                label.pack(pady=20)

                # Função para tratar a escolha de Alterar o registro
                # Esta função será chamada quando o usuário clicar no botão "Alterar Registro"
                def alterar():

                    try:

                        # Atualiza o registro existente no banco de dados com os novos valores fornecidos.
                        # O método `update_one` é usado para encontrar o registro pelo ID (`self.animal_id_atual`)
                        #       e substituir os campos existentes pelos novos valores contidos no dicionário `doc`.
                        db.animais.update_one(

                            # Filtro para localizar o registro no banco de dados, utilizando o ID do animal selecionado.
                            {"_id": ObjectId(self.animal_id_atual)},

                            # Comando para substituir os campos no registro existente com os valores de `doc`.
                            {"$set": doc}

                        )

                        # Exibe uma mensagem de sucesso ao usuário informando que o registro foi atualizado.
                        messagebox.showinfo("Sucesso", "Animal atualizado com sucesso!")

                        # Atualiza a lista de animais exibida na interface (TreeView ou tabela).
                        # Este método é responsável por recarregar os dados do banco e exibi-los novamente.
                        self.carregar_animais()

                        # Limpa os campos do formulário para evitar dados residuais.
                        # Isso é útil para preparar o formulário para um novo cadastro ou edição.
                        self.limpar_form()

                    # Caso ocorra algum erro durante a atualização do registro no banco de dados,
                    #       o erro será capturado e exibido ao usuário em uma mensagem de erro.
                    except Exception as e:

                        # Exibe uma mensagem de erro detalhando o problema que ocorreu.
                        messagebox.showerror("Erro", f"Erro ao atualizar o registro:\n{e}")

                    # Finalmente, independente do sucesso ou falha, a janela modal será fechada.
                    # Isso garante que o fluxo da aplicação continue corretamente após a ação.
                    finally:
                        janela.destroy()

                # Função para cadastrar um novo registro no banco de dados
                # Esta função será chamada quando o usuário clicar no botão "Cadastrar Novo" na janela de opções.
                def cadastrar_novo():

                    try:

                        # Reseta o ID atual para garantir que a operação de inserção seja tratada
                        #       como um novo cadastro.
                        # O atributo `self.animal_id_atual` representa o registro selecionado na interface.
                        # Aqui, ele é redefinido como `None` para indicar que nenhum registro está em edição.
                        self.animal_id_atual = None

                        # Insere um novo registro no banco de dados.
                        # O dicionário `doc` contém os dados do novo animal que devem ser armazenados no banco.
                        # O método `insert_one` do MongoDB é utilizado para adicionar o novo registro.
                        db.animais.insert_one(doc)

                        # Exibe uma mensagem de sucesso ao usuário, confirmando que o cadastro foi realizado.
                        messagebox.showinfo("Sucesso", "Novo animal cadastrado com sucesso!")

                        # Atualiza a exibição da lista de animais na interface gráfica.
                        # Este método recarrega os dados do banco de dados e atualiza os
                        #       elementos visuais (como uma TreeView ou tabela).
                        self.carregar_animais()

                        # Limpa os campos do formulário na interface.
                        # Após a conclusão do cadastro, o formulário é limpo para evitar dados residuais.
                        self.limpar_form()

                    # Captura e trata qualquer exceção que possa ocorrer durante a execução do
                    #       processo de cadastro.
                    except Exception as e:

                        # Exibe uma mensagem de erro detalhando o problema que ocorreu durante o cadastro.
                        messagebox.showerror("Erro", f"Erro ao cadastrar novo registro:\n{e}")

                    # O bloco `finally` é executado independentemente de sucesso ou falha no processo.
                    # Fecha a janela modal para que o usuário possa continuar interagindo com a
                    #       interface principal.
                    finally:
                        janela.destroy()

                # Criação do botão "Alterar Registro"
                # Este botão será exibido na janela modal e permitirá ao usuário alterar o
                #       registro atualmente selecionado.
                btn_alterar = Button(
                    janela,  # Define que o botão será exibido na janela modal (`janela`).
                    text="Alterar Registro",  # Define o texto do botão como "Alterar Registro".
                    font=("Arial", 10),  # Define a fonte do texto do botão como Arial, tamanho 10.
                    bg="blue",  # Define a cor de fundo do botão como azul.
                    fg="white",  # Define a cor do texto no botão como branco.
                    command=alterar
                    # Define a função que será chamada quando o botão for clicado (neste caso, `alterar`).
                )

                # Adiciona o botão "Alterar Registro" à janela modal.
                # A função `pack` posiciona o botão automaticamente dentro da janela, centralizando-o.
                # O parâmetro `pady=10` adiciona um espaçamento vertical de 10 pixels acima e abaixo do botão.
                btn_alterar.pack(pady=10)

                # Criação do botão "Cadastrar Novo"
                # Este botão será exibido na janela modal e permitirá ao usuário cadastrar um novo registro.
                btn_cadastrar = Button(
                    janela,  # Define que o botão será exibido na janela modal (`janela`).
                    text="Cadastrar Novo",  # Define o texto do botão como "Cadastrar Novo".
                    font=("Arial", 10),  # Define a fonte do texto do botão como Arial, tamanho 10.
                    bg="green",  # Define a cor de fundo do botão como verde.
                    fg="white",  # Define a cor do texto no botão como branco.
                    command=cadastrar_novo
                    # Define a função que será chamada quando o botão for clicado (neste caso, `cadastrar_novo`).
                )

                # Adiciona o botão "Cadastrar Novo" à janela modal.
                # A função `pack` posiciona o botão automaticamente dentro da janela, centralizando-o.
                # O parâmetro `pady=10` adiciona um espaçamento vertical de 10 pixels acima e abaixo do botão.
                btn_cadastrar.pack(pady=10)


            # Caso nenhum registro esteja selecionado (self.animal_id_atual é None), o
            #       sistema interpretará que o usuário deseja realizar um novo cadastro.
            else:  # Novo cadastro

                # Insere um novo registro no banco de dados.
                # O método `insert_one` do MongoDB é utilizado para adicionar um novo documento à coleção 'animais'.
                # O dicionário `doc` contém os dados do novo animal a ser cadastrado.
                db.animais.insert_one(doc)

                # Exibe uma mensagem ao usuário confirmando que o cadastro foi realizado com sucesso.
                # Isso fornece feedback imediato e positivo sobre a operação realizada.
                messagebox.showinfo("Sucesso", "Animal cadastrado com sucesso!")

                # Atualiza a lista de animais na interface.
                # O método `self.carregar_animais()` recarrega os dados diretamente do banco de dados
                #       e os exibe na tabela ou TreeView da interface gráfica.
                self.carregar_animais()

                # Limpa os campos do formulário após o cadastro.
                # O método `self.limpar_form()` remove qualquer dado preenchido nos campos de entrada
                #       para preparar o formulário para uma nova operação ou cadastro.
                self.limpar_form()


        # Bloco de captura de exceções.
        # Este bloco será executado se qualquer erro ou exceção ocorrer durante a
        #       execução do código no bloco `try`.
        except Exception as e:

            # Exibe uma mensagem de erro ao usuário utilizando o `messagebox.showerror`.
            # O título da mensagem será "Erro", e o conteúdo detalhará que o registro não pôde ser salvo.
            # A variável `e` contém informações sobre o erro específico que ocorreu, o que ajuda no diagnóstico.
            messagebox.showerror("Erro", f"Não foi possível salvar:\n{e}")




    # Define o método 'excluir' para remover um animal selecionado do banco de dados.
    def excluir(self):

        # Verifica se há um animal selecionado para exclusão.
        # Se 'self.animal_id_atual' for None ou vazio, exibe um aviso ao usuário.
        if not self.animal_id_atual:

            # Mostra uma mensagem de alerta informando que nenhum animal foi selecionado.
            messagebox.showwarning("Aviso", "Selecione um animal para excluir.")

            # Interrompe a execução do método.
            return

        # Pergunta ao usuário se ele realmente deseja excluir o animal.
        # Exibe uma janela de confirmação com as opções 'Sim' ou 'Não'.
        resp = messagebox.askyesno("Confirmação", "Deseja excluir este animal?")

        # Verifica se o usuário confirmou a exclusão.
        if resp:

            try:

                # Tenta excluir o registro do animal no banco de dados.
                # O filtro usa o ID do animal selecionado para identificar o registro correto.
                db.animais.delete_one({"_id": ObjectId(self.animal_id_atual)})

                # Exibe uma mensagem de sucesso ao usuário, indicando que o animal foi excluído.
                messagebox.showinfo("Sucesso", "Animal excluído!")

                # Recarrega a lista de animais para refletir a exclusão na interface.
                self.carregar_animais()

                # Limpa o formulário, removendo os dados do animal excluído da interface.
                self.limpar_form()

            except Exception as e:

                # Em caso de erro, exibe uma mensagem informando o problema.
                # O erro é incluído na mensagem para ajudar no diagnóstico.
                messagebox.showerror("Erro", f"Não foi possível excluir:\n{e}")


    # Método para limpar o formulário, redefinindo todos os campos
    #       para seus valores iniciais ou padrão.
    def limpar_form(self):

        # Redefine o ID do animal atual para None, indicando que
        #       nenhum animal está selecionado.
        self.animal_id_atual = None

        # Limpa o campo de entrada para o nome do animal,
        #       removendo qualquer texto digitado.
        self.entry_nome_animal.delete(0, tk.END)

        # Limpa o campo de entrada para a espécie, removendo qualquer texto digitado.
        self.entry_especie.delete(0, tk.END)

        # Limpa o campo de entrada para a raça, removendo qualquer texto digitado.
        self.entry_raca.delete(0, tk.END)

        # Limpa o campo de entrada para a idade, removendo qualquer texto digitado.
        self.entry_idade.delete(0, tk.END)

        # Redefine o combobox de sexo para o valor padrão (primeira opção da lista).
        self.combo_sexo.current(0)

        # Limpa o campo de entrada para o peso, removendo qualquer texto digitado.
        self.entry_peso.delete(0, tk.END)

        # Se o combobox de donos tiver valores disponíveis, redefine-o para o
        #       valor padrão (primeira opção da lista).
        if self.combo_dono['values']:
            self.combo_dono.current(0)


    # Método para visualizar o histórico de atendimentos de um animal selecionado.
    def ver_historico(self):

        # Verifica se algum animal está selecionado. Caso contrário,
        #       exibe uma mensagem de aviso.
        if not self.animal_id_atual:

            # Exibe um aviso ao usuário.
            messagebox.showwarning("Aviso",
                                   "Selecione um animal para ver o histórico.")

            # Encerra o método, pois não há um animal selecionado.
            return

        # Abre a janela de histórico de atendimentos para o animal atualmente selecionado.
        # 'self.animal_id_atual' é passado como referência ao ID do animal.
        JanelaHistoricoAtendimentos(self, self.animal_id_atual)


# ---------------------------------------------------------------------
# TelaMedicos (CRUD)
# ---------------------------------------------------------------------
# Classe responsável por gerenciar a interface dos médicos.
class TelaMedicos(tk.Frame):

    # Método inicializador da classe 'TelaMedicos'.
    def __init__(self, master, medico_logado, *args, **kwargs):

        # Chama o inicializador da classe pai ('tk.Frame') para configurar o frame.
        super().__init__(master, *args, **kwargs)

        # O parâmetro 'self' refere-se à instância atual da classe,
        #       permitindo acessar atributos e métodos internos.
        # O parâmetro 'master' representa o widget pai, neste caso, a
        #       janela ou frame onde esta tela será inserida.
        # O parâmetro 'medico_logado' armazena os dados do médico atualmente logado no sistema.
        # Os parâmetros '*args' e '**kwargs' permitem que argumentos adicionais
        #       sejam passados, oferecendo flexibilidade na criação do objeto.

        # Armazena o médico logado para rastrear ações e alterações feitas pelo mesmo.
        self.medico_logado = medico_logado

        # Variável para armazenar o ID do médico atualmente selecionado.
        # Inicialmente None, pois nenhum está selecionado.
        self.medico_id_atual = None

        # Define a cor de fundo da interface.
        # 'bg="#F7F7F7"' especifica a cor de fundo como um cinza muito claro,
        #       conhecido como "Alabaster".
        self.configure(bg="#F7F7F7")

        # Cria um rótulo (Label) para exibir o título da tela.
        # 'text="Gerenciamento de Médicos"' define o texto do título.
        # 'font=("Arial", 18, "bold")' especifica a fonte como Arial, tamanho 18, em negrito.
        # 'bg="#F7F7F7"' define a cor de fundo do título como "Alabaster" (cinza
        #       muito claro) para combinar com o fundo geral.
        # 'fg="#333333"' define a cor do texto como um cinza escuro.
        titulo = tk.Label(self,
                          text="Gerenciamento de Médicos",
                          font=("Arial", 18, "bold"),
                          bg="#F7F7F7",
                          fg="#333333")

        # Posiciona o título na parte superior da tela.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical acima e abaixo do título.
        titulo.pack(pady=10)

        # Cria um container (LabelFrame) para o formulário de cadastro de médicos.
        # 'text="Cadastro de Médicos"' define o título do frame.
        # 'font=("Arial", 12, "bold")' especifica a fonte como Arial, tamanho 12, em negrito.
        # 'bg="#FFFFFF"' define a cor de fundo do frame como branco.
        # 'fg="#333333"' define a cor do texto do título como cinza escuro.
        # 'padx=10, pady=10' adiciona 10 pixels de margem interna horizontal e vertical dentro do frame.
        frame_form = tk.LabelFrame(self,
                                   text="Cadastro de Médicos",
                                   font=("Arial", 12, "bold"),
                                   bg="#FFFFFF",
                                   fg="#333333",
                                   padx=10,
                                   pady=10)

        # Posiciona o frame do formulário na parte superior da tela.
        # 'side="top"' alinha o frame na parte superior.
        # 'fill="x"' faz o frame se expandir horizontalmente para preencher a largura disponível.
        # 'padx=20, pady=10' adiciona 20 pixels de margem horizontal e 10
        #       pixels de margem vertical ao redor do frame.
        frame_form.pack(side="top",
                        fill="x",
                        padx=20,
                        pady=10)

        # Cria um rótulo (Label) para o campo de entrada do nome do médico.
        # 'text="Nome do Médico:"' define o texto do rótulo.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco, igual ao fundo do frame.
        tk.Label(frame_form,
                 text="Nome do Médico:",
                 font=("Arial", 12),
                 bg="#FFFFFF").grid(

            # Posiciona o rótulo na primeira linha (row=0) e na primeira
            #       coluna (column=0) da grade (grid) do frame.
            # 'sticky="e"' alinha o rótulo à direita (east) dentro da célula.
            # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do rótulo.
            # 'pady=5' adiciona 5 pixels de margem vertical ao redor do rótulo.
            row=0, column=0, sticky="e", padx=5, pady=5)

        # Cria um campo de entrada (Entry) para o usuário digitar o nome do médico.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        # 'width=40' define a largura do campo de entrada para acomodar até 40 caracteres.
        self.entry_nome = tk.Entry(frame_form,
                                   font=("Arial", 12),
                                   width=40)

        # Posiciona o campo de entrada na primeira linha (row=0) e na segunda
        #       coluna (column=1) da grade (grid) do frame.
        # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do campo de entrada.
        # 'pady=5' adiciona 5 pixels de margem vertical ao redor do campo de entrada.
        self.entry_nome.grid(row=0, column=1, padx=5, pady=5)

        # Cria um rótulo (Label) para o campo de seleção da especialidade do médico.
        # 'text="Especialidade:"' define o texto do rótulo.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco, igual ao fundo do frame.
        tk.Label(frame_form,
                 text="Especialidade:",
                 font=("Arial", 12),
                 bg="#FFFFFF").grid(

                            # Posiciona o rótulo na segunda linha (row=1) e na primeira
                            #       coluna (column=0) da grade (grid) do frame.
                            # 'sticky="e"' alinha o rótulo à direita (east) dentro da célula.
                            # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do rótulo.
                            # 'pady=5' adiciona 5 pixels de margem vertical ao redor do rótulo.
                            row=1, column=0, sticky="e", padx=5, pady=5)


        # Cria um campo de seleção (Combobox) para o usuário
        #       escolher a especialidade do médico.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        # 'width=38' define a largura do campo de seleção para acomodar até 38 caracteres.
        self.combo_especialidade = ttk.Combobox(frame_form,
                                                font=("Arial", 12),
                                                width=38)

        # Define as opções disponíveis na Combobox.
        # 'values' contém uma lista de especialidades que o médico pode escolher.
        self.combo_especialidade['values'] = ["Clínico Geral", "Pediatra", "Cardiologista", "Neurologista"]

        # Posiciona o campo de seleção na segunda linha (row=1) e na
        #       segunda coluna (column=1) da grade (grid) do frame.
        # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do campo de seleção.
        # 'pady=5' adiciona 5 pixels de margem vertical ao redor do campo de seleção.
        self.combo_especialidade.grid(row=1,
                                      column=1,
                                      padx=5,
                                      pady=5)

        # Cria um rótulo (Label) para o campo de entrada do CRMV (registro do médico veterinário).
        # 'text="CRMV:"' define o texto do rótulo.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco, igual ao fundo do frame.
        tk.Label(frame_form,
                 text="CRMV:",
                 font=("Arial", 12),
                 bg="#FFFFFF").grid(

                                # Posiciona o rótulo na terceira linha (row=2) e na primeira
                                #       coluna (column=0) da grade (grid) do frame.
                                # 'sticky="e"' alinha o rótulo à direita (east) dentro da célula.
                                # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do rótulo.
                                # 'pady=5' adiciona 5 pixels de margem vertical ao redor do rótulo.
                                row=2, column=0, sticky="e", padx=5, pady=5)


        # Cria um campo de entrada (Entry) para o usuário digitar o CRMV do médico.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        # 'width=40' define a largura do campo de entrada para acomodar até 40 caracteres.
        self.entry_crmv = tk.Entry(frame_form,
                                   font=("Arial", 12),
                                   width=40)

        # Posiciona o campo de entrada na terceira linha (row=2) e na
        #       segunda coluna (column=1) da grade (grid) do frame.
        # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do campo de entrada.
        # 'pady=5' adiciona 5 pixels de margem vertical ao redor do campo de entrada.
        self.entry_crmv.grid(row=2,
                             column=1,
                             padx=5,
                             pady=5)

        # Cria um rótulo (Label) para o campo de entrada da senha.
        # 'text="Senha:"' define o texto que será exibido no rótulo.
        # 'font=("Arial", 12)' define a fonte do texto como Arial, tamanho 12.
        # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco,
        #       para coincidir com o fundo do frame.
        tk.Label(frame_form,
                 text="Senha:",
                 font=("Arial", 12),
                 bg="#FFFFFF").grid(

                                    # Posiciona o rótulo na quarta linha (row=3) e na primeira coluna (column=0) do grid.
                                    # 'sticky="e"' alinha o rótulo à direita (east) dentro da célula.
                                    # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do rótulo.
                                    # 'pady=5' adiciona 5 pixels de margem vertical ao redor do rótulo.
                                    row=3, column=0, sticky="e", padx=5, pady=5)

        # Cria um campo de entrada (Entry) para o usuário digitar a senha.
        # 'font=("Arial", 12)' define a fonte do texto digitado como Arial, tamanho 12.
        # 'width=40' define a largura do campo de entrada para acomodar até 40 caracteres.
        # 'show="*"' oculta os caracteres digitados, substituindo-os por
        #       asteriscos (*) para maior segurança.
        self.entry_senha = tk.Entry(frame_form,
                                    font=("Arial", 12),
                                    width=40,
                                    show="*")

        # Posiciona o campo de entrada na quarta linha (row=3) e na
        #       segunda coluna (column=1) do grid.
        # 'padx=5' adiciona 5 pixels de margem horizontal ao redor do campo de entrada.
        # 'pady=5' adiciona 5 pixels de margem vertical ao redor do campo de entrada.
        self.entry_senha.grid(row=3,
                              column=1,
                              padx=5,
                              pady=5)

        # Cria um frame para agrupar os botões de ação.
        # 'bg="#FFFFFF"' define a cor de fundo do frame como branco, para
        #       coincidir com o fundo do formulário.
        frame_botoes = tk.Frame(frame_form, bg="#FFFFFF")

        # Posiciona o frame de botões na quinta linha (row=4) do grid.
        # 'column=0' coloca o frame na primeira coluna, e 'columnspan=2' faz
        #       com que o frame ocupe duas colunas.
        # 'pady=10' adiciona 10 pixels de margem vertical ao redor do
        #       frame para separação visual.
        frame_botoes.grid(row=4,
                          column=0,
                          columnspan=2,
                          pady=10)

        # Define uma lista de botões, onde cada item contém:
        # - O texto a ser exibido no botão.
        # - A função que será executada quando o botão for clicado.
        # - A cor de fundo (background) do botão.
        botoes = [

            ("Salvar/Atualizar", self.salvar, "#4CAF50"),  # Botão verde para salvar ou atualizar o médico.
            ("Excluir", self.excluir, "#F44336"),  # Botão vermelho para excluir o médico selecionado.
            ("Limpar", self.limpar_form, "#2196F3"),  # Botão azul para limpar o formulário.

        ]

        # Itera sobre a lista de botões definida anteriormente.
        # Cada item da lista contém o texto do botão, o comando a ser
        #       executado, e a cor de fundo.
        for text, command, color in botoes:

            # Cria um botão para cada item da lista.
            # 'text=text' define o texto exibido no botão.
            # 'font=("Arial", 12, "bold")' aplica a fonte Arial, tamanho 12, em negrito.
            # 'bg=color' define a cor de fundo do botão com base na cor
            #       fornecida na lista (verde, vermelho ou azul).
            # 'fg="#FFFFFF"' define a cor do texto do botão como branco.
            # 'activebackground=color' define a cor de fundo quando o botão é
            #       clicado ou está ativo, mantendo a cor inicial.
            # 'activeforeground="#FFFFFF"' define a cor do texto ativo como branco.
            # 'command=command' vincula a função correspondente ao botão, como salvar, excluir ou limpar.
            # 'width=15' define a largura do botão, garantindo consistência no tamanho.
            tk.Button(  frame_botoes,
                        text=text,
                        font=("Arial", 12, "bold"),
                        bg=color,
                        fg="#FFFFFF",
                        activebackground=color,
                        activeforeground="#FFFFFF",
                        command=command,
                        width=15).pack(

                                    # Posiciona o botão dentro do frame_botoes.
                                    # 'side="left"' alinha os botões lado a lado, da esquerda para a direita.
                                    # 'padx=10' adiciona 10 pixels de espaço horizontal entre os botões.
                                    side="left",
                                    padx=10)

        # Cria um LabelFrame chamado 'frame_lista' para exibir a lista de médicos.
        # 'text="Lista de Médicos"' define o título do LabelFrame.
        # 'font=("Arial", 12, "bold")' aplica a fonte Arial, tamanho 12, em negrito.
        # 'bg="#FFFFFF"' define a cor de fundo do frame como branco.
        # 'fg="#333333"' define a cor do texto como cinza escuro.
        # 'padx=10' adiciona 10 pixels de espaço interno horizontal dentro do frame.
        # 'pady=10' adiciona 10 pixels de espaço interno vertical dentro do frame.
        frame_lista = tk.LabelFrame(self,
                                    text="Lista de Médicos",
                                    font=("Arial", 12, "bold"),
                                    bg="#FFFFFF",
                                    fg="#333333",
                                    padx=10,
                                    pady=10)

        # Posiciona o LabelFrame na parte inferior da tela.
        # 'side="bottom"' alinha o frame na parte inferior da janela.
        # 'fill="both"' permite que o frame preencha tanto a largura
        #       quanto a altura disponíveis.
        # 'expand=True' permite que o frame expanda proporcionalmente
        #       quando a janela é redimensionada.
        # 'padx=20' adiciona 20 pixels de margem horizontal externa ao redor do frame.
        # 'pady=10' adiciona 10 pixels de margem vertical externa ao redor do frame.
        frame_lista.pack(side="bottom",
                         fill="both",
                         expand=True,
                         padx=20,
                         pady=10)

        # Define as colunas para a Treeview que exibirá os médicos.
        # Cada elemento da tupla representa uma coluna na
        #       tabela: "_id", "nome", "especialidade", "crmv" e "senha".
        colunas = ("_id", "nome", "especialidade", "crmv", "senha")

        # Cria um estilo para personalizar a aparência da Treeview.
        estilo = ttk.Style()

        # Configura o estilo geral da Treeview.
        # 'font=("Arial", 10)' define a fonte Arial com tamanho 10 para o texto das células.
        # 'rowheight=25' ajusta a altura de cada linha para 25 pixels.
        # 'background="#FFFFFF"' define o fundo padrão das células como branco.
        # 'fieldbackground="#FFFFFF"' define a cor de fundo do campo de edição como branco.
        estilo.configure("Treeview",
                         font=("Arial", 10),
                         rowheight=25,
                         background="#FFFFFF",
                         fieldbackground="#FFFFFF")

        # Configura o estilo do cabeçalho da Treeview.
        # 'font=("Arial", 12, "bold")' define a fonte Arial, tamanho 12,
        #       com negrito para os cabeçalhos.
        # 'background="#F0F0F0"' define o fundo dos cabeçalhos como cinza claro.
        # 'foreground="#333333"' define a cor do texto dos cabeçalhos como cinza escuro.
        estilo.configure("Treeview.Heading",
                         font=("Arial", 12, "bold"),
                         background="#F0F0F0",
                         foreground="#333333")

        # Configura o mapeamento de cores para a Treeview.
        # 'background=[("selected", "#D3D3D3")]' define que, ao selecionar uma
        #       linha, sua cor de fundo muda para cinza claro.
        estilo.map("Treeview",
                   background=[("selected", "#D3D3D3")])


        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado

        # Cria uma Treeview para exibir a lista de médicos.
        # 'frame_lista' é o container onde a Treeview será posicionada.
        # 'columns=colunas' define as colunas da Treeview conforme especificado anteriormente.
        # 'show="headings"' indica que apenas os cabeçalhos e as colunas serão
        #       exibidos (sem a coluna de índice padrão).
        # 'style="Treeview"' aplica o estilo configurado anteriormente.
        self.tree = ttk.Treeview(frame_lista,
                                 columns=colunas,
                                 show="headings",
                                 style="Treeview")

        # Configura os cabeçalhos e as colunas da Treeview.
        for c in colunas:

            # Define o texto de cada cabeçalho da coluna.
            # 'c.title()' converte o texto da coluna para título (primeira letra maiúscula).
            self.tree.heading(c, text=c.title())

            # Define a configuração de cada coluna:
            # 'anchor="center"' alinha o texto ao centro.
            # 'width=150' ajusta a largura das colunas, exceto para a
            #       coluna "_id", que tem largura 120.
            self.tree.column(c, anchor="center", width=150 if c != "_id" else 120)

        # Posiciona a Treeview no container.
        # 'fill="both"' faz com que a Treeview preencha tanto a largura
        #       quanto a altura disponíveis no container.
        # 'expand=True' permite que a Treeview expanda proporcionalmente
        #       quando a janela é redimensionada.
        # 'pady=5' adiciona um espaçamento vertical de 5 pixels
        #       acima e abaixo da Treeview.
        self.tree.pack(fill="both", expand=True, pady=5)

        # Cria uma barra de rolagem vertical para a Treeview.
        # 'frame_lista' é o container onde a barra de rolagem será posicionada.
        # 'orient="vertical"' define a orientação da barra de rolagem como vertical.
        # 'command=self.tree.yview' conecta a barra de rolagem à Treeview,
        #       permitindo rolar verticalmente.
        scrollbar = ttk.Scrollbar(frame_lista,
                                  orient="vertical",
                                  command=self.tree.yview)

        # Configura a Treeview para usar a barra de rolagem.
        # 'yscrollcommand=scrollbar.set' faz com que a barra de rolagem
        #       responda às interações na Treeview.
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do container.
        # 'side="right"' alinha a barra de rolagem à direita do container.
        # 'fill="y"' faz com que a barra de rolagem preencha toda a
        #       altura disponível no container.
        scrollbar.pack(side="right", fill="y")

        # Vincula o evento de seleção de uma linha na Treeview a um método.
        # '<<TreeviewSelect>>' é o evento que ocorre quando o usuário
        #       seleciona uma linha na Treeview.
        # 'self.ao_selecionar_linha' é o método que será chamado ao ocorrer o evento.
        self.tree.bind("<<TreeviewSelect>>", self.ao_selecionar_linha)

        # Chama o método para carregar a lista de médicos na Treeview.
        # Isso preenche a Treeview com os dados dos médicos
        #       disponíveis no banco de dados.
        self.carregar_medicos()


    def carregar_medicos(self):

        # Remove todos os itens existentes na Treeview.
        # 'self.tree.get_children()' retorna os identificadores de
        #       todos os itens na Treeview.
        # 'self.tree.delete(item)' remove cada item, garantindo que a
        #       lista seja limpa antes de ser recarregada.
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Itera sobre todos os registros de médicos no banco de dados.
        # 'db.medicos.find()' consulta o banco de dados e retorna todos os
        #       documentos na coleção 'medicos'.
        for med in db.medicos.find():

            # Insere uma nova linha na Treeview para cada médico encontrado.
            # 'values=()' define as colunas exibidas na Treeview, extraindo
            #       dados do documento 'med'.
            # 'str(med["_id"])' converte o identificador único do
            #       médico (ObjectId) em string para exibição.
            # 'med.get("nome", "")' obtém o valor do campo 'nome' ou uma
            #       string vazia se o campo não existir.
            # Campos como 'especialidade', 'crmv' e 'senha' são tratados de forma semelhante.
            self.tree.insert("",
                             tk.END,
                             values=(str(med["_id"]),
                                    med.get("nome", ""),
                                    med.get("especialidade", ""),
                                    med.get("crmv", ""),
                                    med.get("senha", "")))


    def ao_selecionar_linha(self, event):

        # Obtém a seleção atual na Treeview.
        # 'self.tree.selection()' retorna uma tupla de IDs dos itens selecionados.
        selection = self.tree.selection()

        # Verifica se nenhuma linha foi selecionada.
        # Se a tupla estiver vazia, simplesmente retorna (nada é feito).
        if not selection:
            return

        # Obtém os dados do item selecionado.
        # 'self.tree.item(selection[0])' retorna um dicionário com os detalhes do item.
        # 'selection[0]' é o ID da primeira linha selecionada (apenas um
        #       item pode ser selecionado de cada vez).
        item = self.tree.item(selection[0])

        # Extrai os valores das colunas do item selecionado.
        # 'item["values"]' é uma lista com os valores das colunas na Treeview.
        vals = item["values"]

        # Atualiza a variável que mantém o ID do médico atualmente selecionado.
        # O ID é usado em outras operações como edição ou exclusão.
        self.medico_id_atual = vals[0]

        # Preenche o campo de entrada do nome do médico com o valor selecionado.
        # Remove o texto existente no campo antes de inserir o novo valor.
        self.entry_nome.delete(0, tk.END)
        self.entry_nome.insert(0, vals[1])

        # Define o valor da especialidade no ComboBox correspondente.
        # Usa o método 'set' para atualizar o valor exibido.
        self.combo_especialidade.set(vals[2])

        # Preenche o campo de entrada do CRMV com o valor selecionado.
        # Remove o texto existente antes de inserir o novo valor.
        self.entry_crmv.delete(0, tk.END)
        self.entry_crmv.insert(0, vals[3])

        # Preenche o campo de entrada da senha com o valor selecionado.
        # Remove o texto existente antes de inserir o novo valor.
        self.entry_senha.delete(0, tk.END)
        self.entry_senha.insert(0, vals[4])


    def salvar(self):

        # Obtém o valor do campo de entrada 'Nome'.
        # 'self.entry_nome.get()' pega o texto inserido no campo.
        # '.strip()' remove espaços em branco no início e no fim.
        nome = self.entry_nome.get().strip()

        # Verifica se o campo 'Nome' está vazio.
        # Caso esteja, exibe um aviso e encerra a execução da função.
        if not nome:
            messagebox.showwarning("Aviso",
                                   "O campo 'Nome do Médico' é obrigatório.")
            return

        # Cria um dicionário com os dados do médico.
        # Este dicionário será usado para salvar ou atualizar os dados no banco.
        doc = {

            # Nome do médico, extraído do campo de entrada.
            "nome": nome,

            # Especialidade do médico.
            # 'self.combo_especialidade.get()' obtém o valor selecionado no ComboBox.
            # '.strip()' remove espaços desnecessários.
            "especialidade": self.combo_especialidade.get().strip(),

            # CRMV do médico, extraído do campo de entrada correspondente.
            "crmv": self.entry_crmv.get().strip(),

            # Senha do médico, extraída do campo de entrada.
            "senha": self.entry_senha.get().strip(),

            # ID do usuário que realizou a última alteração.
            # Obtido da variável 'self.medico_logado'.
            "ultima_alteracao_por_id": self.medico_logado["_id"],

            # Nome do usuário que realizou a última alteração.
            # Também extraído de 'self.medico_logado'.
            "ultima_alteracao_por_nome": self.medico_logado["nome"],

            # Data e hora da última alteração.
            # Utiliza a função 'datetime.datetime.now()' para registrar o momento atual.
            "ultima_alteracao_em": datetime.datetime.now()

        }

        try:

            # Verifica se o médico atual está sendo criado ou atualizado.
            # Caso 'self.medico_id_atual' seja 'None', é um novo registro.
            if self.medico_id_atual is None:

                # Insere o novo documento no banco de dados na coleção 'medicos'.
                # O dicionário 'doc' contém os dados do médico.
                db.medicos.insert_one(doc)

            else:

                # Atualiza um registro existente na coleção 'medicos'.
                # A consulta utiliza o ID do médico atual para localizar o registro.
                # O operador '$set' é usado para atualizar apenas os campos especificados em 'doc'.
                db.medicos.update_one({"_id": ObjectId(self.medico_id_atual)}, {"$set": doc})

            # Exibe uma mensagem de sucesso indicando que o médico foi salvo ou atualizado.
            messagebox.showinfo("Sucesso", "Médico salvo/atualizado!")

            # Recarrega a lista de médicos para refletir as alterações no banco de dados.
            self.carregar_medicos()

            # Limpa o formulário para permitir a entrada de novos dados.
            self.limpar_form()

        except Exception as e:

            # Caso ocorra qualquer erro durante a operação, exibe uma mensagem de erro.
            # O erro capturado é incluído na mensagem para facilitar a depuração.
            messagebox.showerror("Erro", f"Não foi possível salvar:\n{e}")



    def excluir(self):

        # Verifica se há um médico selecionado.
        if not self.medico_id_atual:

            # Exibe uma mensagem de aviso caso nenhum médico esteja selecionado.
            messagebox.showwarning("Aviso",
                                   "Selecione um médico para excluir.")
            return

        # Pergunta ao usuário se ele realmente deseja excluir o médico selecionado.
        resp = messagebox.askyesno("Confirmação",
                                   "Deseja excluir este médico?")

        # Caso o usuário confirme a exclusão.
        if resp:

            try:

                # Remove o médico selecionado do banco de dados.
                # Utiliza o ID do médico atual para localizar e excluir o registro.
                db.medicos.delete_one({"_id": ObjectId(self.medico_id_atual)})

                # Exibe uma mensagem informando que o médico foi excluído com sucesso.
                messagebox.showinfo("Sucesso", "Médico excluído!")

                # Recarrega a lista de médicos para refletir a exclusão.
                self.carregar_medicos()

                # Limpa o formulário para evitar que dados do médico
                #       excluído permaneçam visíveis.
                self.limpar_form()

            except Exception as e:

                # Caso ocorra qualquer erro durante a exclusão, exibe uma mensagem de erro.
                # Inclui o erro capturado na mensagem para auxiliar na depuração.
                messagebox.showerror("Erro",
                                     f"Não foi possível excluir:\n{e}")



    def limpar_form(self):

        # Remove a referência ao ID do médico atual, indicando que
        #       nenhum médico está selecionado.
        self.medico_id_atual = None

        # Limpa o campo de entrada para o nome do médico.
        # delete(0, tk.END) remove todo o texto do campo.
        self.entry_nome.delete(0, tk.END)

        # Reseta a seleção da especialidade para um valor vazio.
        # set("") limpa a seleção atual no combobox.
        self.combo_especialidade.set("")

        # Limpa o campo de entrada para o CRMV.
        self.entry_crmv.delete(0, tk.END)

        # Limpa o campo de entrada para a senha.
        # Este campo utiliza 'show="*"' para mascarar os caracteres.
        self.entry_senha.delete(0, tk.END)


# ---------------------------------------------------------------------
# TelaAgenda (CRUD)
# ---------------------------------------------------------------------
class TelaAgenda(tk.Frame):

    # Método inicializador da classe 'TelaAgenda'.
    # Este método é chamado quando uma nova instância da classe é criada.
    # Ele configura os elementos básicos da aba de agenda, como
    #       variáveis e propriedades do frame.
    def __init__(self, master, medico_logado, *args, **kwargs):

        # 'master' é o widget pai ao qual este frame será anexado. É um container maior.
        # Neste caso, 'master' representa o notebook ou janela principal.

        # 'medico_logado' é um objeto que contém informações sobre o
        #       médico atualmente logado no sistema.
        # Isso permite associar ações ou permissões ao usuário logado.

        # '*args' e '**kwargs' são utilizados para aceitar argumentos e
        #       parâmetros adicionais.
        # '*args' é uma lista de argumentos posicionais.
        # '**kwargs' é um dicionário de argumentos nomeados, permitindo
        #       flexibilidade na chamada do método.

        # Chama o inicializador da classe pai ('tk.Frame'), passando 'master', '*args' e '**kwargs'.
        # Isso garante que o frame herde corretamente as propriedades e métodos da classe pai.
        super().__init__(master, *args, **kwargs)

        # Armazena as informações do médico logado.
        # Isso será usado para determinar permissões e acessar informações relacionadas.
        self.medico_logado = medico_logado

        # Inicializa a variável que armazena o ID do agendamento selecionado.
        # Quando nenhum agendamento está selecionado, ela é definida como 'None'.
        self.agenda_id_atual = None

        # Define a cor de fundo do frame principal.
        # '#F7F7F7' é um tom claro de cinza, ideal para uma aparência limpa e profissional.
        self.configure(bg="#F7F7F7")

        # Cria um rótulo (label) para exibir o título da seção.
        # 'text="Gerenciamento de Agendamentos"' define o texto
        #       que será exibido no rótulo.
        # 'font=("Arial", 18, "bold")' configura a fonte como Arial,
        #       tamanho 18, e estilo negrito.
        # 'bg="#F7F7F7"' define a cor de fundo do rótulo como
        #       cinza claro (código hexadecimal).
        # 'fg="#333333"' define a cor do texto como cinza escuro.
        titulo = tk.Label(self,
                          text="Gerenciamento de Agendamentos",
                          font=("Arial", 18, "bold"),
                          bg="#F7F7F7",
                          fg="#333333")

        # Posiciona o rótulo no frame atual.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical acima e abaixo do rótulo.
        titulo.pack(pady=10)

        # Cria um frame para o formulário de cadastro de agendamento.
        # 'text="Cadastro de Agendamento"' define o título do frame.
        # 'font=("Arial", 12, "bold")' configura a fonte como
        #       Arial, tamanho 12, e estilo negrito.
        # 'bg="#FFFFFF"' define o fundo do frame como branco.
        # 'fg="#333333"' define o texto do título como cinza escuro.
        # 'padx=10, pady=10' adiciona 10 pixels de espaçamento interno
        #       horizontal e vertical no frame.
        frame_form = tk.LabelFrame(self,
                                   text="Cadastro de Agendamento",
                                   font=("Arial", 12, "bold"),
                                   bg="#FFFFFF",
                                   fg="#333333",
                                   padx=10,
                                   pady=10)

        # Posiciona o frame no topo da tela (side="top").
        # 'fill="x"' faz o frame ocupar toda a largura disponível.
        # 'padx=20, pady=10' adiciona margens horizontais de 20 pixels e verticais de 10 pixels.
        frame_form.pack(side="top",
                        fill="x",
                        padx=20,
                        pady=10)

        # Define os campos do formulário.
        # Cada tupla contém o texto do rótulo do campo e o nome do
        #       atributo que será associado.
        campos = [

            ("Animal:", "combo_animal"),  # Campo para selecionar o animal.
            ("Dono do Animal:", "label_dono"),  # Campo para exibir o dono do animal.
            ("Médico:", "combo_medico"),  # Campo para selecionar o médico.
            ("Tipo de Consulta:", "entry_tipo"),  # Campo para o tipo de consulta.
            ("Data:", "data_entry"),  # Campo para a data do agendamento.
            ("Hora (HH:MM):", "entry_hora"),  # Campo para a hora do agendamento.
            ("Status:", "combo_status"),  # Campo para o status do agendamento.
            ("Produtos Usados (sep. por vírgula):", "entry_produtos")  # Campo para os produtos usados.

        ]

        # Itera sobre a lista de campos para criar rótulos e entradas dinamicamente.
        for i, (label, attr_name) in enumerate(campos):

            # Cria um rótulo para cada campo.
            # 'text=label' define o texto do rótulo como o primeiro elemento da tupla.
            # 'font=("Arial", 12)' define o estilo da fonte.
            # 'bg="#FFFFFF"' define o fundo do rótulo como branco.
            # 'row=i, column=0' posiciona o rótulo na linha correspondente ao
            #       índice e na primeira coluna.
            # 'sticky="e"' alinha o rótulo à direita.
            # 'padx=5, pady=5' adiciona margens internas ao redor do rótulo.
            tk.Label(frame_form,
                     text=label,
                     font=("Arial", 12),
                     bg="#FFFFFF").grid(row=i,
                                        column=0,
                                        sticky="e",
                                        padx=5,
                                        pady=5)

            # Verifica se o campo atual é o "combo_animal".
            if attr_name == "combo_animal":

                # Cria uma caixa de seleção para o campo "Animal".
                # 'font=("Arial", 12)' define o estilo da fonte da caixa de seleção.
                # 'width=50' define a largura da caixa de seleção.
                combo = ttk.Combobox(frame_form,
                                     font=("Arial", 12),
                                     width=50)

                # Associa um evento à caixa de seleção que será disparado ao selecionar um item.
                # O método '_quando_seleciona_animal' será chamado sempre que o valor da caixa mudar.
                combo.bind("<<ComboboxSelected>>", self._quando_seleciona_animal)

                # Define o atributo 'combo_animal' como a caixa de seleção criada.
                setattr(self, attr_name, combo)

                # Posiciona a caixa de seleção no formulário.
                # 'row=i, column=1' coloca a caixa na mesma linha do rótulo e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas ao redor da caixa.
                combo.grid(row=i,
                           column=1,
                           padx=5,
                           pady=5)

            # Verifica se o campo atual é o "label_dono".
            elif attr_name == "label_dono":

                # Cria um rótulo para exibir o nome do dono do animal.
                # 'text="---"' define o texto inicial do rótulo.
                # 'font=("Arial", 12)' define o estilo da fonte.
                # 'bg="#FFFFFF"' define o fundo do rótulo como branco.
                # 'anchor="w"' alinha o texto do rótulo à esquerda.
                # 'width=40' define a largura do rótulo.
                label = tk.Label(frame_form,
                                 text="---",
                                 font=("Arial", 12),
                                 bg="#FFFFFF",
                                 anchor="w",
                                 width=40)

                # Define o atributo 'label_dono' como o rótulo criado.
                setattr(self, attr_name, label)

                # Posiciona o rótulo no formulário.
                # 'row=i, column=1' coloca o rótulo na mesma linha do
                #       rótulo de "Dono do Animal" e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas ao redor do rótulo.
                # 'sticky="w"' alinha o rótulo à esquerda.
                label.grid(row=i, column=1, padx=5, pady=5, sticky="w")

            # Verifica se o campo atual é o "combo_medico".
            elif attr_name == "combo_medico":

                # Cria uma caixa de seleção para o campo "Médico".
                # 'font=("Arial", 12)' define o estilo da fonte da caixa de seleção.
                # 'width=50' define a largura da caixa de seleção.
                combo = ttk.Combobox(frame_form,
                                     font=("Arial", 12),
                                     width=50)

                # Define o atributo 'combo_medico' como a caixa de seleção criada.
                setattr(self, attr_name, combo)

                # Posiciona a caixa de seleção no formulário.
                # 'row=i, column=1' coloca a caixa na mesma linha do rótulo
                #       de "Médico" e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas ao redor da caixa.
                combo.grid(row=i, column=1, padx=5, pady=5)

            # Verifica se o campo atual é o "entry_tipo".
            elif attr_name == "entry_tipo":

                # Cria um campo de entrada para o tipo de consulta.
                # 'font=("Arial", 12)' define o estilo da fonte.
                # 'width=52' define a largura do campo.
                entry = tk.Entry(frame_form, font=("Arial", 12), width=52)

                # Define o atributo 'entry_tipo' como o campo de entrada criado.
                setattr(self, attr_name, entry)

                # Posiciona o campo de entrada no formulário.
                # 'row=i, column=1' coloca o campo na mesma linha do rótulo
                #       de "Tipo de Consulta" e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas ao redor do campo.
                entry.grid(row=i, column=1, padx=5, pady=5)

            # Verifica se o campo atual é o "data_entry".
            elif attr_name == "data_entry":

                # Cria um seletor de data para o campo "Data".
                # 'date_pattern="dd/MM/yyyy"' define o formato da data exibida no campo.
                # 'font=("Arial", 12)' define o estilo da fonte do seletor.
                # 'width=20' define a largura do seletor de data.
                # 'background="#4CAF50"' define a cor de fundo do seletor de data.
                # 'foreground="white"' define a cor do texto no seletor.
                calendar = DateEntry(frame_form,
                                     date_pattern='dd/MM/yyyy',
                                     font=("Arial", 12),
                                     width=20,
                                     background="#4CAF50",
                                     foreground="white")

                # Define o atributo 'data_entry' como o seletor de data criado.
                setattr(self, attr_name, calendar)

                # Posiciona o seletor de data no formulário.
                # 'row=i, column=1' coloca o seletor na mesma linha do
                #       rótulo de "Data" e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas ao redor do seletor.
                # 'sticky="w"' alinha o seletor à esquerda.
                calendar.grid(row=i,
                              column=1,
                              padx=5,
                              pady=5,
                              sticky="w")

            # Verifica se o campo atual é "entry_hora".
            elif attr_name == "entry_hora":

                # Cria um campo de entrada para a hora do agendamento.
                # 'font=("Arial", 12)' define a fonte utilizada no campo.
                # 'width=10' define a largura do campo de entrada.
                entry = tk.Entry(frame_form, font=("Arial", 12), width=10)

                # Insere um valor padrão no campo de entrada ("09:00").
                # Isso facilita a seleção de horários comuns para o usuário.
                entry.insert(0, "09:00")

                # Define o atributo 'entry_hora' como o campo de entrada criado.
                setattr(self, attr_name, entry)

                # Posiciona o campo de entrada no formulário.
                # 'row=i, column=1' posiciona o campo na mesma linha do
                #       rótulo "Hora (HH:MM)" e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas horizontais e
                #       verticais ao redor do campo.
                # 'sticky="w"' alinha o campo à esquerda da célula.
                entry.grid(row=i,
                           column=1,
                           padx=5,
                           pady=5,
                           sticky="w")

            # Verifica se o campo atual é "combo_status".
            elif attr_name == "combo_status":

                # Cria um combobox para selecionar o status do agendamento.
                # 'values=["Pendente", "Confirmado", "Realizado", "Cancelado"]'
                #       define as opções disponíveis no combobox.
                # 'font=("Arial", 12)' define o estilo da fonte no combobox.
                # 'width=50' define a largura do combobox.
                combo = ttk.Combobox(frame_form,
                                     values=["Pendente", "Confirmado", "Realizado", "Cancelado"],
                                     font=("Arial", 12),
                                     width=50)

                # Define a opção padrão do combobox como a primeira ("Pendente").
                combo.current(0)

                # Define o atributo 'combo_status' como o combobox criado.
                setattr(self, attr_name, combo)

                # Posiciona o combobox no formulário.
                # 'row=i, column=1' posiciona o combobox na mesma linha do
                #       rótulo "Status" e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas horizontais e
                #       verticais ao redor do combobox.
                combo.grid(row=i, column=1, padx=5, pady=5)

            # Verifica se o campo atual é "entry_produtos".
            elif attr_name == "entry_produtos":

                # Cria um campo de entrada para listar os produtos usados no agendamento.
                # 'font=("Arial", 12)' define o estilo da fonte no campo de entrada.
                # 'width=52' define a largura do campo, permitindo que o
                #       usuário insira uma lista de produtos.
                entry = tk.Entry(frame_form, font=("Arial", 12), width=52)

                # Define o atributo 'entry_produtos' como o campo de entrada criado.
                setattr(self, attr_name, entry)

                # Posiciona o campo de entrada no formulário.
                # 'row=i, column=1' posiciona o campo na mesma linha do
                #       rótulo "Produtos Usados (sep. por vírgula):" e na segunda coluna.
                # 'padx=5, pady=5' adiciona margens internas horizontais e
                #       verticais ao redor do campo.
                entry.grid(row=i, column=1, padx=5, pady=5)


        # Cria um frame para organizar os botões de ação.
        # 'bg="#FFFFFF"' define a cor de fundo do frame como branco.
        frame_botoes = tk.Frame(frame_form, bg="#FFFFFF")

        # Posiciona o frame no formulário.
        # 'row=len(campos)' coloca o frame na linha logo abaixo dos campos do formulário.
        # 'column=0, columnspan=2' faz com que o frame ocupe ambas as colunas do formulário.
        # 'pady=10' adiciona 10 pixels de margem vertical abaixo do frame para espaçamento.
        frame_botoes.grid(row=len(campos),
                          column=0,
                          columnspan=2,
                          pady=10)

        # Define a lista de botões com seu texto, comando
        #       associado e cor de fundo.
        botoes = [

            ("Salvar/Atualizar", self.salvar, "#4CAF50"),
            # Botão para salvar ou atualizar um agendamento, com fundo verde.
            ("Excluir", self.excluir, "#F44336"),  # Botão para excluir um agendamento, com fundo vermelho.
            ("Limpar", self.limpar_form, "#2196F3")  # Botão para limpar o formulário, com fundo azul.

        ]

        # Itera sobre a lista de botões para criar cada botão dinamicamente.
        for text, command, color in botoes:

            # Cria um botão com as seguintes configurações:
            # 'text=text' define o texto exibido no botão.
            # 'font=("Arial", 12, "bold")' define a fonte do texto
            #       como Arial, tamanho 12, em negrito.
            # 'bg=color' define a cor de fundo do botão com base na cor fornecida na lista.
            # 'fg="#FFFFFF"' define a cor do texto como branco.
            # 'activebackground=color' define a cor de fundo ao clicar
            #       no botão como a mesma cor do fundo.
            # 'activeforeground="#FFFFFF"' mantém o texto branco ao clicar no botão.
            # 'command=command' associa o botão à função fornecida na lista de botões.
            # 'width=15' define uma largura fixa para o botão, garantindo consistência no layout.
            tk.Button(frame_botoes,
                      text=text,
                      font=("Arial", 12, "bold"),
                      bg=color,
                      fg="#FFFFFF",
                      activebackground=color,
                      activeforeground="#FFFFFF",
                      command=command,
                      width=15).pack(side="left",  # Alinha os botões horizontalmente (à esquerda do container).
                                     padx=10)  # Adiciona 10 pixels de espaçamento horizontal entre os botões.

        # Cria um LabelFrame para a exibição da lista de agendamentos.
        frame_lista = tk.LabelFrame(

            self,  # Define o container pai como o frame principal da aba.
            text="Lista de Agendamentos",  # Define o título exibido no LabelFrame.
            font=("Arial", 12, "bold"),  # Define a fonte como Arial, tamanho 12, em negrito.
            bg="#FFFFFF",  # Define a cor de fundo como branco (#FFFFFF).
            fg="#333333",  # Define a cor do texto como cinza escuro (#333333).
            padx=10,  # Adiciona 10 pixels de margem interna horizontal.
            pady=10  # Adiciona 10 pixels de margem interna vertical.

        )

        # Posiciona o LabelFrame na parte inferior da tela.
        # 'side="bottom"' alinha o LabelFrame à parte inferior do container pai.
        # 'fill="both"' faz com que o LabelFrame preencha a largura e a altura disponíveis.
        # 'expand=True' permite que o LabelFrame expanda proporcionalmente ao redimensionar a janela.
        # 'padx=20' adiciona 20 pixels de margem externa horizontal ao redor do LabelFrame.
        # 'pady=10' adiciona 10 pixels de margem externa vertical ao redor do LabelFrame.
        frame_lista.pack(side="bottom",
                         fill="both",
                         expand=True,
                         padx=20,
                         pady=10)

        # Define as colunas para a exibição da tabela de agendamentos.
        colunas = ("_id", "nome_animal", "nome_dono", "nome_medico", "data_hora", "tipo", "status", "produtos")

        # Configura o estilo visual da Treeview (tabela).
        estilo = ttk.Style()

        # Define o estilo para as células da tabela.
        # 'font=("Arial", 10)' usa a fonte Arial, tamanho 10.
        # 'rowheight=25' define a altura das linhas como 25 pixels.
        # 'background="#FFFFFF"' define a cor de fundo das células como branco.
        # 'fieldbackground="#FFFFFF"' também define o fundo de edição como branco.
        estilo.configure("Treeview",
                         font=("Arial", 10),
                         rowheight=25,
                         background="#FFFFFF",
                         fieldbackground="#FFFFFF")

        # Define o estilo para os cabeçalhos da tabela.
        # 'font=("Arial", 12, "bold")' usa a fonte Arial, tamanho 12, em negrito.
        # 'background="#F0F0F0"' define a cor de fundo dos cabeçalhos como cinza claro.
        # 'foreground="#333333"' define a cor do texto dos cabeçalhos como cinza escuro.
        estilo.configure("Treeview.Heading",
                         font=("Arial", 12, "bold"),
                         background="#F0F0F0",
                         foreground="#333333")

        # Configura o comportamento visual para quando uma linha for selecionada.
        # 'background=[("selected", "#D3D3D3")]' define a cor de fundo como
        #       cinza claro quando a linha é selecionada.
        estilo.map("Treeview",
                   background=[("selected", "#D3D3D3")])


        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria a Treeview (tabela) no frame 'frame_lista'.
        # 'columns=colunas' define as colunas previamente configuradas.
        # 'show="headings"' indica que apenas os cabeçalhos e dados serão
        #       exibidos, sem a coluna extra para ícones.
        # 'style="Treeview"' aplica o estilo configurado anteriormente.
        self.tree = ttk.Treeview(frame_lista,
                                 columns=colunas,
                                 show="headings",
                                 style="Treeview")

        # Configura cada coluna na tabela.
        for c in colunas:

            # Define o cabeçalho da coluna.
            # 'text=c.replace("_", " ").title()' substitui underscores por
            #       espaços e aplica capitalização em cada palavra.
            self.tree.heading(c, text=c.replace("_", " ").title())

            # Configura o alinhamento e largura da coluna.
            # 'anchor="center"' alinha o texto ao centro.
            # 'width=150' define a largura padrão de cada coluna como 150 pixels.
            self.tree.column(c, anchor="center", width=150)

        # Adiciona a Treeview ao frame com preenchimento.
        # 'fill="both"' faz com que a tabela se ajuste tanto à
        #       largura quanto à altura do frame.
        # 'expand=True' permite que a tabela cresça proporcionalmente ao
        #       redimensionar a janela.
        # 'pady=5' adiciona 5 pixels de margem vertical acima e abaixo da tabela.
        self.tree.pack(fill="both", expand=True, pady=5)

        # Cria uma barra de rolagem vertical para a Treeview (tabela).
        # 'orient="vertical"' indica que a barra será vertical.
        # 'command=self.tree.yview' conecta a barra à rolagem vertical da Treeview.
        scrollbar = ttk.Scrollbar(frame_lista,
                                  orient="vertical",
                                  command=self.tree.yview)

        # Configura a Treeview para usar a barra de rolagem.
        # 'yscrollcommand=scrollbar.set' conecta a barra de rolagem à
        #       Treeview, permitindo que ela seja sincronizada.
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do frame.
        # 'side="right"' alinha a barra à direita.
        # 'fill="y"' faz com que a barra preencha verticalmente
        #       todo o espaço disponível.
        scrollbar.pack(side="right", fill="y")

        # Adiciona um evento de seleção à Treeview.
        # Quando uma linha na tabela é selecionada, o
        #       método 'ao_selecionar_linha' é chamado.
        self.tree.bind("<<TreeviewSelect>>", self.ao_selecionar_linha)

        # Vincula um evento de visibilidade à aba.
        # Quando a aba se torna visível, o método 'on_tab_visible' é
        #       chamado para carregar ou atualizar os dados.
        self.bind("<Visibility>", self.on_tab_visible)


    # Método 'on_tab_visible' é chamado quando a aba da agenda se torna visível.
    # Este método é usado para recarregar as listas de animais, médicos e
    #       agendamentos, garantindo que os dados exibidos estejam atualizados.
    def on_tab_visible(self, event):

        # Recarrega a lista de animais para atualizar as opções no
        #       combobox de seleção de animais.
        # Isso garante que novos animais adicionados em outras
        #       abas sejam refletidos nesta aba.
        self.recarregar_animais()

        # Recarrega a lista de médicos para atualizar as opções no
        #       combobox de seleção de médicos.
        # Isso assegura que novos médicos adicionados em outras
        #       abas sejam exibidos corretamente.
        self.recarregar_medicos()

        # Carrega os agendamentos existentes para preencher a
        #       lista de agendamentos na aba.
        # Isso atualiza a tabela de agendamentos com
        #       qualquer alteração feita.
        self.carregar_agendamentos()


    # Método 'recarregar_animais' é responsável por atualizar a
    #       lista de animais disponíveis no combobox.
    def recarregar_animais(self):

        # Obtém a lista atualizada de animais chamando uma função externa 'obter_lista_animais'.
        # A função retorna uma lista de tuplas com o ID e nome de cada animal.
        lista = obter_lista_animais()

        # Preenche o combobox de seleção de animais com os valores formatados como "Nome (ID:ID)".
        # Cada item é exibido no formato amigável ao usuário.
        self.combo_animal['values'] = [f"{name} (ID:{_id})" for (_id, name) in lista]

        # Se a lista de animais não estiver vazia:
        if lista:

            # Define o primeiro animal da lista como o valor selecionado no combobox.
            self.combo_animal.current(0)

            # Aciona o método '_quando_seleciona_animal' para carregar
            #       informações do animal selecionado.
            # Passa 'None' como evento porque essa chamada não vem
            #       diretamente de uma interação do usuário.
            self._quando_seleciona_animal(None)


    # Método 'recarregar_medicos' atualiza a lista de médicos disponíveis no combobox.
    def recarregar_medicos(self):

        # Obtém a lista atualizada de médicos chamando uma função
        #       externa 'obter_lista_medicos'.
        # A função retorna uma lista de tuplas contendo o ID e o nome de cada médico.
        lista = obter_lista_medicos()

        # Extrai apenas os nomes dos médicos da lista e armazena em uma nova lista.
        nomes = [m for (_id, m) in lista]

        # Atualiza os valores do combobox de médicos com a lista de nomes.
        self.combo_medico['values'] = nomes

        # Se a lista de nomes não estiver vazia:
        if nomes:

            # Define o primeiro médico da lista como o valor selecionado no combobox.
            self.combo_medico.current(0)


    # Método que é chamado ao selecionar uma linha na Treeview.
    def ao_selecionar_linha(self, event):

        # Obtém a seleção atual da Treeview.
        sel = self.tree.selection()

        # Verifica se nenhuma linha está selecionada; se for o
        #       caso, a função termina aqui.
        if not sel:
            return

        # Obtém os dados do item selecionado na Treeview.
        item = self.tree.item(sel[0])
        vals = item["values"]

        # Armazena o ID do agendamento selecionado para futuras
        #       operações (como edição ou exclusão).
        self.agenda_id_atual = vals[0]

        # Obtém o ID e o nome do animal associado ao agendamento.
        id_animal_str = vals[1]  # ID do animal como string.
        nome_animal = vals[2]  # Nome do animal.

        # Atualiza o valor do combobox do animal para exibir o
        #       nome e o ID do animal selecionado.
        self.combo_animal.set(f"{nome_animal} (ID:{id_animal_str})")

        # Chama o método `_quando_seleciona_animal` para
        #       atualizar as informações do dono.
        self._quando_seleciona_animal(None)

        # Obtém o nome do médico associado ao agendamento e
        #       atualiza o combobox do médico.
        nome_medico = vals[4]  # Nome do médico.
        self.combo_medico.set(nome_medico)

        # Armazena a data e hora do agendamento selecionado.
        # 'vals[5]' contém o valor da coluna correspondente à
        #       data/hora do agendamento.
        data_hora = vals[5]  # Exemplo: "25/12/2025 14:30".

        # Armazena o tipo de consulta do agendamento.
        # 'vals[6]' contém o tipo de consulta definido no
        #       agendamento (ex: "Consulta de Rotina").
        tipo = vals[6]  # Exemplo: "Consulta de Rotina".

        # Armazena o status do agendamento selecionado.
        # 'vals[7]' contém o status atual do agendamento (ex: "Pendente", "Confirmado").
        status = vals[7]  # Exemplo: "Pendente".

        # Armazena os produtos associados ao agendamento como string.
        # 'vals[8]' contém a lista de produtos usados, separados por vírgulas.
        produtos_str = vals[8]  # Exemplo: "Vacina A, Remédio X".

        # Limpa o campo de entrada onde será exibido o tipo de consulta.
        # Remove qualquer texto atualmente presente no campo de entrada.
        self.entry_tipo.delete(0, tk.END)

        # Insere o tipo de consulta do agendamento selecionado no campo de entrada.
        # Garante que o campo exiba o tipo correto com base na seleção.
        self.entry_tipo.insert(0, tipo)  # Exemplo: insere "Consulta de Rotina".

        # Verifica se há um valor para data e hora no agendamento selecionado.
        # 'data_hora' deve conter um valor no formato "dd/MM/yyyy HH:mm".
        if data_hora:

            try:

                # Divide a string de data e hora em duas partes: data e hora.
                # Exemplo: "25/12/2025 14:30" -> data_part = "25/12/2025", hora_part = "14:30".
                data_part, hora_part = data_hora.split(" ")

                # Divide a data em dia, mês e ano.
                # Exemplo: "25/12/2025" -> dia = "25", mes = "12", ano = "2025".
                dia, mes, ano = data_part.split("/")

                # Define o calendário para a data extraída.
                # Converte os valores de dia, mês e ano em inteiros antes de criar o objeto datetime.
                self.data_entry.set_date(datetime.datetime(int(ano), int(mes), int(dia)))

                # Limpa o campo de entrada de hora.
                # Garante que valores antigos sejam removidos antes de inserir novos.
                self.entry_hora.delete(0, tk.END)

                # Insere a hora extraída no campo de entrada de hora.
                # Exemplo: insere "14:30" no campo.
                self.entry_hora.insert(0, hora_part)

            except:

                # Caso ocorra um erro na extração ou formatação da
                #       data e hora, ignora silenciosamente.
                pass

        # Define o valor do combobox de status com base no status do agendamento.
        # Exemplo: "Pendente", "Confirmado", etc.
        self.combo_status.set(status)

        # Limpa o campo de entrada de produtos.
        # Remove qualquer lista de produtos previamente exibida.
        self.entry_produtos.delete(0, tk.END)

        # Insere a lista de produtos associados ao agendamento no campo de entrada.
        # Exemplo: insere "Vacina A, Remédio X".
        self.entry_produtos.insert(0, produtos_str)

    # Função para salvar ou atualizar o agendamento no banco de dados.
    def salvar(self):

        # Obtém o valor selecionado no combobox de animais.
        # Exemplo: "Rex (ID:12345)".
        sel = self.combo_animal.get()

        try:

            # Divide a string selecionada para extrair o ID do animal.
            # A string "(ID:12345)" será isolada e limpa, resultando em "12345".
            parte_id = sel.split("(ID:")[1]  # Obtém a parte após "(ID:".

            # Remove o parêntese de fechamento e espaços extras.
            animal_id_str = parte_id.replace(")", "").strip()

        except:

            # Caso ocorra um erro (por exemplo, se o formato da string não
            #       for o esperado), define o ID do animal como `None`.
            animal_id_str = None

        # Verifica se o ID do animal foi obtido corretamente.
        # Caso contrário, exibe uma mensagem de aviso e interrompe a execução da função.
        if not animal_id_str:

            # Alerta o usuário sobre a necessidade de selecionar um animal.
            messagebox.showwarning("Aviso", "Selecione um animal.")

            # Sai da função sem executar o restante do código.
            return

        # Obtém o nome do médico selecionado no combobox.
        # O método .get() retorna o texto selecionado, e .strip()
        #       remove quaisquer espaços em branco extras.
        nome_medico = self.combo_medico.get().strip()

        # Realiza uma busca no banco de dados de médicos com base no nome selecionado.
        doc_medico = db.medicos.find_one({"nome": nome_medico})

        # Verifica se o médico foi encontrado no banco de dados.
        # Se encontrado, armazena o ID do médico; caso contrário, define o ID como None.
        if doc_medico:

            # Obtém o ID do médico a partir do documento retornado.
            id_medico = doc_medico["_id"]

        else:

            # Define como None se nenhum médico correspondente for encontrado.
            id_medico = None

        # Obtém o tipo de consulta do campo de entrada.
        # .get() retorna o texto inserido pelo usuário, e .strip()
        #       remove espaços em branco extras.
        tipo_consulta = self.entry_tipo.get().strip()

        # Obtém a data selecionada no componente de calendário (DateEntry).
        # O método .get_date() retorna um objeto de data (datetime.date)
        #       representando a data escolhida.
        data_sel = self.data_entry.get_date()

        # Obtém o texto do campo de hora (entrada do usuário).
        # .get() retorna o texto inserido, e .strip() remove
        #       espaços em branco adicionais.
        hora_txt = self.entry_hora.get().strip()

        # Obtém o status selecionado no combobox de status.
        # .get() retorna o texto atualmente selecionado no componente.
        status = self.combo_status.get()

        # Obtém os produtos usados, informados no campo de entrada.
        # .get() retorna o texto inserido, e .strip() remove espaços em branco extras.
        produtos_str = self.entry_produtos.get().strip()

        # Verifica se o campo de hora está vazio ou não preenchido.
        # Caso esteja vazio, exibe uma mensagem de aviso e
        #       interrompe a execução da função.
        if not hora_txt:

            # Exibe um alerta para o usuário.
            messagebox.showwarning("Aviso", "Informe a hora (HH:MM).")

            # Interrompe a execução da função para evitar erros subsequentes.
            return

        # Tenta criar um objeto datetime com base na data selecionada e
        #       na hora informada pelo usuário.
        try:

            # Extrai o dia, mês e ano da data selecionada no calendário.
            dia = data_sel.day  # Obtém o dia como um inteiro.
            mes = data_sel.month  # Obtém o mês como um inteiro.
            ano = data_sel.year  # Obtém o ano como um inteiro.

            # Divide o texto da hora informada pelo usuário no formato "HH:MM" em partes.
            hh, mm = hora_txt.split(":")  # 'hh' para horas e 'mm' para minutos.

            # Cria um objeto datetime combinando data e hora.
            # Os valores de 'ano', 'mes', 'dia', 'hh' e 'mm' são usados para criar o timestamp.
            data_hora = datetime.datetime(ano, mes, dia, int(hh), int(mm))

        except:

            # Caso ocorra um erro (ex.: formato de hora inválido),
            #       exibe uma mensagem de erro.
            messagebox.showerror("Erro",
                                 "Formato de hora inválido (use HH:MM).")  # Alerta o usuário sobre o erro.

            # Encerra a execução da função para evitar continuar com dados inválidos.
            return

        # Processa a lista de produtos usados no agendamento.
        # Divide a string 'produtos_str' por vírgulas e remove espaços em
        #       branco extras em cada item.
        # Se 'produtos_str' estiver vazia, a lista resultante será vazia.
        produtos_usados = [p.strip() for p in produtos_str.split(",")] if produtos_str else []

        # Cria um dicionário 'doc' que representa o agendamento no banco de dados.
        doc = {

            # Associa o ID do animal ao agendamento, convertendo-o para ObjectId.
            "id_animal": ObjectId(animal_id_str),

            # Armazena o ID do médico associado ao agendamento (pode
            #       ser None se não houver médico).
            "id_medico": id_medico,

            # Registra a data e hora do agendamento, no formato datetime.
            "data_agendamento": data_hora,

            # Tipo de consulta, fornecido pelo usuário.
            "tipo_consulta": tipo_consulta,

            # Status do agendamento, como "Pendente", "Confirmado", etc.
            "status": status,

            # Lista de produtos utilizados no agendamento.
            "produtos_usados": produtos_usados,

            # ID do médico logado que realizou ou alterou o agendamento.
            "ultima_alteracao_por_id": self.medico_logado["_id"],

            # Nome do médico logado que realizou ou alterou o agendamento.
            "ultima_alteracao_por_nome": self.medico_logado["nome"],

            # Data e hora da última alteração no agendamento.
            "ultima_alteracao_em": datetime.datetime.now()
        }

        try:

            # Verifica se o agendamento atual já possui um ID associado.
            # Caso 'self.agenda_id_atual' seja None, o agendamento é
            #       novo e será inserido no banco.
            if self.agenda_id_atual is None:

                # Insere o dicionário 'doc' como um novo documento na coleção 'agenda'.
                db.agenda.insert_one(doc)

            else:

                # Atualiza o documento existente na coleção 'agenda' com base no ID atual.
                # Utiliza o método '$set' para substituir os campos pelo conteúdo de 'doc'.
                db.agenda.update_one({"_id": ObjectId(self.agenda_id_atual)}, {"$set": doc})

            # Exibe uma mensagem de sucesso ao usuário indicando que o
            #       agendamento foi salvo ou atualizado.
            messagebox.showinfo("Sucesso", "Agendamento salvo/atualizado!")

            # Recarrega a lista de agendamentos para refletir as alterações feitas.
            self.carregar_agendamentos()

            # Limpa os campos do formulário para permitir um novo cadastro ou edição.
            self.limpar_form()

        except Exception as e:

            # Caso ocorra qualquer erro durante a operação, exibe uma
            #       mensagem de erro detalhando o problema.
            messagebox.showerror("Erro",
                                 f"Não foi possível salvar:\n{e}")



    def carregar_agendamentos(self):

        # Remove todos os itens atualmente exibidos na Treeview, limpando a lista.
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Busca todos os agendamentos no banco de dados, ordenados pela data de agendamento.
        for ag in db.agenda.find().sort("data_agendamento", 1):

            # Obtém o documento do animal relacionado ao ID armazenado no agendamento.
            animal_doc = db.animais.find_one({"_id": ag["id_animal"]})

            # Se encontrado, obtém o nome do animal; caso contrário, usa uma string vazia.
            nome_animal = animal_doc.get("nome_animal", "") if animal_doc else ""

            # Obtém o ID do médico associado ao agendamento.
            id_medico = ag.get("id_medico")

            # Inicializa o nome do médico como uma string vazia.
            nome_medico = ""

            # Verifica se o ID do médico está presente no documento do agendamento.
            # Isso indica que o agendamento está associado a um médico específico.
            if id_medico:

                # Realiza uma consulta no banco de dados 'medicos'
                #       para encontrar o documento do médico
                # com o ID correspondente ao armazenado no agendamento.
                medico_doc = db.medicos.find_one({"_id": id_medico})

                # Após recuperar o documento do médico:
                # - Tenta obter o valor do campo "nome" do documento do médico.
                # - Caso o documento seja encontrado e o campo "nome" exista,
                #       atribui o nome ao `nome_medico`.
                # - Caso contrário (documento não encontrado ou campo ausente),
                #       `nome_medico` permanece como uma string vazia.
                nome_medico = medico_doc.get("nome", "") if medico_doc else ""

            # Obtém a data e hora do agendamento no formato armazenado no banco de dados.
            # Utiliza uma função `formatar_data_hora` para formatar o valor em
            #       uma string amigável ao usuário,
            # por exemplo, "dd/mm/aaaa HH:MM".
            data_str = formatar_data_hora(ag.get("data_agendamento"))

            # Recupera a lista de produtos utilizados no agendamento,
            #       armazenada no campo 'produtos_usados'.
            # Se o campo não existir ou estiver vazio, atribui uma
            #       lista vazia como valor padrão.
            produtos_list = ag.get("produtos_usados", [])

            # Combina os itens da lista 'produtos_list' em uma única string,
            #       separando os itens por vírgulas.
            # Exemplo: Se `produtos_list` for ['Vacina A', 'Medicamento B'], `
            #       produtos_str` será 'Vacina A, Medicamento B'.
            produtos_str = ", ".join(produtos_list)

            # Insere um novo item na Treeview com os valores obtidos do agendamento.
            self.tree.insert("", tk.END, values=(

                # O ID do agendamento é convertido para string e exibido na primeira coluna.
                str(ag["_id"]),

                # O ID do animal associado ao agendamento é convertido para
                #       string e exibido na segunda coluna.
                # Isso ajuda na identificação interna do animal no banco de dados.
                str(ag["id_animal"]),

                # O nome do animal é exibido na terceira coluna. Caso o animal
                #       não seja encontrado, o valor permanece vazio.
                nome_animal,

                # O ID do médico responsável pelo agendamento é exibido na quarta coluna.
                # Caso não exista um médico associado, a célula permanece vazia.
                str(id_medico) if id_medico else "",

                # O nome do médico responsável pelo agendamento é exibido na quinta coluna.
                # Caso o nome do médico não seja encontrado, a célula permanece vazia.
                nome_medico,

                # A data e hora do agendamento, formatadas, são exibidas na sexta coluna.
                data_str,

                # O tipo de consulta (por exemplo, "Vacinação", "Check-up") é
                #       exibido na sétima coluna.
                # Se o campo não estiver presente no banco de dados, uma
                #       string vazia é usada como valor padrão.
                ag.get("tipo_consulta", ""),

                # O status do agendamento (por exemplo, "Pendente", "Realizado") é
                #       exibido na oitava coluna.
                # Se o campo não estiver presente, a célula permanece vazia.
                ag.get("status", ""),

                # A lista de produtos utilizados, formatada como uma string
                #       separada por vírgulas, é exibida na nona coluna.
                produtos_str))


    def excluir(self):

        # Verifica se um agendamento foi selecionado. Caso contrário,
        #       exibe um aviso ao usuário.
        if not self.agenda_id_atual:
            messagebox.showwarning("Aviso",
                                   "Selecione um agendamento para excluir.")
            return

        # Solicita uma confirmação do usuário antes de excluir o agendamento.
        # Exibe uma caixa de diálogo com as opções "Sim" e "Não".
        resp = messagebox.askyesno("Confirmação",
                                   "Deseja excluir este agendamento?")

        # Verifica se o usuário confirmou a exclusão do agendamento.
        if resp:

            try:

                # Tenta excluir o agendamento correspondente no banco de dados usando o ID atual.
                db.agenda.delete_one({"_id": ObjectId(self.agenda_id_atual)})

                # Exibe uma mensagem de sucesso caso a exclusão seja realizada corretamente.
                messagebox.showinfo("Sucesso",
                                    "Agendamento excluído!")

                # Recarrega a lista de agendamentos para refletir a exclusão.
                self.carregar_agendamentos()

                # Limpa o formulário de entrada de dados para garantir que
                #       nenhum dado do agendamento excluído permaneça.
                self.limpar_form()

            except Exception as e:

                # Caso ocorra um erro durante a exclusão, exibe uma mensagem
                #       de erro detalhando o problema.
                messagebox.showerror("Erro",
                                     f"Não foi possível excluir:\n{e}")


    def limpar_form(self):

        # Reseta o ID do agendamento atual, indicando que não há
        #       nenhum agendamento selecionado para edição.
        self.agenda_id_atual = None

        # Verifica se há valores disponíveis no combobox de animais.
        if self.combo_animal['values']:

            # Seleciona o primeiro item do combobox de animais como padrão.
            self.combo_animal.current(0)

            # Atualiza o campo de dono baseado no animal selecionado.
            self._quando_seleciona_animal(None)

        # Verifica se há valores disponíveis no combobox de médicos.
        if self.combo_medico['values']:

            # Seleciona o primeiro item do combobox de médicos como padrão.
            self.combo_medico.current(0)

        # Limpa o campo de texto para o tipo de consulta.
        self.entry_tipo.delete(0, tk.END)

        # Define a data de agendamento como a data atual.
        self.data_entry.set_date(datetime.datetime.now())

        # Limpa o campo de hora de agendamento e define o valor padrão como "09:00".
        self.entry_hora.delete(0, tk.END)
        self.entry_hora.insert(0, "09:00")

        # Seleciona o primeiro item do combobox de status como padrão.
        self.combo_status.current(0)

        # Limpa o campo de produtos utilizados.
        self.entry_produtos.delete(0, tk.END)



    # Método '_quando_seleciona_animal' é chamado quando um animal é selecionado no combobox.
    # Ele atualiza o rótulo do dono correspondente ao animal selecionado.
    def _quando_seleciona_animal(self, event):

        # Obtém o texto selecionado no combobox de animais.
        sel = self.combo_animal.get()

        try:

            # Tenta extrair o ID do animal do texto selecionado.
            # O texto esteja no formato "Nome do Animal (ID:123456)".
            parte_id = sel.split("(ID:")[1]  # Obtém a parte após "(ID:".

            # Remove o fechamento ")" e espaços extras.
            animal_id_str = parte_id.replace(")", "").strip()

        except:

            # Se ocorrer algum erro na extração do ID, define o ID como None.
            animal_id_str = None

        # Se o ID do animal não for válido (None ou vazio):
        if not animal_id_str:

            # Atualiza o rótulo do dono para indicar que nenhum dono está associado.
            self.label_dono.config(text="---")

            # Finaliza o método, pois não há ID válido para consultar.
            return

        # Busca no banco de dados o documento correspondente ao animal pelo ID.
        animal_doc = db.animais.find_one({"_id": ObjectId(animal_id_str)})

        # Verifica se o documento do animal não foi encontrado no banco de dados.
        if not animal_doc:

            # Atualiza o rótulo do dono para "---", indicando que o
            #       animal não possui um dono associado.
            self.label_dono.config(text="---")

            # Finaliza o método, pois não há mais informações para processar.
            return

        # Obtém o ID do dono associado ao animal a partir do documento.
        dono_id = animal_doc.get("id_dono")

        # Define um nome padrão para o dono caso ele não seja encontrado.
        dono_nome = "Sem Dono"

        # Verifica se o animal possui um ID de dono válido.
        if dono_id:

            # Busca no banco de dados o documento do cliente
            #       correspondente ao ID do dono.
            dono_doc = db.clientes.find_one({"_id": dono_id})

            # Se o cliente for encontrado, obtém o nome.
            # Caso contrário, usa "Desconhecido".
            dono_nome = dono_doc.get("nome", "") if dono_doc else "Desconhecido"

        # Atualiza o rótulo do dono com o nome obtido (ou "Sem Dono"
        #       caso não tenha dono).
        self.label_dono.config(text=dono_nome)


# Definição da função para formatar apenas a data, sem incluir a hora.
def formatar_data_somente(data):

    # Verifica se a variável 'data' está vazia ou é None.
    if not data:

        # Retorna uma string vazia se não houver data.
        return ""

    # Se a data existir, formata para o formato dia/mês/ano.
    return data.strftime("%d/%m/%Y")


# Definição da função para formatar a data e a hora em um formato específico.
def formatar_data_hora(data):

    # Verifica se a variável 'data' está vazia ou é None.
    if not data:

        # Retorna uma string vazia se não houver data.
        return ""

    # Se a data existir, formata para o formato dia/mês/ano e hora:minuto.
    return data.strftime("%d/%m/%Y %H:%M")


# Definição da função para obter uma lista de animais do banco de dados.
def obter_lista_animais():

    # Cria uma lista vazia para armazenar os resultados.
    resultados = []

    # Loop que percorre cada animal encontrado na coleção 'animais' do MongoDB.
    for ani in db.animais.find():

        # Adiciona na lista um tuplo com o ID e o nome do animal, se
        #       não tiver nome, usa "Sem Nome".
        resultados.append((ani["_id"], ani.get("nome_animal", "Sem Nome")))

    # Retorna a lista de resultados.
    return resultados



# Definição da função para obter uma lista de médicos do banco de dados.
def obter_lista_medicos():

    # Cria uma lista vazia para armazenar os resultados.
    resultados = []

    # Loop que percorre cada médico encontrado na coleção 'medicos' do MongoDB.
    for med in db.medicos.find():

        # Adiciona na lista um tuplo com o ID e o nome do médico, se
        #       não tiver nome, usa "Sem Nome".
        resultados.append((med["_id"], med.get("nome", "Sem Nome")))

    # Retorna a lista de resultados.
    return resultados


# ---------------------------------------------------------------------
# JanelaHistoricoProduto
# ---------------------------------------------------------------------
class JanelaHistoricoProduto(tk.Toplevel):

    """
    Classe para exibir uma janela com o histórico de transações de um produto específico.
    """

    def __init__(self, parent, produto_id_str, *args, **kwargs):

        """
        Inicializa a janela de histórico de transações do produto.

        Args:
            parent (tk.Widget): Janela ou frame pai para esta janela.
            produto_id_str (str): ID do produto para carregar o histórico.
            *args: Argumentos adicionais para a classe base.
            **kwargs: Argumentos nomeados adicionais para a classe base.
        """

        # Chama o inicializador da classe pai ('tk.Toplevel') para configurar a janela.
        super().__init__(parent, *args, **kwargs)

        # Define o título da janela.
        self.title("Histórico de Transações do Produto")

        # Configura as dimensões da janela para 700x400 pixels.
        self.geometry("700x400")

        # Centraliza a janela na tela.
        centralizar_janela(self, 700, 400)

        # Armazena o ID do produto como string para uso posterior.
        self.produto_id_str = produto_id_str

        # Adiciona um título à janela.
        # 'text="Histórico de Transações"' define o texto exibido no título.
        # 'font=("Arial", 16, "bold")' configura a fonte
        #       como Arial, tamanho 16, em negrito.
        tk.Label(self,
                text="Histórico de Transações",
                font=("Arial", 16, "bold")

        # 'pady=10' adiciona 10 pixels de espaço vertical
        #       acima e abaixo do rótulo.
        ).pack(pady=10)

        # Cria um LabelFrame para conter a tabela de transações.
        # 'text="Transações"' define o título exibido no frame.
        # 'padx=10' e 'pady=10' adicionam 10 pixels de margem
        #       horizontal e vertical interna ao frame.
        frame_tabela = tk.LabelFrame(self,
                                    text="Transações",
                                    padx=10,
                                    pady=10)

        # Posiciona o frame na janela.
        # 'fill="both"' faz o frame expandir para preencher tanto a
        #       largura quanto a altura disponíveis.
        # 'expand=True' permite que o frame ajuste seu tamanho conforme a janela é redimensionada.
        # 'padx=10' e 'pady=10' adicionam 10 pixels de margem externa ao redor do frame.
        frame_tabela.pack(fill="both",
                          expand=True,
                          padx=10,
                          pady=10)

        # Define as colunas da tabela de transações.
        # As colunas incluem "Data/Hora", "Tipo", "Quantidade", e "Observações".
        colunas = ("data", "tipo", "quantidade", "observacoes")

        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria um Treeview dentro do frame de tabela para exibir os dados.
        # 'columns=colunas' define as colunas da tabela.
        # 'show="headings"' exibe apenas os cabeçalhos das colunas,
        #       sem uma coluna de índice.
        self.tree_transacoes = ttk.Treeview(frame_tabela,
                                            columns=colunas,
                                            show="headings")

        # Configura o cabeçalho da coluna "data".
        # 'text="Data/Hora"' define o título exibido no cabeçalho da coluna.
        self.tree_transacoes.heading("data", text="Data/Hora")

        # Configura o cabeçalho da coluna "tipo".
        self.tree_transacoes.heading("tipo", text="Tipo")

        # Configura o cabeçalho da coluna "quantidade".
        self.tree_transacoes.heading("quantidade", text="Quantidade")

        # Configura o cabeçalho da coluna "observacoes".
        self.tree_transacoes.heading("observacoes", text="Observações")

        # Configura a largura e o alinhamento das colunas.
        # 'width=120' define a largura da coluna "data" como 120 pixels.
        # 'anchor="center"' alinha o texto ao centro.
        self.tree_transacoes.column("data", width=120, anchor="center")

        # Configura a coluna "tipo" com largura de 80 pixels e alinhamento central.
        self.tree_transacoes.column("tipo", width=80, anchor="center")

        # Configura a coluna "quantidade" com largura de 100 pixels e alinhamento central.
        self.tree_transacoes.column("quantidade", width=100, anchor="center")

        # Configura a coluna "observacoes" com largura de 300 pixels e alinhamento à esquerda.
        self.tree_transacoes.column("observacoes", width=300, anchor="w")

        # Posiciona o Treeview dentro do frame de tabela.
        # 'side="left"' alinha o Treeview à esquerda do frame.
        # 'fill="both"' faz com que o Treeview preencha a largura e a
        #       altura disponíveis.
        # 'expand=True' permite que o Treeview expanda proporcionalmente
        #       quando a janela for redimensionada.
        self.tree_transacoes.pack(side="left", fill="both", expand=True)

        # Cria uma barra de rolagem vertical para o Treeview.
        # 'orient="vertical"' especifica que a barra de rolagem é vertical.
        # 'command=self.tree_transacoes.yview' conecta a barra de rolagem ao
        #       movimento vertical do Treeview.
        scrollbar = ttk.Scrollbar(  frame_tabela,
                                    orient="vertical",
                                    command=self.tree_transacoes.yview)

        # Configura o Treeview para usar a barra de rolagem criada.
        # 'yscrollcommand=scrollbar.set' conecta o movimento vertical do
        #       Treeview à barra de rolagem.
        self.tree_transacoes.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do frame.
        # 'side="right"' alinha a barra de rolagem à direita do frame.
        # 'fill="y"' faz com que a barra de rolagem preencha
        #       toda a altura disponível.
        scrollbar.pack(side="right", fill="y")

        # Chama o método para carregar as transações do banco de
        #       dados e exibi-las no Treeview.
        self.carregar_transacoes()

    def carregar_transacoes(self):

        # Remove todas as entradas existentes no Treeview antes de
        #       carregar novas transações.
        for item in self.tree_transacoes.get_children():
            self.tree_transacoes.delete(item)

        try:

            # Recupera as transações do banco de dados MongoDB filtrando pelo ID do produto.
            # 'find({"id_produto": ObjectId(self.produto_id_str)})' busca
            #       apenas transações associadas ao produto atual.
            # 'sort("data", -1)' ordena as transações por data em ordem decrescente.
            cursor = db.estoque_transacoes.find({"id_produto": ObjectId(self.produto_id_str)}).sort("data", -1)

            # Itera sobre cada transação encontrada no cursor.
            for trans in cursor:

                # Obtém a data da transação e formata para o padrão "dd/MM/yyyy HH:mm".
                dt = trans.get("data")
                data_str = dt.strftime("%d/%m/%Y %H:%M") if dt else ""

                # Obtém o tipo de transação, como "Entrada" ou "Saída".
                tipo = trans.get("tipo", "")

                # Obtém a quantidade envolvida na transação.
                qtd = trans.get("quantidade", 0)

                # Obtém observações associadas à transação.
                obs = trans.get("observacoes", "")

                # Insere os dados da transação no Treeview.
                # 'values=(data_str, tipo, qtd, obs)' define as colunas
                #       exibidas na linha correspondente.
                self.tree_transacoes.insert("",
                                            tk.END,
                                            values=(data_str, tipo, qtd, obs))

        except Exception as e:

            # Exibe uma mensagem de erro caso haja problemas ao
            #       recuperar ou processar as transações.
            messagebox.showerror("Erro",
                                 f"Não foi possível carregar transações:\n{e}")



# ---------------------------------------------------------------------
# Janela para exibir o histórico de atendimentos de um animal
# ---------------------------------------------------------------------
class JanelaHistoricoAtendimentos(tk.Toplevel):

    # Método inicializador da classe 'JanelaHistoricoAtendimentos',
    #       responsável por configurar a nova janela.
    # Este método é chamado automaticamente quando uma nova
    #       instância desta classe é criada.
    def __init__(self, parent, animal_id, *args, **kwargs):

        # 'self' representa a instância atual da classe.
        # É usado para acessar atributos e métodos da própria classe.

        # 'parent' é a janela ou widget pai que cria esta nova janela.
        # Permite que esta janela seja associada a outra já
        #       existente (janela principal ou outro contexto).

        # 'animal_id' é o identificador único do animal cujo
        #       histórico de atendimentos será exibido.
        # Esse parâmetro será usado para buscar e carregar os
        #       dados relacionados no banco de dados.

        # '*args' é usado para aceitar um número variável de
        #       argumentos posicionais adicionais.
        # Esses argumentos podem ser passados sem necessidade de
        #       nomeá-los explicitamente.

        # '**kwargs' aceita argumentos nomeados adicionais que
        #       não foram previamente definidos.
        # Isso permite flexibilidade para configurar opções extras na janela.

        # Chama o inicializador da classe pai ('tk.Toplevel').
        # Isso garante que as funcionalidades padrão de uma janela 'Toplevel'
        #       sejam aplicadas corretamente.
        super().__init__(parent, *args, **kwargs)

        # Define o título da janela que será exibido na barra superior.
        # Aqui, indica que o conteúdo da janela é relacionado ao "Histórico
        #       de Atendimentos do Animal".
        self.title("Histórico de Atendimentos do Animal")

        # Define as dimensões da janela para 800 pixels de
        #       largura por 500 pixels de altura.
        # Essas dimensões determinam o tamanho inicial da janela na tela.
        self.geometry("800x500")

        # Centraliza a janela na tela do usuário.
        # A função 'centralizar_janela' utiliza as dimensões da janela e
        #       da tela para calcular a posição ideal.
        centralizar_janela(self, 800, 500)

        # Armazena o identificador único do animal cujo histórico será carregado.
        # O 'animal_id' é passado como parâmetro e é necessário
        #       para buscar os dados no banco de dados.
        self.animal_id = animal_id

        # Inicializa uma lista vazia que armazenará os registros de
        #       históricos de atendimento.
        # Essa lista será preenchida com os dados do banco de
        #       dados posteriormente.
        self.historicos = []

        # Cria um rótulo para o título da janela.
        # O texto "Histórico de Atendimentos" será exibido em uma
        #       fonte Arial, tamanho 16, em negrito.
        # O método 'pack' posiciona o rótulo na parte superior da janela.
        # 'pady=10' adiciona 10 pixels de margem vertical acima e
        #       abaixo do rótulo para separação visual.
        tk.Label(self,
                 text="Histórico de Atendimentos",
                 font=("Arial", 16, "bold")).pack(pady=10)

        # Cria um LabelFrame (um frame com um título) para exibir os atendimentos registrados.
        # O título "Atendimentos Registrados" será exibido no topo do LabelFrame.
        # 'padx=10' adiciona 10 pixels de margem interna horizontal dentro do frame.
        # 'pady=10' adiciona 10 pixels de margem interna vertical dentro do frame.
        frame_tabela = tk.LabelFrame(self,
                                     text="Atendimentos Registrados",
                                     padx=10, pady=10)

        # Posiciona o LabelFrame na janela principal.
        # 'fill="both"' permite que o frame preencha a largura e a altura disponíveis.
        # 'expand=True' faz com que o frame se expanda proporcionalmente
        #       quando a janela é redimensionada.
        # 'padx=10' adiciona 10 pixels de margem horizontal externa ao redor do frame.
        # 'pady=10' adiciona 10 pixels de margem vertical externa ao redor do frame.
        frame_tabela.pack(fill="both",
                          expand=True,
                          padx=10,
                          pady=10)

        # Define as colunas que serão exibidas na tabela de histórico.
        # Cada coluna representa um campo relevante do histórico de atendimentos.
        colunas = ("data_atendimento", "valor_consulta", "total_produtos", "total_geral", "medico", "observacoes")


        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria uma Treeview para exibir os dados do histórico de atendimentos.
        # 'columns=colunas' define as colunas que serão exibidas.
        # 'show="headings"' exibe apenas os cabeçalhos das colunas,
        #       sem uma coluna de índice padrão.
        self.tree_historico = ttk.Treeview(frame_tabela,
                                           columns=colunas,
                                           show="headings")

        # Configura o cabeçalho da coluna "data_atendimento".
        # 'text="Data/Hora"' define o texto que será exibido no cabeçalho da coluna.
        self.tree_historico.heading("data_atendimento",
                                    text="Data/Hora")

        # Configura o cabeçalho da coluna "valor_consulta".
        # 'text="Consulta (R$)"' define o texto exibido, indicando o
        #       valor da consulta em reais.
        self.tree_historico.heading("valor_consulta", text="Consulta (R$)")

        # Configura o cabeçalho da coluna "total_produtos".
        # 'text="Prod. (R$)"' define o texto exibido, indicando o
        #       total gasto com produtos em reais.
        self.tree_historico.heading("total_produtos", text="Prod. (R$)")

        # Configura o cabeçalho da coluna "total_geral".
        # 'text="Total (R$)"' define o texto exibido, indicando o
        #       total geral do atendimento em reais.
        self.tree_historico.heading("total_geral", text="Total (R$)")

        # Configura o cabeçalho da coluna "medico".
        # 'text="Médico"' define o texto exibido, indicando o
        #       médico responsável pelo atendimento.
        self.tree_historico.heading("medico", text="Médico")

        # Configura o cabeçalho da coluna "observacoes".
        # 'text="Observações"' define o texto exibido, indicando as
        #       observações registradas para o atendimento.
        self.tree_historico.heading("observacoes", text="Observações")

        # Configura a coluna "data_atendimento".
        # 'width=120' define a largura da coluna em pixels.
        # 'anchor="center"' alinha o texto da coluna ao centro.
        self.tree_historico.column("data_atendimento", width=120, anchor="center")

        # Configura a coluna "valor_consulta".
        # 'width=100' define a largura da coluna em pixels.
        # 'anchor="center"' alinha o texto da coluna ao centro.
        self.tree_historico.column("valor_consulta", width=100, anchor="center")

        # Configura a coluna "total_produtos".
        # 'width=100' define a largura da coluna em pixels.
        # 'anchor="center"' alinha o texto da coluna ao centro.
        self.tree_historico.column("total_produtos", width=100, anchor="center")

        # Configura a coluna "total_geral".
        # 'width=100' define a largura da coluna em pixels.
        # 'anchor="center"' alinha o texto da coluna ao centro.
        self.tree_historico.column("total_geral", width=100, anchor="center")

        # Configura a coluna "medico".
        # 'width=150' define a largura da coluna em pixels.
        # 'anchor="center"' alinha o texto da coluna ao centro.
        self.tree_historico.column("medico", width=150, anchor="center")

        # Configura a coluna "observacoes".
        # 'width=250' define a largura da coluna em pixels.
        # 'anchor="w"' alinha o texto da coluna à esquerda, pois as
        #       observações podem ser mais longas.
        self.tree_historico.column("observacoes", width=250, anchor="w")

        # Posiciona o Treeview (tabela) no lado esquerdo do frame.
        # 'side="left"' coloca o Treeview à esquerda dentro do frame.
        # 'fill="both"' permite que o Treeview preencha todo o espaço
        #       disponível, tanto em largura quanto em altura.
        # 'expand=True' faz com que o Treeview expanda quando a janela é redimensionada.
        self.tree_historico.pack(side="left",
                                 fill="both",
                                 expand=True)

        # Cria uma barra de rolagem vertical para o Treeview.
        # 'orient="vertical"' define a orientação da barra como vertical.
        # 'command=self.tree_historico.yview' conecta a barra de
        #       rolagem ao Treeview, permitindo rolar os dados.
        scrollbar = ttk.Scrollbar(frame_tabela,
                                  orient="vertical",
                                  command=self.tree_historico.yview)

        # Conecta a barra de rolagem ao Treeview.
        # 'yscrollcommand=scrollbar.set' sincroniza a barra de rolagem com o Treeview.
        self.tree_historico.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do frame.
        # 'side="right"' alinha a barra de rolagem à direita dentro do frame.
        # 'fill="y"' faz com que a barra de rolagem preencha toda a altura do frame.
        scrollbar.pack(side="right", fill="y")

        # Adiciona um evento para capturar duplo clique em um item do Treeview.
        # '<Double-1>' é o evento de duplo clique do botão esquerdo do mouse.
        # 'self.abrir_detalhes' é o método que será executado quando o evento ocorrer.
        self.tree_historico.bind("<Double-1>", self.abrir_detalhes)

        # Carrega os dados do histórico no Treeview ao iniciar a janela.
        self.carregar_historico()


    def carregar_historico(self):

        # Limpa todos os itens atuais do Treeview para garantir
        #       que os dados exibidos sejam atualizados.
        # 'self.tree_historico.get_children()' retorna todos os
        #       itens atualmente exibidos no Treeview.
        # 'self.tree_historico.delete(item)' remove cada item do Treeview.
        for item in self.tree_historico.get_children():
            self.tree_historico.delete(item)

        # Limpa a lista interna de históricos, garantindo que
        #       apenas os dados mais recentes sejam mantidos.
        self.historicos.clear()

        # Busca o documento do animal no banco de dados usando o ID do animal.
        # 'db.animais.find_one()' retorna o documento do animal se ele for encontrado.
        animal_doc = db.animais.find_one({"_id": ObjectId(self.animal_id)})

        # Verifica se o animal foi encontrado no banco de dados.
        # Se não for encontrado, exibe uma mensagem de erro para o usuário.
        if not animal_doc:
            messagebox.showerror("Erro",
                                 "Animal não encontrado no banco de dados.")
            return

        # Obtém a lista de históricos do documento do animal.
        # 'animal_doc.get("historicos", [])' retorna a lista de históricos do
        #       animal ou uma lista vazia, caso não exista o campo "historicos".
        historicos = animal_doc.get("historicos", [])

        # Itera sobre cada histórico encontrado na lista de históricos do animal.
        for h in historicos:

            # Verifica se o tipo do histórico é "Atendimento".
            # Se o tipo não for "Atendimento", a iteração continua para o
            #       próximo item, ignorando o atual.
            if h.get("tipo") != "Atendimento":
                continue

            # Adiciona o histórico à lista local 'self.historicos', que
            #       será usada para exibição no Treeview.
            self.historicos.append(h)

        # Itera sobre a lista de históricos do animal.
        # 'enumerate(self.historicos)' retorna o índice e o item
        #       atual em cada iteração.
        for i, h in enumerate(self.historicos):

            # Formata a data e hora do atendimento, convertendo o valor
            #       armazenado no campo "data" para o formato desejado.
            # 'formatar_data_hora(h.get("data"))' garante que a data
            #       será exibida no formato legível.
            data_str = formatar_data_hora(h.get("data"))

            # Obtém o valor da consulta do histórico. Caso não
            #       exista, retorna 0.0 como padrão.
            val_cons = h.get("valor_consulta", 0.0)

            # Obtém o valor total dos produtos usados no atendimento.
            #       Caso não exista, retorna 0.0 como padrão.
            val_prod = h.get("total_produtos", 0.0)

            # Calcula o valor total geral do atendimento (consulta + produtos).
            #       Caso não exista, retorna 0.0 como padrão.
            val_geral = h.get("total_geral", 0.0)

            # Obtém o nome do médico responsável pelo atendimento.
            #       Se não existir, retorna uma string vazia.
            medico_nome = h.get("realizado_por_nome", "")

            # Obtém as observações do atendimento, caso existam.
            # Se não houver, retorna uma string vazia.
            obs = h.get("observacoes", "")

            # Insere os dados do atendimento na Treeview para exibição ao usuário.
            # 'iid=str(i)' usa o índice atual como identificador único para o item na Treeview.
            # 'values' é uma tupla que contém as informações a serem exibidas nas colunas.
            # Cada campo é formatado ou ajustado para exibição apropriada.
            self.tree_historico.insert(

                "",  # Insere como um item "filho" direto da raiz da Treeview.
                tk.END,  # Adiciona o item no final da lista existente.
                iid=str(i),  # Define um identificador único para o item usando o índice da iteração.
                values=(
                    data_str,  # Data e hora do atendimento formatadas.
                    f"{val_cons:.2f}",  # Valor da consulta formatado com duas casas decimais.
                    f"{val_prod:.2f}",  # Valor dos produtos formatado com duas casas decimais.
                    f"{val_geral:.2f}",  # Valor total geral formatado com duas casas decimais.
                    medico_nome,  # Nome do médico responsável.
                    obs  # Observações do atendimento.
                )
            )


    # Define o método para abrir os detalhes de um
    #       atendimento selecionado na Treeview.
    def abrir_detalhes(self, event):

        # Obtém os itens selecionados na Treeview.
        # 'self.tree_historico.selection()' retorna uma tupla com os
        #       identificadores (iid) dos itens selecionados.
        selection = self.tree_historico.selection()

        # Verifica se nenhum item foi selecionado.
        # Se 'selection' estiver vazio, a função simplesmente retorna sem fazer nada.
        if not selection:
            return

        # Obtém o primeiro identificador (iid) da seleção.
        # Neste caso, como apenas um item pode ser selecionado,
        #       pegamos o primeiro elemento da tupla.
        iid = selection[0]

        # Converte o identificador (iid) para um índice numérico.
        # Como usamos índices numéricos como iid ao inserir os itens na
        #       Treeview, podemos convertê-lo diretamente.
        idx = int(iid)

        # Localiza o registro correspondente ao índice
        # ``selecionado na lista de históricos.
        # 'self.historicos' contém a lista de atendimentos do animal,
        #       onde 'idx' representa a posição do atendimento selecionado.
        registro = self.historicos[idx]

        # Cria uma nova janela para exibir os detalhes do atendimento selecionado.
        # 'JanelaDetalheHistorico' é a classe responsável por exibir essas informações.
        # Passa 'self' como a janela pai, e 'registro' como os
        #       dados do atendimento a serem exibidos.
        JanelaDetalheHistorico(self, registro)


# Define a classe 'JanelaDetalheHistorico', que é responsável por
#       exibir os detalhes de um atendimento específico.
class JanelaDetalheHistorico(tk.Toplevel):

    # Método inicializador da classe.
    # 'parent' é a janela que abriu essa (geralmente a janela anterior).
    # 'registro' contém os dados do atendimento a serem exibidos.
    # '*args' e '**kwargs' permitem que parâmetros adicionais sejam
    #       passados para a classe pai ('tk.Toplevel').
    def __init__(self, parent, registro, *args, **kwargs):

        # Chama o inicializador da classe pai ('tk.Toplevel'), garantindo
        #       que todas as configurações padrão sejam aplicadas.
        super().__init__(parent, *args, **kwargs)

        # Define o título da janela para "Detalhes do Atendimento".
        self.title("Detalhes do Atendimento")

        # Define o tamanho da janela para 600x400 pixels.
        self.geometry("600x600")

        # Centraliza a janela na tela, chamando a função 'centralizar_janela'.
        # Passa as dimensões desejadas (600x400) como argumentos.
        centralizar_janela(self, 600, 600)

        # Obtém a data do atendimento a partir do registro.
        # Se o campo 'data' não estiver presente, será retornado 'None'.
        data_atend = registro.get("data")

        # Formata a data/hora para o padrão desejado (dd/mm/yyyy hh:mm).
        # Se 'data_atend' for nulo, 'data_str' será uma string vazia.
        data_str = formatar_data_hora(data_atend) if data_atend else ""

        # Obtém o nome do médico que realizou o atendimento.
        # Se o campo 'realizado_por_nome' não estiver presente, retorna uma string vazia.
        medico_nome = registro.get("realizado_por_nome", "")

        # Obtém o valor da consulta a partir do registro.
        # Se 'valor_consulta' não estiver presente, será atribuído o valor padrão de 0.0.
        valor_consulta = registro.get("valor_consulta", 0.0)

        # Obtém o valor total dos produtos utilizados no atendimento.
        # Se 'total_produtos' não estiver presente, será atribuído 0.0.
        total_produtos = registro.get("total_produtos", 0.0)

        # Obtém o valor total geral do atendimento, incluindo a consulta e os produtos.
        # Se 'total_geral' não estiver presente, será atribuído 0.0.
        total_geral = registro.get("total_geral", 0.0)

        # Obtém as observações registradas no atendimento.
        # Se 'observacoes' não estiver presente, será atribuída uma string vazia.
        observacoes = registro.get("observacoes", "")

        # Cria um frame para agrupar as informações principais do atendimento.
        # 'padx=10' adiciona 10 pixels de margem horizontal interna ao frame.
        # 'pady=10' adiciona 10 pixels de margem vertical interna ao frame.
        frame_info = tk.Frame(self, padx=10, pady=10)

        # Posiciona o frame na parte superior da janela.
        # 'side="top"' posiciona o frame na parte superior.
        # 'fill="x"' faz o frame preencher toda a largura disponível na janela.
        frame_info.pack(side="top", fill="x")

        # Cria um rótulo para exibir a data e hora do atendimento.
        # 'text=f"Data/Hora: {data_str}"' define o texto exibido
        #       com a data e hora formatadas.
        # 'font=("Arial", 12, "bold")' configura a fonte com
        #       tamanho 12 e estilo negrito para destaque.
        lbl_data = tk.Label(frame_info,
                            text=f"Data/Hora: {data_str}",
                            font=("Arial", 12, "bold"))

        # Posiciona o rótulo no frame de informações.
        # 'row=0' posiciona o rótulo na primeira linha do grid.
        # 'column=0' posiciona o rótulo na primeira coluna do grid.
        # 'sticky="w"' alinha o rótulo à esquerda da célula do grid.
        # 'padx=5' adiciona 5 pixels de margem horizontal externa ao rótulo.
        # 'pady=3' adiciona 3 pixels de margem vertical externa ao rótulo.
        lbl_data.grid(row=0, column=0, sticky="w", padx=5, pady=3)

        # Cria um rótulo para exibir o nome do médico responsável pelo atendimento.
        # 'text=f"Médico: {medico_nome}"' exibe o nome do médico
        #       extraído do registro do atendimento.
        # 'font=("Arial", 12, "bold")' configura o texto do rótulo em
        #       fonte Arial, tamanho 12, com estilo negrito.
        lbl_medico = tk.Label(frame_info,
                              text=f"Médico: {medico_nome}",
                              font=("Arial", 12, "bold"))

        # Posiciona o rótulo no frame de informações.
        # 'row=1' posiciona o rótulo na segunda linha do grid.
        # 'column=0' posiciona o rótulo na primeira coluna do grid.
        # 'sticky="w"' alinha o rótulo à esquerda da célula do grid.
        # 'padx=5' adiciona 5 pixels de margem horizontal externa ao rótulo.
        # 'pady=3' adiciona 3 pixels de margem vertical externa ao rótulo.
        lbl_medico.grid(row=1, column=0, sticky="w", padx=5, pady=3)

        # Cria um rótulo para exibir o valor da consulta.
        # 'text=f"Valor da Consulta: R$ {valor_consulta:.2f}"' exibe o
        #       valor da consulta formatado com duas casas decimais.
        # 'font=("Arial", 12)' configura o texto do rótulo em fonte Arial,
        #       tamanho 12, sem negrito para diferenciação visual.
        lbl_valor = tk.Label(frame_info,
                             text=f"Valor da Consulta: R$ {valor_consulta:.2f}",
                             font=("Arial", 12))

        # Posiciona o rótulo no frame de informações.
        # 'row=2' posiciona o rótulo na terceira linha do grid.
        # 'column=0' posiciona o rótulo na primeira coluna do grid.
        # 'sticky="w"' alinha o rótulo à esquerda da célula do grid.
        # 'padx=5' adiciona 5 pixels de margem horizontal externa ao rótulo.
        # 'pady=3' adiciona 3 pixels de margem vertical externa ao rótulo.
        lbl_valor.grid(row=2, column=0, sticky="w", padx=5, pady=3)

        # Cria um rótulo para exibir o total de produtos utilizados no atendimento.
        # 'text=f"Total Produtos: R$ {total_produtos:.2f}"' exibe o valor
        #       total dos produtos formatado com duas casas decimais.
        # 'font=("Arial", 12)' configura o texto do rótulo em
        #       fonte Arial, tamanho 12, sem negrito.
        lbl_produtos = tk.Label(frame_info,
                                text=f"Total Produtos: R$ {total_produtos:.2f}",
                                font=("Arial", 12))

        # Posiciona o rótulo no frame de informações.
        # 'row=3' posiciona o rótulo na quarta linha do grid.
        # 'column=0' posiciona o rótulo na primeira coluna do grid.
        # 'sticky="w"' alinha o rótulo à esquerda da célula do grid.
        # 'padx=5' adiciona 5 pixels de margem horizontal externa ao rótulo.
        # 'pady=3' adiciona 3 pixels de margem vertical externa ao rótulo.
        lbl_produtos.grid(row=3, column=0, sticky="w", padx=5, pady=3)

        # Cria um rótulo para exibir o valor total geral do atendimento.
        # 'text=f"Total Geral: R$ {total_geral:.2f}"' exibe o valor total
        #       geral (consulta + produtos) formatado com duas casas decimais.
        # 'font=("Arial", 12, "bold")' configura o texto do rótulo em
        #       fonte Arial, tamanho 12, com estilo negrito para dar
        #       destaque ao total geral.
        lbl_geral = tk.Label(frame_info,
                             text=f"Total Geral: R$ {total_geral:.2f}",
                             font=("Arial", 12, "bold"))

        # Posiciona o rótulo no frame de informações.
        # 'row=4' posiciona o rótulo na quinta linha do grid.
        # 'column=0' posiciona o rótulo na primeira coluna do grid.
        # 'sticky="w"' alinha o rótulo à esquerda da célula do grid.
        # 'padx=5' adiciona 5 pixels de margem horizontal externa ao rótulo.
        # 'pady=3' adiciona 3 pixels de margem vertical externa ao rótulo.
        lbl_geral.grid(row=4, column=0, sticky="w", padx=5, pady=3)

        # Cria um rótulo para indicar o campo de observações.
        # 'text="Observações:"' define o texto do rótulo como "Observações".
        # 'font=("Arial", 12, "bold")' configura o texto do rótulo em
        #       fonte Arial, tamanho 12, com estilo negrito para dar destaque.
        tk.Label(frame_info, text="Observações:", font=("Arial", 12, "bold")).grid(

            # Posiciona o rótulo no frame de informações.
            # 'row=5' posiciona o rótulo na sexta linha do grid.
            # 'column=0' posiciona o rótulo na primeira coluna do grid.
            # 'sticky="w"' alinha o rótulo à esquerda da célula do grid.
            # 'padx=5' adiciona 5 pixels de margem horizontal externa ao rótulo.
            # 'pady=(10,2)' adiciona 10 pixels de margem superior e 2
            #       pixels de margem inferior ao rótulo.
            row=5, column=0, sticky="w", padx=5, pady=(10, 2)

        )

        # Cria um widget de texto para exibir as observações.
        # 'width=60' define a largura do widget em 60 caracteres.
        # 'height=3' define a altura do widget em 3 linhas de texto.
        # 'font=("Arial", 11)' configura o texto do widget em fonte Arial,
        #       tamanho 11, para melhorar a legibilidade.
        txt_obs = tk.Text(frame_info,
                          width=60,
                          height=3,
                          font=("Arial", 11))

        # Posiciona o widget de texto no frame de informações.
        # 'row=6' posiciona o widget na sétima linha do grid.
        # 'column=0' posiciona o widget na primeira coluna do grid.
        # 'sticky="w"' alinha o widget à esquerda da célula do grid.
        # 'pady=5' adiciona 5 pixels de margem vertical externa ao widget.
        txt_obs.grid(row=6,
                     column=0,
                     pady=5,
                     sticky="w")

        # Insere o texto das observações no widget de texto.
        # 'tk.END' insere o texto no final do conteúdo existente (inicialmente vazio).
        txt_obs.insert(tk.END, observacoes)

        # Configura o widget de texto como "somente leitura".
        # 'state="disabled"' desativa a edição do texto pelo usuário.
        txt_obs.config(state="disabled")

        # Cria um LabelFrame para agrupar e exibir a tabela de
        #       produtos utilizados no atendimento.
        # 'text="Produtos Utilizados"' define o título do
        #       LabelFrame como "Produtos Utilizados".
        # 'padx=10' adiciona 10 pixels de margem interna horizontal ao LabelFrame.
        # 'pady=10' adiciona 10 pixels de margem interna vertical ao LabelFrame.
        frame_prods = tk.LabelFrame(self,
                                    text="Produtos Utilizados",
                                    padx=10,
                                    pady=10)

        # Posiciona o LabelFrame na parte inferior da janela.
        # 'side="bottom"' alinha o LabelFrame na parte inferior.
        # 'fill="both"' faz com que o LabelFrame preencha tanto a
        #       largura quanto a altura disponíveis.
        # 'expand=True' permite que o LabelFrame se expanda
        #       proporcionalmente quando a janela é redimensionada.
        # 'padx=10' adiciona 10 pixels de margem horizontal externa ao LabelFrame.
        # 'pady=10' adiciona 10 pixels de margem vertical externa ao LabelFrame.
        frame_prods.pack(side="bottom",
                         fill="both",
                         expand=True,
                         padx=10,
                         pady=10)

        # Define as colunas da tabela de produtos.
        # Cada coluna representa um campo relevante de um produto utilizado.
        colunas = ("nome", "preco_unitario", "quantidade", "subtotal")

        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado



        # Cria um Treeview dentro do LabelFrame para exibir os produtos em formato de tabela.
        # 'columns=colunas' define as colunas do Treeview com base na
        #       lista de colunas especificada.
        # 'show="headings"' configura o Treeview para exibir apenas os
        #       cabeçalhos das colunas, sem uma coluna hierárquica inicial.
        self.tree_prods = ttk.Treeview(frame_prods,
                                       columns=colunas,
                                       show="headings")

        # Define o cabeçalho da coluna "nome".
        # 'heading("nome", text="Produto")' define o texto do cabeçalho como "Produto".
        self.tree_prods.heading("nome", text="Produto")

        # Define o cabeçalho da coluna "preco_unitario".
        # 'heading("preco_unitario", text="Preço Unit.")' define o
        #       texto do cabeçalho como "Preço Unit.".
        self.tree_prods.heading("preco_unitario", text="Preço Unit.")

        # Define o cabeçalho da coluna "quantidade".
        # 'heading("quantidade", text="Qtd")' define o texto do cabeçalho como "Qtd".
        self.tree_prods.heading("quantidade", text="Qtd")

        # Define o cabeçalho da coluna "subtotal".
        # 'heading("subtotal", text="Subtotal")' define o
        #       texto do cabeçalho como "Subtotal".
        self.tree_prods.heading("subtotal", text="Subtotal")

        # Configura a largura e o alinhamento da coluna "nome" na tabela de produtos.
        # 'width=200' define a largura da coluna como 200 pixels.
        self.tree_prods.column("nome", width=200)

        # Configura a largura e o alinhamento da coluna "preco_unitario".
        # 'width=100' define a largura da coluna como 100 pixels.
        # 'anchor="center"' alinha os valores da coluna no centro.
        self.tree_prods.column("preco_unitario", width=100, anchor="center")

        # Configura a largura e o alinhamento da coluna "quantidade".
        # 'width=80' define a largura da coluna como 80 pixels.
        # 'anchor="center"' alinha os valores da coluna no centro.
        self.tree_prods.column("quantidade", width=80, anchor="center")

        # Configura a largura e o alinhamento da coluna "subtotal".
        # 'width=100' define a largura da coluna como 100 pixels.
        # 'anchor="center"' alinha os valores da coluna no centro.
        self.tree_prods.column("subtotal", width=100, anchor="center")

        # Posiciona a tabela de produtos dentro do LabelFrame.
        # 'side="left"' alinha a tabela no lado esquerdo do LabelFrame.
        # 'fill="both"' faz com que a tabela preencha toda a largura e
        #       altura disponíveis no LabelFrame.
        # 'expand=True' permite que a tabela expanda proporcionalmente
        #       quando o LabelFrame for redimensionado.
        self.tree_prods.pack(side="left",
                             fill="both",
                             expand=True)

        # Cria uma barra de rolagem vertical para a tabela de produtos.
        # 'orient="vertical"' especifica que a barra será vertical.
        # 'command=self.tree_prods.yview' conecta a barra de rolagem à
        #       função de visualização vertical do Treeview.
        sb_vertical = ttk.Scrollbar(frame_prods,
                                    orient="vertical",
                                    command=self.tree_prods.yview)

        # Configura a barra de rolagem como o controlador de rolagem vertical da tabela.
        # 'yscrollcommand=sb_vertical.set' associa a barra de rolagem ao Treeview.
        self.tree_prods.configure(yscrollcommand=sb_vertical.set)

        # Posiciona a barra de rolagem dentro do LabelFrame.
        # 'side="right"' alinha a barra no lado direito do LabelFrame.
        # 'fill="y"' faz com que a barra preencha toda a altura disponível do LabelFrame.
        sb_vertical.pack(side="right", fill="y")

        # Carrega os produtos utilizados no atendimento para exibição na tabela.
        # 'registro' contém os dados do atendimento atual,
        #       incluindo os produtos utilizados.
        self.carregar_produtos(registro)


    def carregar_produtos(self, registro):

        # Obtém a lista de produtos utilizados no atendimento.
        # 'produtos_usados' é uma chave no dicionário 'registro' que
        #       armazena os detalhes dos produtos.
        produtos = registro.get("produtos_usados", [])

        # Itera sobre a lista de produtos para adicionar cada item à tabela.
        for item in produtos:

            # Obtém o nome do produto. Se o nome não estiver disponível,
            #       utiliza uma string vazia como valor padrão.
            nome = item.get("nome", "")

            # Obtém o preço unitário do produto. Se não estiver
            #       disponível, utiliza 0.0 como valor padrão.
            preco = item.get("preco_unitario", 0.0)

            # Obtém a quantidade utilizada do produto. Se não estiver
            #       disponível, utiliza 0.0 como valor padrão.
            qtd = item.get("quantidade", 0.0)

            # Obtém o subtotal do produto (preço unitário x quantidade).
            # Se não estiver disponível, utiliza 0.0 como valor padrão.
            st = item.get("subtotal", 0.0)

            # Insere os detalhes do produto na tabela Treeview.
            # '""' indica que o item será inserido na raiz do Treeview.
            # 'tk.END' posiciona o item no final da tabela.
            # 'values=(nome, f"{preco:.2f}", f"{qtd:.2f}", f"{st:.2f}")'
            #       define os valores das colunas:
            # - 'nome': Nome do produto.
            # - 'f"{preco:.2f}"': Formata o preço com duas casas decimais.
            # - 'f"{qtd:.2f}"': Formata a quantidade com duas casas decimais.
            # - 'f"{st:.2f}"': Formata o subtotal com duas casas decimais.
            self.tree_prods.insert("",
                                   tk.END,
                                   values=(nome, f"{preco:.2f}", f"{qtd:.2f}", f"{st:.2f}"))




# ---------------------------------------------------------------------
# TelaEstoque
# ---------------------------------------------------------------------
class TelaEstoque(tk.Frame):

    # Método inicializador da classe 'TelaEstoque'.
    # Configura os elementos iniciais da interface de gerenciamento de estoque.
    def __init__(self, master, medico_logado, *args, **kwargs):

        # `self`: Referência ao objeto atual da classe 'TelaEstoque', permitindo
        #       acesso a seus atributos e métodos.
        # `master`: O container pai da tela, uma janela principal ou
        #       notebook onde a aba será exibida.
        # `medico_logado`: Um dicionário ou objeto representando o médico
        #       atualmente logado no sistema, usado para controle de alterações e auditorias.
        # `*args`: Argumentos posicionais adicionais que podem ser passados
        #       para o método da classe base.
        # `**kwargs`: Argumentos nomeados adicionais para maior flexibilidade, como
        #       configurar propriedades específicas do frame.

        # Chama o inicializador da classe pai ('tk.Frame') para herdar e configurar o frame base.
        super().__init__(master, *args, **kwargs)

        # Armazena as informações do médico logado no atributo `self.medico_logado`.
        # Isso pode ser útil para rastrear quem está realizando alterações no estoque.
        self.medico_logado = medico_logado

        # Inicializa o atributo `produto_id_atual` como `None`.
        # Este atributo será usado para armazenar o ID do produto selecionado na interface.
        self.produto_id_atual = None

        # Define a cor de fundo do frame como '#F7F7F7', um tom de cinza claro.
        # A escolha dessa cor ajuda a criar um layout limpo e profissional.
        self.configure(bg="#F7F7F7")

        # Cria um rótulo (Label) para exibir o título da tela de
        #       gerenciamento de estoque.
        titulo = tk.Label(

            # O rótulo será adicionado como filho do frame 'TelaEstoque'.
            self,

            # Define o texto que será exibido no rótulo.
            text="Gerenciamento de Estoque",

            # Define a fonte usada no texto:
            # - Fonte: 'Arial'.
            # - Tamanho: 18.
            # - Estilo: 'bold' (negrito).
            font=("Arial", 18, "bold"),

            # Configura o fundo do rótulo para a mesma cor de fundo do
            #       frame '#F7F7F7' (cinza claro).
            bg="#F7F7F7",

            # Configura a cor do texto do rótulo como '#333333' (cinza escuro),
            #       proporcionando contraste adequado.
            fg="#333333"
        )

        # Posiciona o rótulo no frame:
        # - 'pady=10': Adiciona um espaçamento vertical de 10 pixels
        #       acima e abaixo do rótulo.
        titulo.pack(pady=10)

        # Cria um LabelFrame para o formulário de cadastro/edição de produtos.
        frame_form = tk.LabelFrame(

            self,  # O frame é adicionado como filho do frame 'TelaEstoque'.

            # Define o título exibido no topo do frame.
            text="Cadastrar/Editar Produto",

            # Configurações de fonte:
            # - Fonte: 'Arial'.
            # - Tamanho: 12.
            # - Estilo: 'bold' (negrito).
            font=("Arial", 12, "bold"),

            # Define a cor de fundo do frame como '#FFFFFF' (branco).
            bg="#FFFFFF",

            # Define a cor do texto do título como '#333333' (cinza escuro),
            #       garantindo boa legibilidade.
            fg="#333333",

            # Configura o espaçamento interno do frame:
            # - 'padx=10': Adiciona 10 pixels de margem interna nas
            #       laterais esquerda e direita.
            # - 'pady=10': Adiciona 10 pixels de margem interna acima e abaixo.
            padx=10,
            pady=10
        )

        # Posiciona o frame no layout:
        # - 'side="top"': Alinha o frame na parte superior do layout.
        # - 'fill="x"': Faz com que o frame preencha toda a largura disponível horizontalmente.
        # - 'padx=20': Adiciona uma margem horizontal de 20 pixels ao redor do frame.
        # - 'pady=10': Adiciona uma margem vertical de 10 pixels acima e abaixo do frame.
        frame_form.pack(side="top", fill="x", padx=20, pady=10)

        # Define os campos do formulário como uma lista de tuplas.
        # Cada tupla contém:
        # 1. O texto do rótulo que será exibido.
        # 2. O nome do atributo que será associado ao campo correspondente.
        campos = [
            ("Nome do Produto:", "entry_nome_produto"),  # Campo para o nome do produto.
            ("Tipo do Produto:", "combo_tipo_produto"),  # Combobox para selecionar o tipo do produto.
            ("Quantidade:", "entry_quantidade"),  # Campo para a quantidade disponível no estoque.
            ("Unidade (kg, ml, un):", "entry_unidade"),  # Campo para definir a unidade de medida.
            ("Preço Unitário (R$):", "entry_preco_unit"),  # Campo para o preço unitário do produto.
            ("Data de Validade:", "data_validade"),  # Campo para a data de validade.
            ("Observações:", "entry_obs")  # Campo para inserir observações adicionais.
        ]

        # Itera sobre os campos definidos na lista 'campos', onde cada item
        #       contém um texto do rótulo e o nome do atributo associado.
        for i, (label_txt, attr_name) in enumerate(campos):

            # Cria um rótulo (Label) para cada campo no formulário.
            # 'frame_form' é o frame onde o rótulo será posicionado.
            # 'text=label_txt' define o texto do rótulo (ex.: "Nome do Produto:").
            # 'font=("Arial", 12)' define o estilo e tamanho da fonte do rótulo.
            # 'bg="#FFFFFF"' define o fundo do rótulo como branco.
            tk.Label(frame_form,
                     text=label_txt,
                     font=("Arial", 12),
                     bg="#FFFFFF").grid(

                # Posiciona o rótulo na interface usando o gerenciador de layout 'grid'.
                # 'row=i' posiciona o rótulo na linha correspondente ao índice atual.
                # 'column=0' posiciona o rótulo na primeira coluna.
                # 'sticky="e"' alinha o texto do rótulo à direita.
                # 'padx=5' e 'pady=5' adicionam espaçamento ao redor do rótulo.
                row=i, column=0, sticky="e", padx=5, pady=5)

            # Verifica se o atributo atual é "combo_tipo_produto",
            #       que exige um ComboBox.
            if attr_name == "combo_tipo_produto":

                # Cria um ComboBox (caixa de seleção) para selecionar o tipo do produto.
                # 'values' define as opções disponíveis para o ComboBox.
                # 'font=("Arial", 12)' aplica o estilo e tamanho da fonte no ComboBox.
                # 'width=40' define a largura do ComboBox.
                combo = ttk.Combobox(frame_form,
                                     values=["Medicamento", "Ração", "Acessório", "Outros"],
                                     font=("Arial", 12),
                                     width=40)

                # Define o item selecionado por padrão como o primeiro da lista.
                combo.current(0)

                # Atribui o ComboBox como um atributo da instância da classe,
                #       utilizando o nome especificado.
                setattr(self, attr_name, combo)

                # Posiciona o ComboBox na interface.
                # 'row=i' posiciona o ComboBox na mesma linha do rótulo correspondente.
                # 'column=1' posiciona o ComboBox na segunda coluna.
                # 'padx=5' e 'pady=5' adicionam espaçamento ao redor do ComboBox.
                combo.grid(row=i, column=1, padx=5, pady=5)


            # Verifica se o atributo atual é "data_validade", que exige um
            #       widget de calendário (DateEntry).
            elif attr_name == "data_validade":

                # Cria um widget DateEntry para seleção de datas, configurado
                #       para exibir o formato 'dd/mm/yyyy'.
                # 'font=("Arial", 12)' define o estilo e tamanho da fonte do calendário.
                # 'width=20' define a largura do campo de entrada para a data.
                # 'background="#4CAF50"' define a cor de fundo do calendário.
                # 'foreground="white"' define a cor do texto no calendário.
                cal = DateEntry(frame_form,
                                date_pattern='dd/mm/yyyy',
                                font=("Arial", 12),
                                width=20,
                                background="#4CAF50",
                                foreground="white")

                # Atribui o widget de calendário como um atributo da instância da
                #       classe, utilizando o nome especificado.
                setattr(self, attr_name, cal)

                # Posiciona o widget DateEntry na interface.
                # 'row=i' posiciona o calendário na mesma linha do rótulo correspondente.
                # 'column=1' posiciona o calendário na segunda coluna.
                # 'padx=5' e 'pady=5' adicionam espaçamento ao redor do widget.
                # 'sticky="w"' alinha o calendário à esquerda dentro da célula.
                cal.grid(row=i, column=1, padx=5, pady=5, sticky="w")

            # Caso contrário, para os demais atributos, cria um campo de
            #       entrada de texto (Entry).
            else:

                # Cria um widget Entry para entrada de texto.
                # 'font=("Arial", 12)' define o estilo e tamanho da fonte do campo.
                # 'width=40' define a largura do campo de entrada.
                entry = tk.Entry(frame_form, font=("Arial", 12), width=40)

                # Atribui o widget Entry como um atributo da instância da
                #       classe, utilizando o nome especificado.
                setattr(self, attr_name, entry)

                # Posiciona o widget Entry na interface.
                # 'row=i' posiciona o campo de entrada na mesma linha
                #       do rótulo correspondente.
                # 'column=1' posiciona o campo de entrada na segunda coluna.
                # 'padx=5' e 'pady=5' adicionam espaçamento ao redor do campo.
                entry.grid(row=i, column=1, padx=5, pady=5)

        # Cria um container (Frame) para organizar os botões
        #       dentro do formulário.
        # 'bg="#FFFFFF"' define a cor de fundo do frame para branco,
        #       mantendo a consistência visual.
        frame_botoes = tk.Frame(frame_form, bg="#FFFFFF")

        # Posiciona o frame de botões abaixo dos campos do formulário.
        # 'row=len(campos)' coloca o frame na linha seguinte ao último campo.
        # 'column=0' e 'columnspan=2' fazem o frame ocupar duas
        #       colunas, centralizando os botões.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical entre o
        #       frame e os outros elementos.
        frame_botoes.grid(row=len(campos),
                          column=0,
                          columnspan=2,
                          pady=10)

        # Define uma lista de tuplas contendo o texto, a função
        #       associada e a cor de fundo de cada botão.
        botoes = [
            ("Salvar/Atualizar", self.salvar, "#4CAF50"),  # Botão para salvar ou atualizar o produto, com cor verde.
            ("Excluir", self.excluir, "#F44336"),  # Botão para excluir o produto, com cor vermelha.
            ("Limpar", self.limpar_form, "#2196F3")  # Botão para limpar o formulário, com cor azul.
        ]

        # Itera sobre a lista de botões definida anteriormente.
        # Para cada botão, 'text' contém o texto exibido, 'command' a
        #       função associada e 'color' a cor de fundo.
        for text, command, color in botoes:

            # Cria um botão dentro do frame de botões (frame_botoes).
            # 'text=text' define o texto exibido no botão.
            # 'font=("Arial", 12, "bold")' utiliza a fonte Arial em tamanho 12, com estilo negrito.
            # 'bg=color' define a cor de fundo do botão com base na cor especificada na lista.
            # 'fg="#FFFFFF"' define a cor do texto do botão como branco.
            # 'activebackground=color' altera a cor de fundo do botão para a mesma cor ao ser clicado.
            # 'activeforeground="#FFFFFF"' mantém a cor do texto como branco durante a interação.
            # 'command=command' vincula a função associada ao evento de clique no botão.
            # 'width=15' define a largura do botão, garantindo uniformidade no layout.
            tk.Button(  frame_botoes,
                        text=text,
                        font=("Arial", 12, "bold"),
                        bg=color,
                        fg="#FFFFFF",
                        activebackground=color,
                        activeforeground="#FFFFFF",
                        command=command,
                        width=15).pack( side="left",  # Posiciona os botões lado a lado no frame.
                                        padx=10) # Adiciona 10 pixels de espaçamento horizontal entre os botões.


        # Criação do frame que irá conter a lista de produtos do estoque.
        frame_lista = tk.LabelFrame(self,  # Define o elemento pai como sendo a própria classe/tela em que o frame será inserido.
                                    text="Estoque Atual",  # Define o título do LabelFrame como "Estoque Atual".
                                    font=("Arial", 12, "bold"),  # Define a fonte utilizada no título: Arial, tamanho 12, em negrito.
                                    bg="#FFFFFF",  # Define o fundo do LabelFrame como branco (#FFFFFF).
                                    fg="#333333",  # Define a cor do texto do título como cinza escuro (#333333).
                                    padx=10,  # Adiciona 10 pixels de preenchimento horizontal interno (espaçamento entre a borda e o conteúdo).
                                    pady=10)  # Adiciona 10 pixels de preenchimento vertical interno.


        # Posiciona o frame contendo a lista de produtos na parte inferior da janela.
        frame_lista.pack(   side="bottom",  # Alinha o frame na parte inferior da janela.
                            fill="both",  # Permite que o frame preencha tanto a largura quanto a altura disponíveis.
                            expand=True,  # Faz com que o frame se expanda proporcionalmente ao redimensionar a janela.
                            padx=20,  # Adiciona 20 pixels de margem horizontal externa ao redor do frame.
                            pady=10)  # Adiciona 10 pixels de margem vertical externa ao redor do frame.


        # Define as colunas da tabela que exibirá os produtos do estoque.
        colunas = (
            "_id",  # Identificador único do produto.
            "nome_produto",  # Nome do produto.
            "tipo_produto",  # Tipo do produto (ex.: Medicamento, Ração).
            "quantidade",  # Quantidade disponível do produto.
            "unidade",  # Unidade de medida (ex.: kg, ml, un).
            "preco_unitario",  # Preço unitário do produto.
            "data_validade",  # Data de validade do produto.
            "observacoes"  # Observações adicionais sobre o produto.
        )

        # Configuração do estilo da tabela (Treeview) para
        #       personalizar sua aparência.
        # Cria um objeto de estilo para personalizar elementos visuais.
        estilo = ttk.Style()

        # Configura o estilo das células da tabela (Treeview).
        estilo.configure("Treeview",  # Aplica as configurações à tabela (Treeview).
                            font=("Arial", 10),  # Define a fonte das células como Arial, tamanho 10.
                            rowheight=25,  # Define a altura de cada linha da tabela como 25 pixels.
                            background="#FFFFFF",  # Define o fundo das células como branco.
                            fieldbackground="#FFFFFF") # Define o fundo interno dos campos das células como branco.


        # Configura o estilo do cabeçalho da tabela (Treeview.Heading).
        estilo.configure("Treeview.Heading",  # Aplica as configurações ao cabeçalho da tabela.
                            font=("Arial", 12, "bold"),  # Define a fonte do cabeçalho como Arial, tamanho 12, em negrito.
                            background="#F0F0F0",  # Define o fundo do cabeçalho como cinza claro.
                            foreground="#333333")  # Define a cor do texto do cabeçalho como cinza escuro.


        # Configura o estilo para o estado selecionado das células da tabela.
        estilo.map("Treeview",  # Aplica as configurações à tabela.
            background=[("selected", "#D3D3D3")])  # Define o fundo das células selecionadas como cinza claro.


        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria a tabela (Treeview) no frame destinado à lista de produtos.
        self.tree = ttk.Treeview(frame_lista,  # Define o frame pai onde a tabela será inserida.
                                columns=colunas,  # Configura as colunas da tabela com os nomes especificados.
                                show="headings",  # Mostra apenas os cabeçalhos, sem coluna inicial vazia.
                                style="Treeview")  # Aplica o estilo previamente configurado.


        # Configura os cabeçalhos e colunas da tabela.
        for c in colunas:

            # Define o texto exibido no cabeçalho da coluna, formatando o nome da coluna.
            self.tree.heading(c, text=c.replace("_", " ").title())

            # Configura o alinhamento e largura das colunas.
            if c == "_id":

                # Para a coluna '_id', centraliza o texto e define largura de 100 pixels.
                self.tree.column(c, anchor="center", width=100)

            else:

                # Para as demais colunas, centraliza o texto e define largura de 150 pixels.
                self.tree.column(c, anchor="center", width=150)

        # Insere a tabela no frame, configurando o preenchimento e espaçamento.
        self.tree.pack( fill="both",  # A tabela preenche todo o espaço disponível, tanto horizontal quanto vertical.
            expand=True,  # Permite que a tabela expanda proporcionalmente quando o frame é redimensionado.
            pady=5)  # Adiciona um espaçamento vertical de 5 pixels acima e abaixo da tabela.


        # Cria uma barra de rolagem vertical para a tabela.
        # 'frame_lista' é o container da tabela e a barra de
        #       rolagem será usada para navegar pelos itens.
        scrollbar = ttk.Scrollbar(frame_lista,
                                  orient="vertical",
                                  command=self.tree.yview)

        # Configura a tabela para que a barra de rolagem acompanhe as
        #       mudanças de visualização.
        # Define 'yscrollcommand' como a função da barra de rolagem.
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Posiciona a barra de rolagem no lado direito do frame.
        # 'side="right"' posiciona a barra à direita,
        #       enquanto 'fill="y"' a estende verticalmente.
        scrollbar.pack(side="right", fill="y")

        # Associa o evento de seleção de uma linha na tabela ao método 'ao_selecionar_linha'.
        # Isso será chamado sempre que o usuário clicar em uma linha.
        self.tree.bind("<<TreeviewSelect>>", self.ao_selecionar_linha)

        # Associa o evento de duplo clique em uma linha ao método 'abrir_historico_produto'.
        # Este método abrirá os detalhes do histórico do produto selecionado.
        self.tree.bind("<Double-1>", self.abrir_historico_produto)

        # Chama o método 'carregar_estoque' para preencher a
        #       tabela com os dados do estoque.
        # Esse método busca os dados no banco e os insere na tabela.
        self.carregar_estoque()


    # Define o método para carregar os dados do estoque na tabela.
    def carregar_estoque(self):

        # Remove todos os itens atuais da tabela antes de inserir novos.
        # 'self.tree.get_children()' retorna todos os IDs das
        #       linhas existentes na tabela.
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Recupera todos os produtos da coleção 'estoque' no banco de dados.
        produtos = db.estoque.find()

        # Itera pelos produtos encontrados no banco.
        for prod in produtos:

            # Inicializa a string de validade como vazia.
            val_str = ""

            # Verifica se o produto possui uma data de validade registrada.
            if prod.get("data_validade"):

                # Formata a data de validade para exibição no formato apropriado.
                val_str = formatar_data_somente(prod["data_validade"])

            # Insere uma nova linha na tabela com os dados do produto.
            self.tree.insert(

                "",  # Especifica que a nova linha será inserida na raiz da árvore (sem pai).
                tk.END,  # Define que a linha será adicionada ao final da tabela.
                values=(  # Define os valores das colunas da tabela.
                    str(prod["_id"]),  # Converte o ID do produto para string e o insere na primeira coluna.
                    prod.get("nome_produto", ""),
                    # Obtém o nome do produto ou insere uma string vazia se não existir.
                    prod.get("tipo_produto", ""),  # Obtém o tipo do produto ou insere uma string vazia.
                    prod.get("quantidade", 0),  # Insere a quantidade do produto ou 0 como padrão.
                    prod.get("unidade", ""),  # Insere a unidade do produto (ex.: kg, ml, un) ou uma string vazia.
                    f"{prod.get('preco_unitario', 0.0):.2f}",
                    # Formata o preço unitário como moeda com duas casas decimais.
                    val_str,  # Insere a data de validade formatada ou uma string vazia.
                    prod.get("observacoes", "")  # Insere as observações do produto ou uma string vazia.
                )
            )

    def abrir_historico_produto(self, event):

        print("...")


    # Define o método para lidar com a seleção de uma linha na tabela.
    def ao_selecionar_linha(self, event):

        # Obtém a seleção atual da Treeview.
        selection = self.tree.selection()

        # Verifica se não há nenhuma linha selecionada.
        if not selection:
            return  # Sai da função se não houver seleção.

        # Obtém o item correspondente à seleção.
        item = self.tree.item(selection[0])

        # Extrai os valores das colunas da linha selecionada.
        vals = item["values"]

        # Armazena o ID do produto atual como uma string.
        self.produto_id_atual = vals[0]  # Primeiro valor representa o `_id`.

        # Extrai e armazena os dados restantes da linha selecionada.
        nome_produto = vals[1]  # Nome do produto (segunda coluna).
        tipo_produto = vals[2]  # Tipo do produto (terceira coluna).
        quantidade = vals[3]  # Quantidade em estoque (quarta coluna).
        unidade = vals[4]  # Unidade de medida (ex.: kg, ml, un) (quinta coluna).
        preco = vals[5]  # Preço unitário formatado como string (sexta coluna).
        data_val = vals[6]  # Data de validade, formatada (sétima coluna).
        obs = vals[7]  # Observações do produto (oitava coluna).

        # Limpa o campo de entrada "Nome do Produto" antes de inserir um novo valor.
        self.entry_nome_produto.delete(0, tk.END)

        # Insere o nome do produto extraído da linha selecionada.
        self.entry_nome_produto.insert(0, nome_produto)

        # Define o tipo do produto no combobox com base no valor extraído.
        self.combo_tipo_produto.set(tipo_produto)

        # Limpa o campo de entrada "Quantidade" antes de inserir um novo valor.
        self.entry_quantidade.delete(0, tk.END)

        # Insere a quantidade do produto extraída da linha selecionada.
        self.entry_quantidade.insert(0, quantidade)

        # Limpa o campo de entrada "Unidade" antes de inserir um novo valor.
        self.entry_unidade.delete(0, tk.END)

        # Insere a unidade do produto extraída da linha selecionada.
        self.entry_unidade.insert(0, unidade)

        # Limpa o campo de entrada "Preço Unitário" antes de inserir um novo valor.
        self.entry_preco_unit.delete(0, tk.END)

        # Insere o preço unitário do produto extraído da linha selecionada.
        self.entry_preco_unit.insert(0, preco)

        # Verifica se há uma data de validade fornecida.
        if data_val:

            try:

                # Divide a string da data no formato "dd/mm/yyyy" em dia, mês e ano.
                dia, mes, ano = data_val.split("/")

                # Cria um objeto datetime com os valores extraídos.
                dt = datetime.datetime(int(ano), int(mes), int(dia))

                # Define a data no widget "DateEntry" com a data extraída.
                self.data_validade.set_date(dt)

            except:

                # Em caso de erro no formato da data, define a data atual no widget.
                self.data_validade.set_date(datetime.datetime.now())

        else:

            # Caso nenhuma data de validade seja fornecida, define a data atual no widget.
            self.data_validade.set_date(datetime.datetime.now())

        # Limpa o campo de entrada "Observações" antes de inserir um novo valor.
        self.entry_obs.delete(0, tk.END)

        # Insere as observações extraídas da linha selecionada.
        self.entry_obs.insert(0, obs)


    def abrir_historico_produto(self, event):

        # Obtém o item atualmente selecionado na Treeview.
        selection = self.tree.selection()

        # Verifica se não há nenhum item selecionado. Se for o
        #       caso, retorna sem fazer nada.
        if not selection:
            return

        # Obtém os detalhes do item selecionado, incluindo seus valores.
        item = self.tree.item(selection[0])
        vals = item["values"]

        # O ID do produto é o primeiro valor na linha selecionada.
        produto_id_str = vals[0]

        # Abre uma nova janela para exibir o histórico do produto selecionado.
        JanelaHistoricoProduto(self, produto_id_str)


    def salvar(self):

        # Obtém o nome do produto do campo de entrada, removendo espaços em branco.
        nome_produto = self.entry_nome_produto.get().strip()

        # Verifica se o nome do produto está vazio. Exibe um aviso e
        #       interrompe o processo caso esteja.
        if not nome_produto:
            messagebox.showwarning("Aviso",
                                   "Nome do Produto é obrigatório.")
            return

        # Obtém o tipo do produto selecionado no combobox, removendo espaços em branco.
        tipo_produto = self.combo_tipo_produto.get().strip()

        # Obtém o texto da quantidade do produto, removendo espaços em branco.
        qtd_txt = self.entry_quantidade.get().strip()

        # Obtém a unidade do produto (ex.: "kg", "ml"), removendo espaços em branco.
        unidade = self.entry_unidade.get().strip()

        # Obtém o preço unitário do produto como texto, removendo espaços em branco.
        preco_txt = self.entry_preco_unit.get().strip()

        # Obtém a data de validade selecionada no widget DateEntry.
        data_val = self.data_validade.get_date()

        # Obtém as observações sobre o produto do campo de entrada,
        #       removendo espaços em branco.
        obs = self.entry_obs.get().strip()

        try:

            # Tenta converter o texto de quantidade para um número inteiro.
            # Caso ocorra um erro (ex.: texto vazio ou caractere
            #       inválido), define a quantidade como 0.
            quantidade = int(qtd_txt)

        except:

            quantidade = 0

        try:

            # Tenta converter o texto do preço unitário para um número de ponto flutuante.
            # Substitui vírgulas por pontos para compatibilidade numérica.
            # Caso ocorra um erro, define o preço unitário como 0.0.
            preco_unit = float(preco_txt.replace(",", "."))

        except:

            preco_unit = 0.0

        # Converte a data selecionada (apenas ano, mês e dia) para um objeto datetime.
        dt_val = datetime.datetime(data_val.year, data_val.month, data_val.day)

        # Cria o dicionário que representa o documento a ser
        #       salvo no banco de dados.
        doc = {

            # Nome do produto.
            "nome_produto": nome_produto,

            # Tipo do produto (ex.: Medicamento, Ração, etc.).
            "tipo_produto": tipo_produto,

            # Quantidade disponível no estoque.
            "quantidade": quantidade,

            # Unidade de medida (ex.: kg, ml, un).
            "unidade": unidade,

            # Preço unitário do produto.
            "preco_unitario": preco_unit,

            # Data de validade do produto.
            "data_validade": dt_val,

            # Data e hora do registro no sistema (entrada no estoque).
            "data_entrada": datetime.datetime.now(),

            # Observações adicionais sobre o produto.
            "observacoes": obs,

            # ID do médico responsável pela última alteração.
            "ultima_alteracao_por_id": self.medico_logado["_id"],

            # Nome do médico responsável pela última alteração.
            "ultima_alteracao_por_nome": self.medico_logado["nome"],

            # Data e hora da última alteração no registro.
            "ultima_alteracao_em": datetime.datetime.now()

        }

        try:

            # Verifica se o produto já existe no banco de dados.
            if self.produto_id_atual is None:

                # Caso o produto seja novo, insere o documento no
                #       banco de dados na coleção "estoque".
                db.estoque.insert_one(doc)

            else:

                # Caso o produto já exista, atualiza o documento
                #       correspondente com base no "_id".
                db.estoque.update_one({"_id": ObjectId(self.produto_id_atual)}, {"$set": doc})

            # Exibe uma mensagem informando que o produto foi salvo ou atualizado com sucesso.
            messagebox.showinfo("Sucesso", "Produto salvo/atualizado!")

            # Recarrega os dados do estoque para atualizar a tabela
            #       com as informações mais recentes.
            self.carregar_estoque()

            # Limpa o formulário de entrada para preparar para um
            #       novo registro ou edição.
            self.limpar_form()

        except Exception as e:

            # Caso ocorra algum erro durante a operação de salvar ou
            #       atualizar, exibe uma mensagem de erro.
            messagebox.showerror("Erro", f"Não foi possível salvar:\n{e}")



    def excluir(self):

        # Verifica se um produto foi selecionado na tabela.
        if not self.produto_id_atual:

            # Exibe um aviso caso nenhum produto esteja selecionado.
            messagebox.showwarning("Aviso", "Selecione um produto para excluir.")
            return

        # Pergunta ao usuário se ele realmente deseja excluir o produto selecionado.
        resp = messagebox.askyesno("Confirmação", "Deseja excluir este produto?")
        if resp:

            try:

                # Realiza a exclusão do produto no banco de dados com base no "_id" selecionado.
                db.estoque.delete_one({"_id": ObjectId(self.produto_id_atual)})

                # Exibe uma mensagem de sucesso após a exclusão.
                messagebox.showinfo("Sucesso", "Produto excluído!")

                # Atualiza a tabela de estoque para refletir a exclusão.
                self.carregar_estoque()

                # Limpa o formulário de entrada para garantir que
                #       nenhum dado antigo permaneça.
                self.limpar_form()

            except Exception as e:

                # Caso ocorra algum erro durante a exclusão, exibe uma
                #       mensagem de erro com detalhes.
                messagebox.showerror("Erro", f"Não foi possível excluir:\n{e}")



    def limpar_form(self):

        # Redefine o ID do produto atual para None, indicando que
        #       nenhum produto está selecionado.
        self.produto_id_atual = None

        # Limpa o campo de entrada do nome do produto.
        self.entry_nome_produto.delete(0, tk.END)

        # Reseta o combobox do tipo de produto para o valor
        #       padrão (primeira opção).
        self.combo_tipo_produto.current(0)

        # Limpa o campo de entrada da quantidade do produto.
        self.entry_quantidade.delete(0, tk.END)

        # Limpa o campo de entrada da unidade de medida do produto.
        self.entry_unidade.delete(0, tk.END)

        # Limpa o campo de entrada do preço unitário do produto.
        self.entry_preco_unit.delete(0, tk.END)

        # Reseta a data de validade para a data atual.
        self.data_validade.set_date(datetime.datetime.now())

        # Limpa o campo de entrada de observações do produto.
        self.entry_obs.delete(0, tk.END)



# ---------------------------------------------------------------------
# TelaRelatorios
# ---------------------------------------------------------------------

class TelaRelatorios(tk.Frame):

    """
    Classe que representa a Tela de Relatórios, exibindo atendimentos com detalhes
    como data, animal, médico e valores. Permite a interação por duplo clique para
    abrir uma janela detalhada (JanelaDetalheRelatorio) com informações adicionais
    e uma tabela de produtos usados.
    """

    def __init__(self, master, medico_logado, *args, **kwargs):

        """
        Inicializa a Tela de Relatórios.

        Parâmetros:
        - master: Componente pai (normalmente uma janela ou notebook).
        - medico_logado: Dicionário com informações do médico atualmente logado.
        - *args: Argumentos adicionais para a classe pai.
        - **kwargs: Argumentos nomeados adicionais para a classe pai.
        """

        # Chama o inicializador da classe pai (tk.Frame), garantindo que
        # todas as propriedades herdadas sejam configuradas.
        super().__init__(master, *args, **kwargs)

        # Armazena as informações do médico logado para uso na classe.
        self.medico_logado = medico_logado

        # Lista interna que guardará cada registro de atendimento encontrado,
        # inclusive um "id_interno" ou algo que identifique de onde veio.
        self.resultados_encontrados = []

        # Criação do Frame de Filtros
        # Este frame será utilizado para agrupar os componentes de
        #       filtro da tela de relatórios.

        frame_filtros = tk.LabelFrame(self,  # Define o frame como um componente filho do frame principal.
                                      text="Filtros de Relatório",  # Título exibido no topo do frame.
                                      padx=10,  # Define um padding interno horizontal de 10 pixels.
                                      pady=10)  # Define um padding interno vertical de 10 pixels.

        # Posicionamento do Frame de Filtros
        # 'side="top"' posiciona o frame na parte superior do layout.
        # 'fill="x"' faz com que o frame preencha horizontalmente toda a largura disponível.
        # 'padx=10' adiciona um espaço de 10 pixels em ambos os lados do frame.
        # 'pady=10' adiciona um espaço de 10 pixels acima e abaixo do frame.
        frame_filtros.pack(side="top",  # Posiciona o frame no topo da tela.
                           fill="x",  # Faz o frame ocupar toda a largura disponível.
                           padx=10,  # Adiciona margem horizontal ao redor do frame.
                           pady=10)  # Adiciona margem vertical ao redor do frame.

        # Cria um rótulo (label) com o texto "Data Inicial:" dentro do frame de filtros.
        # Este rótulo serve para indicar ao usuário onde selecionar a data
        #       inicial para o filtro do relatório.
        # 'row=0' posiciona o rótulo na primeira linha do grid.
        # 'column=0' posiciona o rótulo na primeira coluna do grid.
        # 'padx=5' adiciona um espaço horizontal de 5 pixels à esquerda e à direita do rótulo.
        # 'pady=5' adiciona um espaço vertical de 5 pixels acima e abaixo do rótulo.
        # 'sticky="e"' alinha o texto à direita dentro da célula do
        #       grid para uma melhor organização visual.
        tk.Label(frame_filtros,
                 text="Data Inicial:").grid(row=0,
                                            column=0,
                                            sticky="e",
                                            padx=5,
                                            pady=5)

        # Cria um campo de entrada de data (DateEntry) para a "Data Inicial".
        # Este campo permite que o usuário selecione uma data no
        #       formato 'dd/mm/yyyy', útil para o filtro do relatório.
        # 'width=12' define a largura do campo como 12 caracteres.
        # 'font=("Arial", 12)' aplica a fonte Arial com tamanho 12 para garantir legibilidade.
        self.data_ini = DateEntry(frame_filtros,
                                  date_pattern='dd/mm/yyyy',
                                  width=12,
                                  font=("Arial", 12))

        # Posiciona o campo de entrada de data na primeira linha e segunda coluna do grid.
        # 'padx=5' adiciona um espaço horizontal de 5 pixels à esquerda e à direita do campo.
        # 'pady=5' adiciona um espaço vertical de 5 pixels acima e abaixo do campo.
        self.data_ini.grid(row=0, column=1, padx=5, pady=5)

        # Cria um rótulo (label) com o texto "Data Final:" dentro do frame de filtros.
        # Este rótulo serve para indicar ao usuário onde selecionar a
        #       data final para o filtro do relatório.
        # 'row=0' posiciona o rótulo na primeira linha do grid.
        # 'column=2' posiciona o rótulo na terceira coluna do grid.
        # 'padx=5' adiciona um espaço horizontal de 5 pixels à esquerda e à direita do rótulo.
        # 'pady=5' adiciona um espaço vertical de 5 pixels acima e abaixo do rótulo.
        # 'sticky="e"' alinha o texto à direita dentro da célula do grid para uma melhor organização visual.
        tk.Label(frame_filtros,
                 text="Data Final:").grid(row=0,
                                          column=2,
                                          sticky="e",
                                          padx=5,
                                          pady=5)

        # Cria um campo de entrada de data (DateEntry) para a "Data Final".
        # Este campo permite que o usuário selecione uma data no formato 'dd/mm/yyyy',
        #       útil para definir o período do relatório.
        # 'width=12' define a largura do campo como 12 caracteres.
        # 'font=("Arial", 12)' aplica a fonte Arial com tamanho 12 para
        #       garantir legibilidade.
        self.data_fim = DateEntry(frame_filtros,
                                  date_pattern='dd/mm/yyyy',
                                  width=12,
                                  font=("Arial", 12))

        # Posiciona o campo de entrada de data na primeira linha e quarta coluna do grid.
        # 'padx=5' adiciona um espaço horizontal de 5 pixels à esquerda e à direita do campo.
        # 'pady=5' adiciona um espaço vertical de 5 pixels acima e abaixo do campo.
        self.data_fim.grid(row=0,
                           column=3,
                           padx=5,
                           pady=5)

        # Cria um rótulo (label) com o texto "Médico:" dentro do frame de filtros.
        # Este rótulo indica ao usuário onde selecionar o médico para filtrar os relatórios.
        # 'row=0' posiciona o rótulo na primeira linha do grid.
        # 'column=4' posiciona o rótulo na quinta coluna do grid.
        # 'padx=5' adiciona um espaço horizontal de 5 pixels à esquerda e à direita do rótulo.
        # 'pady=5' adiciona um espaço vertical de 5 pixels acima e abaixo do rótulo.
        # 'sticky="e"' alinha o texto à direita dentro da célula do grid
        #       para manter a organização visual.
        tk.Label(frame_filtros,
                 text="Médico:").grid(row=0,
                                      column=4,
                                      sticky="e",
                                      padx=5,
                                      pady=5)

        # Cria uma caixa de combinação (Combobox) para selecionar o médico.
        # Este campo permite que o usuário escolha um médico específico para o filtro do relatório.
        # 'width=20' define a largura da caixa como 20 caracteres.
        # 'font=("Arial", 12)' aplica a fonte Arial com tamanho 12 para uma leitura confortável.
        self.combo_medico = ttk.Combobox(frame_filtros,
                                         width=20,
                                         font=("Arial", 12))

        # Posiciona a caixa de combinação na primeira linha e sexta coluna do grid.
        # 'padx=5' adiciona um espaço horizontal de 5 pixels à esquerda e à direita da caixa.
        # 'pady=5' adiciona um espaço vertical de 5 pixels acima e abaixo da caixa.
        self.combo_medico.grid(row=0, column=5, padx=5, pady=5)

        # Cria um botão para aplicar o filtro no relatório.
        # O texto "Filtrar" indica claramente sua funcionalidade ao usuário.
        # 'command=self.filtrar_relatorio' associa a funcionalidade do
        #       botão ao método `filtrar_relatorio`.
        # 'font=("Arial", 12, "bold")' aplica a fonte Arial em tamanho 12
        #       com estilo em negrito para maior destaque.
        btn_filtrar = tk.Button(frame_filtros,
                                text="Filtrar",
                                command=self.filtrar_relatorio,
                                font=("Arial", 12, "bold"))

        # Posiciona o botão "Filtrar" na primeira linha e sétima coluna do grid no frame de filtros.
        # 'padx=5' adiciona um espaço horizontal de 5 pixels à esquerda e à direita do botão.
        # 'pady=5' adiciona um espaço vertical de 5 pixels acima e abaixo do botão.
        btn_filtrar.grid(row=0, column=6, padx=5, pady=5)

        # Cria um frame com borda rotulada para exibir os resultados do filtro.
        # O texto "Resultados" no rótulo informa ao usuário que este espaço exibe os dados filtrados.
        # 'padx=10' e 'pady=10' criam margens internas de 10 pixels ao redor do conteúdo no frame.
        frame_resultados = tk.LabelFrame(self,
                                         text="Resultados",
                                         padx=10,
                                         pady=10)

        # Posiciona o frame de resultados na interface.
        # 'side="top"' coloca o frame no topo da janela.
        # 'fill="both"' faz com que o frame preencha todo o espaço disponível em largura e altura.
        # 'expand=True' permite que o frame expanda proporcionalmente quando a janela é redimensionada.
        # 'padx=10' e 'pady=10' adicionam margens externas de 10 pixels ao redor do frame.
        frame_resultados.pack(side="top",
                              fill="both",
                              expand=True,
                              padx=10,
                              pady=10)

        # Define as colunas do Treeview para exibir os dados filtrados.
        # As colunas incluem: índice ("idx"), data, animal, médico, valor da
        #       consulta, total de produtos, total geral, e observações.
        # O tuple `colunas` lista as identificações das colunas, que
        #       serão usadas no Treeview.
        colunas = ("idx", "data", "animal", "medico", "valor_consulta", "total_produtos", "total_geral", "observacoes")

        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria um widget Treeview para exibir os resultados dentro do frame de resultados.
        # O parâmetro `columns=colunas` configura as colunas a serem exibidas no Treeview.
        # O parâmetro `show="headings"` faz com que apenas os cabeçalhos das
        #       colunas sejam exibidos, sem uma coluna inicial extra.
        self.tree_result = ttk.Treeview(frame_resultados,
                                        columns=colunas,
                                        show="headings")

        # Configura o cabeçalho da coluna "idx" para exibir o texto "ID".
        # Isto ajuda o usuário a identificar que esta coluna representa os
        #       identificadores dos registros.
        self.tree_result.heading("idx", text="ID")

        # Configura a largura e o alinhamento da coluna "idx".
        # 'width=80' define a largura em pixels da coluna.
        # 'anchor="center"' centraliza o conteúdo dentro da célula desta coluna.
        self.tree_result.column("idx",
                                width=80,
                                anchor="center")

        # Configura o cabeçalho da coluna "data" para exibir o texto "Data/Hora".
        # Isto identifica que a coluna exibe a data e hora dos registros.
        self.tree_result.heading("data", text="Data/Hora")

        # Define a largura e o alinhamento da coluna "data".
        # 'width=150' ajusta a largura para 150 pixels.
        # 'anchor="center"' centraliza o conteúdo da célula.
        self.tree_result.column("data", width=150, anchor="center")

        # Configura o cabeçalho da coluna "animal" para exibir o texto "Animal".
        # Esta coluna irá mostrar o nome do animal relacionado ao registro.
        self.tree_result.heading("animal", text="Animal")

        # Define a largura e o alinhamento da coluna "animal".
        # 'width=150' ajusta a largura da coluna.
        # 'anchor="center"' centraliza o texto.
        self.tree_result.column("animal", width=150, anchor="center")

        # Configura o cabeçalho da coluna "medico" para exibir o texto "Médico".
        # Indica que a coluna exibe o nome do médico associado ao registro.
        self.tree_result.heading("medico", text="Médico")

        # Define a largura e o alinhamento da coluna "medico".
        # 'width=150' ajusta a largura da coluna para comportar nomes de médicos.
        # 'anchor="center"' centraliza o texto.
        self.tree_result.column("medico", width=150, anchor="center")

        # Configura o cabeçalho da coluna "valor_consulta" para exibir o texto "Cons. (R$)".
        # Informa que a coluna exibe o valor cobrado pela consulta.
        self.tree_result.heading("valor_consulta", text="Cons. (R$)")

        # Define a largura e o alinhamento da coluna "valor_consulta".
        # 'width=120' ajusta a largura da coluna para valores monetários.
        # 'anchor="center"' centraliza os valores dentro da coluna.
        self.tree_result.column("valor_consulta", width=120, anchor="center")

        # Configura o cabeçalho da coluna "total_produtos" para exibir o texto "Prod. (R$)".
        # Indica que esta coluna mostra o total em reais dos produtos utilizados.
        self.tree_result.heading("total_produtos", text="Prod. (R$)")

        # Define a largura e o alinhamento da coluna "total_produtos".
        # 'width=120' ajusta a largura para valores monetários.
        # 'anchor="center"' centraliza o texto dentro da célula.
        self.tree_result.column("total_produtos", width=120, anchor="center")

        # Configura o cabeçalho da coluna "total_geral" para exibir o texto "Total (R$)".
        # Identifica que a coluna exibe o total geral (consulta + produtos) em reais.
        self.tree_result.heading("total_geral", text="Total (R$)")

        # Define a largura e o alinhamento da coluna "total_geral".
        # 'width=120' ajusta a largura para valores monetários.
        # 'anchor="center"' centraliza os valores.
        self.tree_result.column("total_geral", width=120, anchor="center")

        # Configura o cabeçalho da coluna "observacoes" para exibir o texto "Observações".
        # Indica que a coluna exibe quaisquer observações associadas ao registro.
        self.tree_result.heading("observacoes", text="Observações")

        # Define a largura e o alinhamento da coluna "observacoes".
        # 'width=200' ajusta a largura para comportar textos maiores.
        # 'anchor="w"' alinha o texto à esquerda dentro da célula.
        self.tree_result.column("observacoes", width=200, anchor="w")

        # Exibe o Treeview na interface gráfica.
        # 'fill="both"' faz o Treeview preencher toda a largura e altura disponíveis.
        # 'expand=True' permite que o Treeview expanda
        #       proporcionalmente ao redimensionar a janela.
        self.tree_result.pack(fill="both", expand=True)

        # Cria um rótulo para exibir o total geral dos valores listados no relatório.
        # O texto inicial é "Total Geral: R$ 0.00", indicando que o total começa zerado.
        # 'font=("Arial", 14, "bold")' aplica uma fonte maior e em negrito para dar destaque.
        self.label_soma = tk.Label(self,
                                   text="Total Geral: R$ 0.00",
                                   font=("Arial", 14, "bold"))

        # Posiciona o rótulo na interface com um espaçamento vertical (padding) de 5 pixels.
        # 'pady=5' adiciona espaço acima e abaixo do rótulo para melhorar a aparência.
        self.label_soma.pack(pady=5)

        # Adiciona um evento de "duplo clique" no Treeview de resultados.
        # O evento chama o método 'abrir_detalhes_atendimento',
        #       que exibe detalhes do registro clicado.
        self.tree_result.bind("<Double-1>", self.abrir_detalhes_atendimento)

        # Adiciona um evento para quando a aba de relatórios se torna visível.
        # O evento chama o método 'on_tab_visible' para recarregar os dados exibidos.
        self.bind("<Visibility>", self.on_tab_visible)

    # Evento disparado quando a aba de relatórios se torna visível.
    def on_tab_visible(self, event):

        # Obtém a lista de médicos do banco de dados ou de uma função auxiliar.
        # 'obter_lista_medicos()' retorna uma lista de tuplas, onde cada tupla contém
        # o ID e o nome do médico. Exemplo: [(id1, "Dr. João"), (id2, "Dr. Maria")].
        lista = obter_lista_medicos()

        # Extrai apenas os nomes dos médicos da lista para exibição no combobox.
        # Usando list comprehension, pega o segundo elemento de cada tupla (o nome).
        nomes = [m for (_id, m) in lista]

        # Configura o combobox 'self.combo_medico' para mostrar a lista de nomes,
        # adicionando a opção "(Todos)" como primeira entrada.
        self.combo_medico["values"] = ["(Todos)"] + nomes

        # Define a seleção inicial do combobox para a primeira opção "(Todos)".
        # Apenas aplica a seleção se houver nomes na lista.
        if len(nomes) > 0:
            self.combo_medico.current(0)



    # Inicia o processo de filtragem do relatório, limpando dados anteriores.
    def filtrar_relatorio(self):

        # Remove todas as linhas exibidas na treeview.
        for item in self.tree_result.get_children():
            self.tree_result.delete(item)

        # Limpa a lista interna que armazena os resultados encontrados.
        self.resultados_encontrados.clear()

        # Obtém a data inicial e final selecionadas nos campos de data.
        dt_ini = self.data_ini.get_date()
        dt_fim = self.data_fim.get_date()

        # Converte as datas em objetos datetime com horários ajustados:
        # Data inicial às 00:00:00 e data final às 23:59:59.
        dt_ini = datetime.datetime(dt_ini.year, dt_ini.month, dt_ini.day, 0, 0, 0)
        dt_fim = datetime.datetime(dt_fim.year, dt_fim.month, dt_fim.day, 23, 59, 59)

        # Obtém o nome do médico selecionado no combobox de filtro.
        nome_medico_filtro = self.combo_medico.get()

        # Determina se o filtro por médico será aplicado.
        # O filtro será ativo se um médico específico for selecionado (excluindo "(Todos)").
        filtrar_medico = (nome_medico_filtro and nome_medico_filtro != "(Todos)")

        # Inicializa a soma geral para calcular o total de valores exibidos no relatório.
        soma_geral = 0.0

        # Exemplo: Percorre todos os animais cadastrados no banco de dados.
        # O objetivo é verificar os históricos de cada animal e filtrar
        #       apenas os do tipo "Atendimento".
        for ani in db.animais.find():

            # Obtém o nome do animal. Se não houver, usa uma string vazia como fallback.
            nome_animal = ani.get("nome_animal", "")

            # Obtém a lista de históricos do animal. Se não houver, usa
            #       uma lista vazia como fallback.
            historicos = ani.get("historicos", [])

            # Itera sobre cada histórico associado ao animal.
            for i, h in enumerate(historicos):

                # Verifica se o tipo do histórico é "Atendimento". Caso contrário, ignora.
                if h.get("tipo") != "Atendimento":
                    continue

                # Obtém a data do atendimento. Se não houver, pula para o próximo histórico.
                data_atend = h.get("data")
                if not data_atend:
                    continue

                # Verifica se a data do atendimento está dentro do intervalo
                #       especificado (dt_ini e dt_fim).
                # Se estiver fora do intervalo, ignora o registro.
                if not (dt_ini <= data_atend <= dt_fim):
                    continue

                # Se o filtro por médico estiver ativo, verifica se o médico
                #       do atendimento corresponde
                #       ao médico selecionado. Caso contrário, ignora o registro.
                if filtrar_medico:
                    if h.get("realizado_por_nome") != nome_medico_filtro:
                        continue

                # Formata a data do atendimento para um formato legível (ex.: DD/MM/AAAA HH:MM).
                # Usa a função `formatar_data_hora` para isso.
                data_str = formatar_data_hora(data_atend)

                # Obtém o valor da consulta do histórico. Se não estiver
                #       definido, usa 0.0 como padrão.
                valor_consulta = h.get("valor_consulta", 0.0)

                # Obtém o total de produtos do histórico. Se não estiver
                #       definido, usa 0.0 como padrão.
                total_produtos = h.get("total_produtos", 0.0)

                # Obtém o valor total (consulta + produtos) do histórico.
                # Se não estiver definido, usa 0.0 como padrão.
                total_geral = h.get("total_geral", 0.0)

                # Obtém o nome do médico que realizou o atendimento.
                # Se não estiver definido, usa uma string vazia como padrão.
                medico_nome = h.get("realizado_por_nome", "")

                # Obtém as observações relacionadas ao atendimento.
                # Se não houver, usa uma string vazia como padrão.
                obs = h.get("observacoes", "")

                # Cria um dicionário que representa o registro completo do atendimento.
                # Inclui todas as informações relevantes para exibição e manipulação.
                registro_completo = {

                    # Um identificador interno único, composto pelo ID do
                    #       animal e o índice do histórico.
                    "idx_interno": f"{ani['_id']}_{i}",

                    # Nome do animal associado ao atendimento.
                    "animal": nome_animal,

                    # Data do atendimento formatada como string.
                    "data_str": data_str,

                    # Valor da consulta do atendimento.
                    "valor_consulta": valor_consulta,

                    # Total em produtos usados no atendimento.
                    "total_produtos": total_produtos,

                    # Valor geral do atendimento (consulta + produtos).
                    "total_geral": total_geral,

                    # Nome do médico que realizou o atendimento.
                    "medico": medico_nome,

                    # Observações adicionais sobre o atendimento.
                    "observacoes": obs,

                    # Lista de produtos usados no atendimento.
                    "produtos_usados": h.get("produtos_usados", []),

                    # Data do atendimento em formato bruto (objeto `datetime`).
                    "hora_crua": data_atend,
                }

                # Adiciona o registro completo à lista de resultados encontrados.
                self.resultados_encontrados.append(registro_completo)

        # Itera por cada registro completo encontrado em `self.resultados_encontrados`.
        for reg in self.resultados_encontrados:

            # Recupera o identificador interno único do registro.
            idx_interno = reg["idx_interno"]

            # Recupera a data formatada como string (ex.: DD/MM/AAAA HH:MM).
            data_str = reg["data_str"]

            # Recupera o nome do animal associado ao atendimento.
            animal = reg["animal"]

            # Recupera o nome do médico que realizou o atendimento.
            medico_nome = reg["medico"]

            # Recupera o valor da consulta do atendimento.
            valor_cons = reg["valor_consulta"]

            # Recupera o total de produtos usados no atendimento.
            tot_prod = reg["total_produtos"]

            # Recupera o valor total do atendimento (consulta + produtos).
            tot_geral = reg["total_geral"]

            # Recupera as observações relacionadas ao atendimento.
            obs = reg["observacoes"]

            # Insere os valores formatados na Treeview, organizados em colunas.
            self.tree_result.insert("",
                                    tk.END,
                                    values=(idx_interno,  # Identificador único interno, usado para referência.

                                            # Data e hora formatadas como string.
                                            data_str,

                                            # Nome do animal associado ao atendimento.
                                            animal,

                                            # Nome do médico que realizou o atendimento.
                                            medico_nome,

                                            # Valor da consulta formatado com duas casas decimais.
                                            f"{valor_cons:.2f}",

                                            # Total dos produtos utilizados no atendimento.
                                            f"{tot_prod:.2f}",

                                            # Valor total do atendimento (consulta + produtos).
                                            f"{tot_geral:.2f}",

                                            # Observações adicionais registradas no atendimento.
                                            obs))

            # Soma o valor total do atendimento ao acumulador geral.
            soma_geral += tot_geral

        # Atualiza o rótulo que exibe o total geral de valores,
        #       formatando com duas casas decimais.
        self.label_soma.config(text=f"Total Geral: R$ {soma_geral:.2f}")


    # Define o método que será acionado ao dar duplo clique em
    #       uma linha da Treeview.
    def abrir_detalhes_atendimento(self, event):

        """
        Este método é responsável por abrir uma janela detalhada de
                atendimento ao clicar duas vezes
                em uma linha da tabela de resultados (Treeview). Ele localiza o
                registro correspondente na lista
                de resultados encontrados e cria a janela com os detalhes.
        """

        # Obtém a linha atualmente selecionada na Treeview.
        selection = self.tree_result.selection()

        # Verifica se alguma linha foi selecionada. Se não, interrompe o método.
        if not selection:
            return

        # Extrai os valores da linha selecionada na Treeview.
        item = self.tree_result.item(selection[0])
        vals = item["values"]

        # Captura o identificador interno da linha (primeira coluna da Treeview).
        idx_interno = vals[0]  # colunas = ("idx", "data", ...)

        # Inicializa a variável para armazenar o registro correspondente.
        registro_encontrado = None

        # Itera pela lista de resultados encontrados para localizar o
        #       registro com o índice interno correspondente.
        for reg in self.resultados_encontrados:

            # Compara o "idx_interno" da linha selecionada com o
            #       índice de cada registro na lista.
            if reg["idx_interno"] == idx_interno:

                # Se encontrar, armazena o registro e interrompe o loop.
                registro_encontrado = reg
                break

        # Verifica se o registro foi encontrado.
        if registro_encontrado is None:

            # Exibe uma mensagem de erro se o registro correspondente
            #       não for localizado.
            messagebox.showerror("Erro",
                                 "Não encontrei detalhes deste atendimento na lista.")
            return

        # Abre a janela de detalhes do relatório, passando o
        #       registro encontrado como parâmetro.
        JanelaDetalheRelatorio(self, registro_encontrado)


class JanelaDetalheRelatorio(tk.Toplevel):

    """
    Classe JanelaDetalheRelatorio:
    - Exibe detalhes de um atendimento específico, incluindo dados gerais e produtos usados.
    - Herda de `tk.Toplevel`, permitindo que seja uma janela secundária independente.
    """

    def __init__(self, parent, registro, *args, **kwargs):

        """
        Inicializador da JanelaDetalheRelatorio:
        - Configura o layout principal, centraliza a janela e exibe os
                dados do registro fornecido.

        Parâmetros:
        - parent: Referência à janela ou frame pai.
        - registro: Dicionário contendo os dados do atendimento a ser exibido.
        - *args, **kwargs: Argumentos opcionais para configurar a janela.

        Ações realizadas:
        - Define o título da janela como "Detalhes do Atendimento (Relatório)".
        - Ajusta o tamanho da janela para 700x500 pixels.
        - Centraliza a janela na tela com a função `centralizar_janela`.
        - Configura o fundo da janela como branco para um layout limpo.
        """

        # Inicializa a classe pai (tk.Toplevel) para garantir que
        #       todos os comportamentos padrões sejam aplicados.
        super().__init__(parent, *args, **kwargs)

        # Define o título da janela para "Detalhes do Atendimento (Relatório)".
        # Este título será exibido na barra superior da janela.
        self.title("Detalhes do Atendimento (Relatório)")

        # Define o tamanho da janela como 700 pixels de largura por 500 pixels de altura.
        # Isso garante uma área suficiente para exibir os detalhes do
        #       atendimento e os produtos usados.
        self.geometry("700x700")

        # Centraliza a janela na tela, utilizando a função personalizada `centralizar_janela`.
        # Isso melhora a experiência do usuário, posicionando a janela no centro da tela.
        centralizar_janela(self, 700, 700)

        # Configura o fundo da janela para a cor branca.
        # Essa escolha de cor proporciona um layout limpo e profissional.
        self.configure(bg="white")

        # Armazena o registro fornecido como um atributo da
        #       instância para uso posterior.
        self.registro = registro

        # Extrai a string de data formatada do registro. Isso será exibido no cabeçalho.
        data_str = registro["data_str"]

        # Extrai o nome do animal associado ao atendimento do registro.
        # Este campo identifica qual animal recebeu o atendimento.
        animal = registro["animal"]

        # Obtém o nome do médico que realizou o atendimento.
        # Esse dado será usado para informar o responsável pelo procedimento.
        medico = registro["medico"]

        # Obtém o valor da consulta do registro. Esse valor é usado no
        #       resumo financeiro do atendimento.
        valor_consulta = registro["valor_consulta"]

        # Obtém o total de produtos usados no atendimento, calculado com base no registro.
        total_produtos = registro["total_produtos"]

        # Obtém o valor total do atendimento, incluindo a consulta e os produtos usados.
        total_geral = registro["total_geral"]

        # Extrai as observações feitas no registro do atendimento.
        # Estas observações podem conter informações adicionais relevantes.
        obs = registro["observacoes"]

        # Extrai a lista de produtos usados durante o atendimento.
        # Essa lista será exibida em uma tabela de detalhes.
        produtos_usados = registro["produtos_usados"]

        # Cria o frame principal onde todos os elementos da janela serão posicionados.
        # Define o fundo como branco para harmonizar com o restante do layout.
        frame_main = tk.Frame(self, bg="white")

        # Posiciona o frame principal para preencher toda a janela.
        # `fill="both"` faz com que o frame expanda tanto horizontal quanto verticalmente.
        # `expand=True` permite que ele use todo o espaço disponível ao redimensionar.
        # `padx=10` e `pady=10` adicionam margens internas ao redor do frame.
        frame_main.pack(fill="both",
                        expand=True,
                        padx=10,
                        pady=10)

        # Adiciona um título descritivo ao frame principal para
        #       indicar a funcionalidade da janela.
        # Define uma fonte grande e negrito para destacar o título.
        # `bg="white"` faz com que o fundo do rótulo combine com o fundo do frame.
        tk.Label(frame_main,
                text="Detalhes do Atendimento",
                font=("Arial", 18, "bold"),
                bg="white").pack(pady=10)  # Posiciona o título com espaçamento vertical (10px) acima e abaixo.

        # Cria um LabelFrame que agrupa as informações principais do atendimento.
        # O texto "Informações Principais" identifica claramente a seção.
        # `bg="white"` define o fundo como branco, mantendo consistência
        #       com o layout geral da janela.
        # `padx=10` e `pady=10` adicionam espaçamento interno horizontal e
        #       vertical, melhorando a legibilidade.
        # `font=("Arial", 12, "bold")` utiliza uma fonte clara, com tamanho 12 e
        #       estilo negrito para o título.
        info_frame = tk.LabelFrame(frame_main,
                                    text="Informações Principais",
                                    bg="white",
                                    padx=10,
                                    pady=10,
                                    font=("Arial", 12, "bold"))

        # Posiciona o LabelFrame no frame principal, preenchendo
        #       horizontalmente (`fill="x"`).
        # `padx=5` adiciona 5 pixels de margem horizontal externa para
        #       espaçamento entre elementos laterais.
        # `pady=5` adiciona 5 pixels de margem vertical externa, separando
        #       este LabelFrame dos demais elementos.
        info_frame.pack(fill="x", padx=5, pady=5)

        # Cria um rótulo para exibir a data e hora do atendimento.
        # `f"Data/Hora: {data_str}"` insere a data e hora formatadas do
        #       atendimento no texto do rótulo.
        # `font=("Arial", 12)` utiliza uma fonte "Arial" de tamanho 12 para leitura confortável.
        # `bg="white"` mantém o fundo do rótulo branco, combinando com o
        #       LabelFrame onde está inserido.
        tk.Label(info_frame,
                text=f"Data/Hora: {data_str}",
                font=("Arial", 12),
                bg="white").pack(anchor="w",  # Alinha o texto à esquerda dentro do LabelFrame.
                                pady=2)  # Adiciona 2 pixels de espaçamento vertical externo,
                                            # separando o rótulo dos demais elementos.

        # Cria um rótulo para exibir o nome do animal relacionado ao atendimento.
        # `f"Animal: {animal}"` insere o nome do animal no texto do
        #       rótulo, extraído do registro fornecido.
        # `font=("Arial", 12)` utiliza uma fonte "Arial" de tamanho 12,
        #       proporcionando boa legibilidade.
        # `bg="white"` ajusta o fundo do rótulo para branco, mantendo o
        #       layout visual uniforme.
        # `.pack(anchor="w", pady=2)` posiciona o rótulo alinhado à
        #       esquerda (`anchor="w"`) com 2 pixels de espaçamento vertical externo.
        tk.Label(info_frame,
                text=f"Animal: {animal}",
                font=("Arial", 12),
                bg="white").pack(anchor="w",  # Alinha o texto à esquerda dentro do LabelFrame.
                                pady=2)  # Adiciona 2 pixels de espaçamento vertical externo.


        # Cria um rótulo para exibir o nome do médico responsável pelo atendimento.
        # `f"Médico: {medico}"` insere o nome do médico no texto do
        #       rótulo, extraído do registro fornecido.
        # `font=("Arial", 12)` e `bg="white"` seguem o mesmo padrão para
        #       consistência visual e de legibilidade.
        # `.pack(anchor="w", pady=2)` posiciona o rótulo alinhado à
        #       esquerda com espaçamento vertical.
        tk.Label(info_frame,
                text=f"Médico: {medico}",
                font=("Arial", 12),
                bg="white").pack(anchor="w",  # Alinha o texto à esquerda dentro do LabelFrame.
                                pady=2)  # Adiciona 2 pixels de espaçamento vertical externo.


        # Cria um rótulo para exibir o valor da consulta.
        # `f"Valor da Consulta: R$ {valor_consulta:.2f}"` formata o
        #       valor da consulta com duas casas decimais.
        # `font=("Arial", 12)` define a fonte "Arial" de
        #       tamanho 12 para manter a legibilidade.
        # `bg="white"` mantém o fundo branco, alinhando com o layout geral.
        # `.pack(anchor="w", pady=2)` posiciona o rótulo alinhado à esquerda
        #       com 2 pixels de espaçamento vertical externo.
        tk.Label(info_frame,
                text=f"Valor da Consulta: R$ {valor_consulta:.2f}",
                font=("Arial", 12),
                bg="white").pack(anchor="w",  # Alinha o texto à esquerda dentro do LabelFrame.
                                pady=2)  # Adiciona 2 pixels de espaçamento vertical externo.


        # Cria um rótulo para exibir o total referente aos produtos
        #       utilizados no atendimento.
        # O texto exibe o total formatado em reais com duas casas decimais.
        # `font=("Arial", 12)` especifica a fonte "Arial" com tamanho 12 para legibilidade.
        # `bg="white"` define o fundo branco. O nome da cor "white" é Branco.
        # `.pack(anchor="w", pady=2)` alinha o texto à esquerda e
        #       adiciona 2 pixels de margem vertical externa.
        tk.Label(info_frame,
                text=f"Total Produtos: R$ {total_produtos:.2f}",
                font=("Arial", 12),
                bg="white"  # Fundo branco (White).
                ).pack(anchor="w",  # Alinha o texto à esquerda dentro do LabelFrame.
                        pady=2)  # Adiciona 2 pixels de espaçamento vertical externo.

        # Cria um rótulo para exibir o valor total geral do atendimento.
        # O texto exibe o total geral formatado em reais com duas casas decimais.
        # `font=("Arial", 12, "bold")` destaca o texto com tamanho 12 em negrito.
        # `bg="white"` mantém o fundo branco. O nome da cor "white" é Branco.
        # `.pack(anchor="w", pady=2)` posiciona o texto à esquerda com
        #       espaçamento vertical para organização.
        tk.Label(info_frame,
                text=f"Total Geral: R$ {total_geral:.2f}",
                font=("Arial", 12, "bold"),
                bg="white"  # Fundo branco (White).
                ).pack(
                    anchor="w",  # Alinha o texto à esquerda dentro do LabelFrame.
                    pady=2  # Adiciona 2 pixels de espaçamento vertical externo.
                )

        # Verifica se há observações para exibir.
        if obs:

            # Cria um LabelFrame para agrupar o campo de observações.
            # `text="Observações"` define o título do frame.
            # `bg="white"` define o fundo como branco (White) para
            #       harmonizar com o layout geral.
            # `padx=10` e `pady=10` adicionam 10 pixels de espaçamento interno
            #       horizontal e vertical, respectivamente.
            # `font=("Arial", 12, "bold")` aplica a fonte Arial com tamanho 12 e
            #       em negrito para o título do frame.
            obs_frame = tk.LabelFrame(
                frame_main,
                text="Observações",
                bg="white",  # Fundo branco (White).
                padx=10,  # Espaçamento interno horizontal de 10 pixels.
                pady=10,  # Espaçamento interno vertical de 10 pixels.
                font=("Arial", 12, "bold"))  # Fonte Arial, tamanho 12, em negrito.

            # Posiciona o LabelFrame no layout principal.
            # `fill="x"` faz com que o frame preencha toda a largura disponível.
            # `padx=5` e `pady=5` adicionam espaçamento externo
            #       horizontal e vertical, respectivamente.
            obs_frame.pack(fill="x", padx=5, pady=5)

            # Cria um widget Text para exibir o conteúdo das observações.
            # `width=80` define a largura do campo em caracteres.
            # `height=4` define a altura do campo em linhas de texto.
            # `font=("Arial", 12)` especifica a fonte Arial com tamanho 12.
            txt_obs = tk.Text(  obs_frame,
                                width=80,  # Largura de 80 caracteres.
                                height=4,  # Altura de 4 linhas.
                                font=("Arial", 12))  # Fonte Arial, tamanho 12.

            # Posiciona o campo de texto dentro do frame.
            # `fill="x"` faz com que o campo preencha a largura disponível no frame.
            # `expand=True` permite que o campo cresça com o frame.
            txt_obs.pack(fill="x", expand=True)

            # Insere o texto das observações no campo de texto.
            # `tk.END` insere o texto no final do campo.
            txt_obs.insert(tk.END, obs)

            # Desabilita a edição do campo de texto para que o conteúdo
            #       seja apenas visualizado.
            txt_obs.config(state="disabled")

        # Cria um LabelFrame para agrupar a tabela que exibirá os
        #       produtos utilizados no atendimento.
        # `text="Produtos Utilizados"` define o título do frame.
        # `bg="white"` define o fundo como branco (White), alinhado ao estilo geral da janela.
        # `padx=10` e `pady=10` adicionam 10 pixels de espaçamento interno
        #       horizontal e vertical, respectivamente.
        # `font=("Arial", 12, "bold")` aplica a fonte Arial com tamanho 12 e
        #       em negrito para o título do frame.
        produtos_frame = tk.LabelFrame( frame_main,
                                        text="Produtos Utilizados",
                                        bg="white",  # Fundo branco (White) para consistência visual.
                                        padx=10,  # Espaçamento interno horizontal de 10 pixels.
                                        pady=10,  # Espaçamento interno vertical de 10 pixels.
                                        font=("Arial", 12, "bold"))  # Fonte Arial, tamanho 12, em negrito.


        # Posiciona o LabelFrame no layout principal.
        # `fill="both"` faz com que o frame preencha tanto a largura
        #       quanto a altura disponíveis.
        # `expand=True` permite que o frame cresça proporcionalmente ao
        #       redimensionamento da janela.
        # `padx=5` e `pady=5` adicionam espaçamento externo
        #       horizontal e vertical ao frame.
        produtos_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Define as colunas da Treeview que exibirá os produtos utilizados no atendimento.
        # `("nome", "preco_unit", "qtd", "subtotal")` são os identificadores das colunas.
        colunas = ("nome", "preco_unit", "qtd", "subtotal")

        # Criando um estilo para personalizar o Treeview
        style = ttk.Style()
        style.theme_use("default")

        # Configurando a cor de fundo e o texto do item selecionado
        style.map("Treeview",
                  background=[("selected", "darkblue")],  # Cor de fundo do item selecionado
                  foreground=[("selected", "white")])  # Cor da fonte do item selecionado


        # Cria a Treeview dentro do `produtos_frame` para exibir os dados tabulares.
        # `columns=colunas` define as colunas da tabela.
        # `show="headings"` exibe apenas os cabeçalhos das
        #       colunas (sem a coluna padrão à esquerda).
        self.tree_prods = ttk.Treeview(
            produtos_frame,
            columns=colunas,
            show="headings"
        )

        # Configura o cabeçalho e as propriedades da coluna "nome".
        # `heading("nome", text="Produto")` define o texto do cabeçalho como "Produto".
        self.tree_prods.heading("nome", text="Produto")

        # `column("nome", width=200, anchor="w")` define a largura
        #       como 200 pixels e alinha o conteúdo à esquerda.
        self.tree_prods.column("nome", width=200, anchor="w")

        # Configura o cabeçalho e as propriedades da coluna "preco_unit".
        # `heading("preco_unit", text="Preço Unitário (R$)")` define o texto do
        #       cabeçalho como "Preço Unitário (R$)".
        self.tree_prods.heading("preco_unit",
                                text="Preço Unitário (R$)")

        # `column("preco_unit", width=120, anchor="center")` define a largura
        #       como 120 pixels e centraliza o conteúdo.
        self.tree_prods.column("preco_unit",
                               width=120,
                               anchor="center")

        # Configura o cabeçalho e as propriedades da coluna "qtd".
        # `heading("qtd", text="Qtd")` define o texto do cabeçalho como "Qtd".
        self.tree_prods.heading("qtd", text="Qtd")

        # `column("qtd", width=80, anchor="center")` define a largura
        #       como 80 pixels e centraliza o conteúdo.
        self.tree_prods.column("qtd",
                               width=80,
                               anchor="center")

        # Configura o cabeçalho e as propriedades da coluna "subtotal".
        # `heading("subtotal", text="Subtotal (R$)")` define o texto do
        #       cabeçalho como "Subtotal (R$)".
        self.tree_prods.heading("subtotal",
                                text="Subtotal (R$)")

        # `column("subtotal", width=120, anchor="center")` define a largura
        #       como 120 pixels e centraliza o conteúdo.
        self.tree_prods.column("subtotal",
                               width=120,
                               anchor="center")

        # Posiciona a Treeview dentro do `produtos_frame`.
        # `side="left"` alinha a Treeview ao lado esquerdo do frame.
        # `fill="both"` faz com que a Treeview preencha tanto a largura
        #       quanto a altura disponíveis.
        # `expand=True` permite que a Treeview cresça proporcionalmente ao
        #       redimensionamento do frame.
        self.tree_prods.pack(side="left",
                             fill="both",
                             expand=True)

        # Cria uma barra de rolagem vertical para a Treeview de produtos.
        # `produtos_frame` é o frame onde a barra de rolagem será posicionada.
        # `orient="vertical"` define que a barra de rolagem será vertical.
        # `command=self.tree_prods.yview` conecta a barra de rolagem com a
        #       visualização vertical da Treeview.
        sb_vertical = ttk.Scrollbar( produtos_frame,
                                    orient="vertical",
                                    command=self.tree_prods.yview)

        # Posiciona a barra de rolagem vertical dentro do `produtos_frame`.
        # `side="right"` alinha a barra de rolagem ao lado direito do frame.
        # `fill="y"` faz com que a barra de rolagem preencha toda a altura disponível.
        sb_vertical.pack(side="right", fill="y")

        # Configura a Treeview para sincronizar sua visualização
        #       vertical com a barra de rolagem.
        # `yscrollcommand=sb_vertical.set` conecta o comando de rolagem da
        #       Treeview à barra de rolagem vertical.
        self.tree_prods.configure(yscrollcommand=sb_vertical.set)

        # Itera sobre a lista de produtos utilizados para preencher a Treeview.
        # Cada produto é representado como um dicionário contendo informações detalhadas.
        for p in produtos_usados:

            # Obtém o nome do produto. Se não existir, retorna uma string vazia.
            nome_prod = p.get("nome", "")

            # Obtém o preço unitário do produto. Se não existir, assume o valor 0.0.
            prec = p.get("preco_unitario", 0.0)

            # Obtém a quantidade utilizada do produto. Se não existir, assume o valor 0.0.
            qtd = p.get("quantidade", 0.0)

            # Obtém o subtotal do produto (preço unitário x quantidade).
            # Se não existir, assume o valor 0.0.
            st = p.get("subtotal", 0.0)

            # Insere os dados do produto na Treeview.
            # Os valores são formatados para exibir números com duas casas decimais.
            # `""` representa o ID do item pai na Treeview, já que não há hierarquia.
            # `tk.END` indica que o item será inserido ao final da lista existente.
            self.tree_prods.insert(
                "",
                tk.END,
                values=(
                    nome_prod,  # Nome do produto.
                    f"{prec:.2f}",  # Preço unitário formatado com duas casas decimais.
                    f"{qtd:.2f}",  # Quantidade formatada com duas casas decimais.
                    f"{st:.2f}"  # Subtotal formatado com duas casas decimais.
                )
            )



# ---------------------------------------------------------------------
# Janela de Login em Tela Cheia
# ---------------------------------------------------------------------

# Define uma nova classe chamada JanelaLogin, que herda de tk.Tk.
# Isso significa que JanelaLogin é uma extensão da classe tk.Tk,
#       que é a janela principal do Tkinter.
# Com isso, podemos adicionar funcionalidades específicas à
#       nossa janela de login.
class JanelaLogin(tk.Tk):

    # O método __init__ é um "inicializador" que configura a classe assim que ela é criada.
    # Ele é usado para inicializar os atributos da classe e configurar o comportamento inicial.
    # Neste caso, estamos configurando a JanelaLogin.
    def __init__(self):

        # O método __init__ é o inicializador da classe. Ele é chamado automaticamente
        # toda vez que criamos uma nova instância da classe JanelaLogin. Ele serve
        # para configurar a janela logo após sua criação.
        super().__init__()  # Chama o inicializador da classe pai (tk.Tk),
                                    # que é a janela principal do Tkinter.

        # Define o título da janela, que aparece na barra de título do sistema operacional.
        # Isso ajuda o usuário a identificar facilmente a finalidade da janela.
        self.title("Login - Clínica Veterinária")

        # Configura a janela para abrir em tela cheia.
        # 'zoomed' faz com que a janela ocupe todo o espaço disponível na tela do usuário.
        # Isso é útil para aplicações onde queremos que os usuários
        #       tenham uma visão ampla do conteúdo.
        self.state('zoomed')

        try:

            # Inicia um bloco try para tentar executar o código que pode gerar erros.
            # Neste caso, estamos tentando carregar e configurar uma imagem de fundo.

            # Abre a imagem de fundo chamada "fundo.jpg".
            imagem_fundo = Image.open("fundo.jpg")
            # Caso o nome ou o caminho da imagem esteja incorreto, o
            #       programa levantará uma exceção.

            # Redimensiona a imagem para cobrir toda a tela.
            # 'self.winfo_screenwidth()' retorna a largura da tela.
            # 'self.winfo_screenheight()' retorna a altura da tela.
            # 'Image.Resampling.LANCZOS' é usado para redimensionar a
            #       imagem com alta qualidade.
            imagem_fundo = imagem_fundo.resize((self.winfo_screenwidth(),
                                                self.winfo_screenheight()),
                                                Image.Resampling.LANCZOS)

            # Converte a imagem em um formato que o Tkinter pode usar.
            self.bg_image = ImageTk.PhotoImage(imagem_fundo)

            # Cria um rótulo para exibir a imagem de fundo.
            # 'self' é a janela principal, então o rótulo será exibido na janela.
            rotulo_imagem = tk.Label(self)

            # Armazena a imagem no atributo 'image' do rótulo para evitar
            #       que o garbage collector a remova.
            rotulo_imagem.image = self.bg_image

            # Configura o rótulo para usar a imagem como fundo.
            rotulo_imagem.configure(image=self.bg_image)

            # Posiciona o rótulo para ocupar toda a janela.
            # 'x=0, y=0' define a posição inicial (canto superior esquerdo).
            # 'relwidth=1, relheight=1' define que o rótulo ocupará 100% da
            #       largura e altura da janela.
            rotulo_imagem.place(x=0, y=0, relwidth=1, relheight=1)

        except Exception as e:

            # Caso ocorra um erro, este bloco será executado.
            # Ele captura a exceção gerada (armazenada em 'e') e exibe
            #       uma mensagem de erro no console.
            print(f"Erro ao carregar a imagem de fundo: {e}")


        # Cria um container centralizado dentro da janela principal.
        # 'self' indica que o container pertence à janela principal.
        # 'bg="#ECECEC"' define a cor de fundo do container como um tom de cinza claro.
        # 'bd=2' define a largura da borda do container.
        # 'relief="groove"' aplica um estilo de borda em forma de sulco (groove).
        container = tk.Frame(self,
                             bg="#ECECEC",
                             bd=2,
                             relief="groove")

        # Posiciona o container no centro da janela principal.
        # 'relx=0.5, rely=0.5' posiciona o centro do container no centro
        #       da janela (50% da largura e altura).
        # 'anchor="center"' ancora o container pelo centro, para garantir
        #       que ele fique realmente centralizado.
        container.place(relx=0.5,
                        rely=0.5,
                        anchor="center")

        # Cria um rótulo de texto para o título dentro do container.
        # 'container' é o widget pai, ou seja, o rótulo será exibido dentro do container.
        # 'text' define o texto que será exibido no rótulo.
        # 'font=("Arial", 18, "bold")' define a fonte como Arial, tamanho 18, em negrito.
        # 'bg="#ECECEC"' ajusta a cor de fundo do rótulo para coincidir com a cor do container.
        label_titulo = tk.Label(container,
                                text="Bem-vindo(a) à Clínica Veterinária",
                                font=("Arial", 18, "bold"),
                                bg="#ECECEC")

        # Posiciona o rótulo de título dentro do container usando o layout de grade.
        # 'row=0' coloca o rótulo na primeira linha da grade.
        # 'column=0, columnspan=2' faz com que o rótulo ocupe duas colunas,
        #       centralizando-o horizontalmente.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical acima e abaixo do rótulo.
        label_titulo.grid(row=0, column=0, columnspan=2, pady=10)

        # Cria um rótulo para o campo "Usuário".
        # 'container' define que o rótulo será exibido dentro do container centralizado.
        # 'text="Usuário (Nome do Médico):"' especifica o texto exibido no rótulo.
        # 'bg="#ECECEC"' define a cor de fundo do rótulo como um cinza claro (#ECECEC),
        #       que é a mesma cor do fundo do container para manter uma aparência uniforme.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        tk.Label(container,
                 text="Usuário (Nome do Médico):",
                 bg="#ECECEC",
                 font=("Arial", 12)).grid(

            # Posiciona o rótulo no layout de grade do container.
            # 'row=1' coloca o rótulo na segunda linha da grade (a primeira é
            #       ocupada pelo título).
            # 'column=0' posiciona o rótulo na primeira coluna.
            # 'sticky="e"' alinha o rótulo à direita (east) da célula.
            # 'padx=5, pady=5' adiciona 5 pixels de espaçamento horizontal e vertical.
            row=1, column=0, sticky="e", padx=5, pady=5)

        # Cria uma caixa de entrada de texto para o campo "Usuário".
        # 'container' define que a entrada será exibida dentro do container centralizado.
        # 'width=30' define a largura da caixa de texto em caracteres.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        self.entry_usuario = tk.Entry(container,
                                      width=30,
                                      font=("Arial", 12))

        # Posiciona a caixa de entrada no layout de grade do container.
        # 'row=1' coloca a entrada na segunda linha da grade.
        # 'column=1' posiciona a entrada na segunda coluna.
        # 'padx=5, pady=5' adiciona 5 pixels de espaçamento horizontal e vertical.
        self.entry_usuario.grid(row=1,
                                column=1,
                                padx=5,
                                pady=5)
        self.entry_usuario.insert(0, "clevison")

        # Cria um rótulo para o campo "Senha".
        # 'container' especifica que o rótulo será exibido dentro do container centralizado.
        # 'text="Senha:"' define o texto exibido no rótulo.
        # 'bg="#ECECEC"' define a cor de fundo do rótulo como cinza claro (#ECECEC),
        # harmonizando com o fundo do container para uma estética uniforme.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        tk.Label(container,
                 text="Senha:",
                 bg="#ECECEC",
                 font=("Arial", 12)).grid(

            # Posiciona o rótulo no layout de grade do container.
            # 'row=2' coloca o rótulo na terceira linha da grade (a primeira
            #       linha contém o título).
            # 'column=0' posiciona o rótulo na primeira coluna.
            # 'sticky="e"' alinha o rótulo à direita (east) da célula para uma aparência organizada.
            # 'padx=5, pady=5' adiciona 5 pixels de espaçamento horizontal e vertical.
            row=2, column=0, sticky="e", padx=5, pady=5)

        # Cria uma caixa de entrada de texto para o campo "Senha".
        # 'container' especifica que a entrada será exibida dentro do container centralizado.
        # 'width=30' define a largura da caixa de entrada como 30 caracteres.
        # 'show="*"' mascara os caracteres digitados, exibindo asteriscos (*)
        #       para manter a confidencialidade da senha.
        # 'font=("Arial", 12)' define a fonte como Arial, tamanho 12.
        self.entry_senha = tk.Entry(container,
                                    width=30,
                                    show="*",
                                    font=("Arial", 12))

        # Posiciona a caixa de entrada no layout de grade do container.
        # 'row=2' coloca a entrada na terceira linha da grade.
        # 'column=1' posiciona a entrada na segunda coluna.
        # 'padx=5, pady=5' adiciona 5 pixels de espaçamento horizontal e vertical.
        self.entry_senha.grid(row=2,
                              column=1,
                              padx=5,
                              pady=5)
        self.entry_senha.insert(0, "555")

        # Cria um botão para a ação "Entrar".
        # 'container' define que o botão será exibido dentro do
        #       container centralizado.
        # 'text="Entrar"' especifica o texto exibido no botão.
        # 'font=("Arial", 12, "bold")' define a fonte como Arial,
        #       tamanho 12, com estilo negrito (bold).
        # 'bg="#4CAF50"' define a cor de fundo do botão como verde (#4CAF50),
        # geralmente usada para representar ações positivas.
        # 'fg="white"' define a cor do texto como branco, garantindo
        #       contraste com o fundo verde.
        # 'width=10' define a largura do botão como equivalente a 10 caracteres.
        # 'command=self.fazer_login' vincula o botão à função self.fazer_login,
        # que será executada quando o botão for clicado.
        btn_entrar = tk.Button(container,
                               text="Entrar",
                               font=("Arial", 12, "bold"),
                               bg="#4CAF50",
                               fg="white",
                               width=10,
                               command=self.fazer_login)

        # Posiciona o botão "Entrar" no layout de grade do container.
        # 'row=3' coloca o botão na quarta linha da grade.
        # 'column=0' posiciona o botão na primeira coluna.
        # 'padx=5, pady=10' adiciona 5 pixels de espaçamento
        #       horizontal e 10 pixels de espaçamento vertical.
        btn_entrar.grid(row=3, column=0, padx=5, pady=10)

        # Cria um botão para a ação "Cadastrar Novo Usuário".
        # 'container' especifica que o botão será exibido dentro do
        #       container centralizado.
        # 'text="Cadastrar Novo Usuário"' define o texto exibido no botão.
        # 'font=("Arial", 12)' define a fonte como Arial com tamanho 12.
        # 'bg="#2196F3"' define a cor de fundo do botão como azul (#2196F3),
        #       uma cor frequentemente usada para representar ações neutras ou informativas.
        # 'fg="white"' define a cor do texto como branco, garantindo um
        #       bom contraste com o fundo azul.
        # 'command=self.abrir_cadastro' vincula o botão à função self.abrir_cadastro,
        #       que será chamada quando o botão for clicado.
        btn_cadastrar = tk.Button(container,
                                  text="Cadastrar Novo Usuário",
                                  font=("Arial", 12),
                                  bg="#2196F3",
                                  fg="white",
                                  command=self.abrir_cadastro)

        # Posiciona o botão "Cadastrar Novo Usuário" no layout de
        #       grade do container.
        # 'row=3' coloca o botão na quarta linha da grade.
        # 'column=1' posiciona o botão na segunda coluna.
        # 'padx=5' adiciona 5 pixels de espaçamento horizontal entre o
        #       botão e os elementos adjacentes.
        # 'pady=10' adiciona 10 pixels de espaçamento vertical acima e abaixo do botão.
        btn_cadastrar.grid(row=3, column=1, padx=5, pady=10)


    # Define a função para realizar o login do usuário.
    # Essa função verifica as credenciais do usuário e realiza
    #       ações com base no resultado.
    def fazer_login(self):

        # Obtém o texto inserido no campo de entrada de usuário,
        # removendo espaços em branco do início e do fim com o método .strip().
        usuario = self.entry_usuario.get().strip()

        # Obtém o texto inserido no campo de entrada de senha,
        # também removendo espaços em branco do início e do fim.
        senha = self.entry_senha.get().strip()

        # Verifica se os campos de usuário ou senha estão vazios.
        # Se estiverem, exibe uma mensagem de aviso e interrompe a execução da função.
        if not usuario or not senha:
            messagebox.showwarning("Aviso", "Informe usuário e senha.")
            return

        # Busca no banco de dados MongoDB por um documento na coleção "medicos"
        # que corresponda ao nome de usuário e senha fornecidos.
        medico_doc = db.medicos.find_one({"nome": usuario, "senha": senha})

        # Se um médico correspondente for encontrado, o login é bem-sucedido.
        if medico_doc:

            # Exibe uma mensagem de sucesso informando que o login foi realizado.
            messagebox.showinfo("Sucesso", "Login bem-sucedido!")

            # Fecha a janela de login, destruindo sua instância.
            self.destroy()

            # Abre a janela principal, passando as informações do médico logado.
            # A função mainloop() é chamada para iniciar o loop principal da nova janela.
            JanelaPrincipal(medico_doc).mainloop()

        # Se nenhum médico correspondente for encontrado,
        #       exibe uma mensagem de erro.
        else:
            messagebox.showerror("Erro", "Usuário ou senha inválidos.")


    # Define a função para abrir a janela de cadastro de médico.
    # Essa função cria uma nova instância da classe JanelaCadastroMedico,
    # passando a janela de login (self) como o pai da nova janela.
    def abrir_cadastro(self):
        JanelaCadastroMedico(self)



# ---------------------------------------------------------------------
# Execução inicial
# ---------------------------------------------------------------------

# Cria uma instância da classe `JanelaLogin`, que inicializa a interface
# de login do sistema. Isso configura os elementos gráficos da janela
# principal e define o comportamento de interação do usuário.
app = JanelaLogin()  # Instância da janela de login.

# Inicia o loop principal da aplicação tkinter.
# O método `mainloop()` entra em um loop de eventos, aguardando
# interações do usuário (como cliques, digitação, etc.).
# É essencial para manter a janela aberta e responsiva.
app.mainloop()